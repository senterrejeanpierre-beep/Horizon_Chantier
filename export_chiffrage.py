from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# =========================
# Modèle poste (minimal)
# =========================
@dataclass
class Poste:
    code: str
    designation: str
    unite: str
    quantite: float


# =========================
# Styles simples (optionnel)
# =========================
def _styles():
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    title_fill = PatternFill("solid", fgColor="D9E1F2")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    return border, header_fill, title_fill, bold, center, left, right


# =========================
# Import bordereau (xlsx)
# =========================
def importer_bordereau_xlsx(
    path: Path,
    sheet_name: Optional[str] = None,
    col_code: str = "A",
    col_designation: str = "B",
    col_unite: str = "C",
    col_quantite: str = "D",
    start_row: int = 2,
) -> list[Poste]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    postes: list[Poste] = []
    r = start_row
    while True:
        code = ws[f"{col_code}{r}"].value
        des = ws[f"{col_designation}{r}"].value
        uni = ws[f"{col_unite}{r}"].value
        qty = ws[f"{col_quantite}{r}"].value

        # stop si ligne vide (adaptable)
        if code is None and des is None:
            break

        code_s = str(code).strip() if code is not None else ""
        des_s = str(des).strip() if des is not None else ""
        uni_s = str(uni).strip() if uni is not None else ""
        try:
            qty_f = float(qty) if qty is not None else 0.0
        except Exception:
            qty_f = 0.0

        # ignore lignes totalement vides
        if code_s or des_s:
            postes.append(Poste(code=code_s, designation=des_s, unite=uni_s, quantite=qty_f))

        r += 1

    return postes


# =========================
# Génération d'un bloc "poste"
# =========================
def ecrire_bloc_poste(ws: Worksheet, top: int, poste: Poste) -> int:
    """
    Écrit un bloc complet pour un poste à partir de la ligne `top`.
    Retourne la prochaine ligne disponible (après le bloc).
    """
    border, header_fill, title_fill, bold, center, left, right = _styles()

    # --- Paramètres du "pack"
    # Largeur: 12 colonnes A..L
    A, L = 1, 12

    # Hauteurs
    header_rows = 4          # en-tête poste
    mat_rows = 10            # 10 lignes matière
    mat_total_rows = 2       # titre + total
    mo_rows = 10             # 10 lignes MO
    mo_total_rows = 2        # titre + total
    recap_rows = 6           # totaux (à droite ou en dessous)
    gap = 2                  # espace entre sections

    # Lignes clés
    r0 = top
    r_header_end = r0 + header_rows - 1

    r_mat_title = r_header_end + 1
    r_mat_start = r_mat_title + 1
    r_mat_end = r_mat_start + mat_rows - 1
    r_mat_total = r_mat_end + 1

    r_mo_title = r_mat_total + gap
    r_mo_start = r_mo_title + 1
    r_mo_end = r_mo_start + mo_rows - 1
    r_mo_total = r_mo_end + 1

    r_recap_start = r_mo_total + gap
    r_recap_end = r_recap_start + recap_rows - 1

    next_top = r_recap_end + 3  # espace entre postes

    # --- Helper: style range
    def style_range(r1: int, r2: int, c1: int, c2: int):
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                cell = ws.cell(rr, cc)
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    # --- EN-TÊTE POSTE (A..L)
    ws.merge_cells(start_row=r0, start_column=A, end_row=r0, end_column=L)
    ws.cell(r0, A).value = f"POSTE  |  {poste.code}  |  {poste.designation}"
    ws.cell(r0, A).font = Font(bold=True, size=12)
    ws.cell(r0, A).fill = title_fill
    ws.cell(r0, A).alignment = left

    # Ligne 2 : Unité / Quantité
    ws.cell(r0 + 1, 1).value = "Unité"
    ws.cell(r0 + 1, 2).value = poste.unite
    ws.cell(r0 + 1, 4).value = "Quantité"
    ws.cell(r0 + 1, 5).value = poste.quantite
    ws.cell(r0 + 1, 5).number_format = "0.00"

    # Totaux rapides (référence vers le recap)
    ws.cell(r0 + 2, 1).value = "PR Matière"
    ws.cell(r0 + 2, 4).value = "PR MO"
    ws.cell(r0 + 2, 7).value = "PR Total"
    ws.cell(r0 + 3, 1).value = "PV Total"

    # formules vers recap (on les remplit après)
    # (les cellules de recap seront: N? etc si tu veux à droite; ici on met en dessous -> plus simple)
    # Donc on pointe vers r_recap_start...
    # On définit les cellules récap ci-dessous et on relie.

    # Style en-tête
    style_range(r0, r_header_end, A, L)
    for rr in range(r0, r_header_end + 1):
        for cc in range(A, L + 1):
            ws.cell(rr, cc).fill = header_fill

    # --- TITRE MATIÈRE
    ws.merge_cells(start_row=r_mat_title, start_column=A, end_row=r_mat_title, end_column=L)
    ws.cell(r_mat_title, A).value = "MATIÈRE (10 lignes)"
    ws.cell(r_mat_title, A).font = bold
    ws.cell(r_mat_title, A).fill = title_fill
    ws.cell(r_mat_title, A).alignment = left

    # Entête colonnes matière (12)
    mat_headers = [
        "Ref", "Désignation", "Fournisseur", "Unité",
        "Qté/u poste", "Qté totale", "PU achat", "% pertes",
        "PU corrigé", "Total", "Sécurité", "Commentaire"
    ]
    for i, h in enumerate(mat_headers, start=A):
        ws.cell(r_mat_start - 1, i).value = h
        ws.cell(r_mat_start - 1, i).font = bold
        ws.cell(r_mat_start - 1, i).fill = header_fill
        ws.cell(r_mat_start - 1, i).alignment = center
        ws.cell(r_mat_start - 1, i).border = border

    # Lignes matière vides + formules
    for rr in range(r_mat_start, r_mat_end + 1):
        # Qté totale = Qté/u poste * Quantité poste (cellule E header = r0+1 col 5)
        ws.cell(rr, 6).value = f"=E{rr}*$E${r0+1}"
        # PU corrigé = PU achat * (1 + % pertes)
        ws.cell(rr, 9).value = f"=G{rr}*(1+H{rr})"
        # Total = Qté totale * PU corrigé
        ws.cell(rr, 10).value = f"=F{rr}*I{rr}"

        # formats
        ws.cell(rr, 5).number_format = "0.00"
        ws.cell(rr, 6).number_format = "0.00"
        ws.cell(rr, 7).number_format = "#,##0.00"
        ws.cell(rr, 8).number_format = "0.00%"
        ws.cell(rr, 9).number_format = "#,##0.00"
        ws.cell(rr, 10).number_format = "#,##0.00"
        ws.cell(rr, 11).number_format = "#,##0.00"

    style_range(r_mat_start, r_mat_end, A, L)

    # Sous-total matière (Total colonne J + sécurité colonne K si tu veux)
    ws.cell(r_mat_total, 9).value = "Sous-total matière"
    ws.cell(r_mat_total, 9).font = bold
    ws.cell(r_mat_total, 10).value = f"=SUM(J{r_mat_start}:J{r_mat_end})"
    ws.cell(r_mat_total, 10).font = bold
    ws.cell(r_mat_total, 10).number_format = "#,##0.00"

    ws.cell(r_mat_total, 11).value = f"=SUM(K{r_mat_start}:K{r_mat_end})"
    ws.cell(r_mat_total, 11).font = bold
    ws.cell(r_mat_total, 11).number_format = "#,##0.00"

    style_range(r_mat_total, r_mat_total, A, L)
    ws.cell(r_mat_total, 9).fill = header_fill
    ws.cell(r_mat_total, 10).fill = header_fill
    ws.cell(r_mat_total, 11).fill = header_fill

    # --- TITRE MO
    ws.merge_cells(start_row=r_mo_title, start_column=A, end_row=r_mo_title, end_column=L)
    ws.cell(r_mo_title, A).value = "MAIN-D’ŒUVRE (10 lignes)"
    ws.cell(r_mo_title, A).font = bold
    ws.cell(r_mo_title, A).fill = title_fill
    ws.cell(r_mo_title, A).alignment = left

    # Entête colonnes MO (12)
    mo_headers = [
        "Métier/Équipe", "Nb pers.", "Rendement", "Mode (h/u ou u/h)",
        "Qté poste", "Heures calc.", "Heures prévues", "Coût horaire",
        "Coût MO", "Sécurité", "Total MO+sécu", "Commentaire"
    ]
    for i, h in enumerate(mo_headers, start=A):
        ws.cell(r_mo_start - 1, i).value = h
        ws.cell(r_mo_start - 1, i).font = bold
        ws.cell(r_mo_start - 1, i).fill = header_fill
        ws.cell(r_mo_start - 1, i).alignment = center
        ws.cell(r_mo_start - 1, i).border = border

    for rr in range(r_mo_start, r_mo_end + 1):
        # Qté poste = Quantité poste (cellule $E$header)
        ws.cell(rr, 5).value = f"=$E${r0+1}"

        # Heures calc:
        # si mode = "h/u" => heures = rendement * qté
        # si mode = "u/h" => heures = qté / rendement
        ws.cell(rr, 6).value = f'=IF(D{rr}="h/u", C{rr}*E{rr}, IF(D{rr}="u/h", E{rr}/C{rr}, ""))'

        # Heures prévues: si G vide => prend heures calc, sinon override
        ws.cell(rr, 7).value = f"=IF(G{rr}=\"\", F{rr}, G{rr})"

        # Coût MO = heures prévues * coût horaire
        ws.cell(rr, 9).value = f"=G{rr}*H{rr}"

        # Total MO + sécu
        ws.cell(rr, 11).value = f"=I{rr}+J{rr}"

        # formats
        ws.cell(rr, 2).number_format = "0"
        ws.cell(rr, 3).number_format = "0.00"
        ws.cell(rr, 5).number_format = "0.00"
        ws.cell(rr, 6).number_format = "0.00"
        ws.cell(rr, 7).number_format = "0.00"
        ws.cell(rr, 8).number_format = "#,##0.00"
        ws.cell(rr, 9).number_format = "#,##0.00"
        ws.cell(rr, 10).number_format = "#,##0.00"
        ws.cell(rr, 11).number_format = "#,##0.00"

    style_range(r_mo_start, r_mo_end, A, L)

    # Sous-total MO
    ws.cell(r_mo_total, 9).value = "Sous-total MO"
    ws.cell(r_mo_total, 9).font = bold
    ws.cell(r_mo_total, 10).value = f"=SUM(I{r_mo_start}:I{r_mo_end})"
    ws.cell(r_mo_total, 10).font = bold
    ws.cell(r_mo_total, 10).number_format = "#,##0.00"
    ws.cell(r_mo_total, 11).value = f"=SUM(J{r_mo_start}:J{r_mo_end})"
    ws.cell(r_mo_total, 11).font = bold
    ws.cell(r_mo_total, 11).number_format = "#,##0.00"

    style_range(r_mo_total, r_mo_total, A, L)
    ws.cell(r_mo_total, 9).fill = header_fill
    ws.cell(r_mo_total, 10).fill = header_fill
    ws.cell(r_mo_total, 11).fill = header_fill

    # --- RÉCAP poste (en dessous pour rester simple)
    # On met les totaux calculés ici, puis on référence dans l'en-tête
    ws.merge_cells(start_row=r_recap_start, start_column=A, end_row=r_recap_start, end_column=L)
    ws.cell(r_recap_start, A).value = "RÉCAPITULATIF POSTE"
    ws.cell(r_recap_start, A).font = bold
    ws.cell(r_recap_start, A).fill = title_fill
    ws.cell(r_recap_start, A).alignment = left
    style_range(r_recap_start, r_recap_start, A, L)

    # Lignes recap
    # PR matière = sous-total matière (J mat_total)
    ws.cell(r_recap_start + 1, 1).value = "PR matière"
    ws.cell(r_recap_start + 1, 2).value = f"=J{r_mat_total}"

    # PR MO = sous-total MO (J mo_total)
    ws.cell(r_recap_start + 2, 1).value = "PR main-d’œuvre"
    ws.cell(r_recap_start + 2, 2).value = f"=J{r_mo_total}"

    # Sécurité totale = K matière + K MO
    ws.cell(r_recap_start + 3, 1).value = "Sécurité"
    ws.cell(r_recap_start + 3, 2).value = f"=K{r_mat_total}+K{r_mo_total}"

    # PR total
    ws.cell(r_recap_start + 4, 1).value = "PR total"
    ws.cell(r_recap_start + 4, 2).value = f"=B{r_recap_start+1}+B{r_recap_start+2}+B{r_recap_start+3}"

    # Marge (ex: % sur PR) + PV (tu adapteras)
    ws.cell(r_recap_start + 1, 4).value = "Marge %"
    ws.cell(r_recap_start + 1, 5).value = 0.15
    ws.cell(r_recap_start + 1, 5).number_format = "0.00%"

    ws.cell(r_recap_start + 2, 4).value = "PV HT"
    ws.cell(r_recap_start + 2, 5).value = f"=B{r_recap_start+4}*(1+E{r_recap_start+1})"
    ws.cell(r_recap_start + 2, 5).number_format = "#,##0.00"

    ws.cell(r_recap_start + 3, 4).value = "TVA %"
    ws.cell(r_recap_start + 3, 5).value = 0.21
    ws.cell(r_recap_start + 3, 5).number_format = "0.00%"

    ws.cell(r_recap_start + 4, 4).value = "PV TTC"
    ws.cell(r_recap_start + 4, 5).value = f"=E{r_recap_start+2}*(1+E{r_recap_start+3})"
    ws.cell(r_recap_start + 4, 5).number_format = "#,##0.00"

    # Style recap
    for rr in range(r_recap_start + 1, r_recap_end + 1):
        for cc in range(1, 12 + 1):
            ws.cell(rr, cc).border = border
            ws.cell(rr, cc).alignment = Alignment(vertical="center")
    for rr in range(r_recap_start + 1, r_recap_end + 1):
        ws.cell(rr, 1).font = bold
        ws.cell(rr, 4).font = bold

    # --- Lier l'en-tête aux recap
    ws.cell(r0 + 2, 2).value = f"=B{r_recap_start+1}"  # PR matière
    ws.cell(r0 + 2, 5).value = f"=B{r_recap_start+2}"  # PR MO
    ws.cell(r0 + 2, 8).value = f"=B{r_recap_start+4}"  # PR total
    ws.cell(r0 + 3, 2).value = f"=E{r_recap_start+2}"  # PV HT (ou TTC si tu veux)

    # Formats en-tête chiffres
    ws.cell(r0 + 2, 2).number_format = "#,##0.00"
    ws.cell(r0 + 2, 5).number_format = "#,##0.00"
    ws.cell(r0 + 2, 8).number_format = "#,##0.00"
    ws.cell(r0 + 3, 2).number_format = "#,##0.00"

    # Largeurs colonnes (une fois suffirait, mais ok)
    widths = [12, 40, 18, 10, 14, 14, 14, 12, 14, 14, 14, 22]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # Hauteurs titres
    ws.row_dimensions[r0].height = 22
    ws.row_dimensions[r_mat_title].height = 18
    ws.row_dimensions[r_mo_title].height = 18
    ws.row_dimensions[r_recap_start].height = 18

    return next_top


# =========================
# Génération fichier chiffrage
# =========================
def generer_chiffrage(postes: list[Poste], out_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Chiffrage"

    top = 1
    for p in postes:
        top = ecrire_bloc_poste(ws, top, p)

    wb.save(out_path)
    return out_path


# =========================
# Exemple d'utilisation
# =========================
if __name__ == "__main__":
    bordereau = Path("bordereau.xlsx")  # <-- ton fichier importé
    postes = importer_bordereau_xlsx(
        bordereau,
        sheet_name=None,
        col_code="A",
        col_designation="B",
        col_unite="C",
        col_quantite="D",
        start_row=2,
    )

    out = Path("CHIFFRAGE_GENERÉ.xlsx")
    generer_chiffrage(postes, out)
    print(f"OK -> {out.resolve()}")
