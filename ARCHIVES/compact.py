from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from pathlib import Path

# Fichier de sortie
out = Path("DATA/tableau_compact.xlsx")

wb = Workbook()
ws = wb.active
ws.title = "Compact"

# Colonnes compactes
headers = [
    "Poste",
    "Désignation",
    "Type",
    "Unité",
    "Quantité",
    "PU revient",
    "Total revient",
    "Marge %",
    "PU vente",
    "Total vente"
]

ws.append(headers)

# Style simple
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# Exemple de lignes (modifiables)
data = [
    ["3.01", "Transport", "Matière", "forfait", 1, 100, "", 20, "", ""],
    ["", "Petit matériel", "Matière", "forfait", 1, 50, "", 20, "", ""],
    ["", "Briques", "Matière", "m2", 10, 12, "", 20, "", ""],
    ["", "Pierres", "Matière", "m2", 10, 20, "", 20, "", ""],
    ["", "Maçonnerie briques", "Main-d’œuvre", "h/homme", 8, 45, "", 20, "", ""],
]

for row in data:
    ws.append(row)

# Formules automatiques
for r in range(2, ws.max_row + 1):
    ws[f"G{r}"] = f"=E{r}*F{r}"                 # Total revient
    ws[f"I{r}"] = f"=F{r}*(1+H{r}/100)"         # PU vente
    ws[f"J{r}"] = f"=E{r}*I{r}"                 # Total vente

# Largeur colonnes
widths = [10, 25, 14, 10, 10, 12, 14, 10, 12, 14]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64+i)].width = w

wb.save(out)
print("✅ Tableau compact créé :", out)
