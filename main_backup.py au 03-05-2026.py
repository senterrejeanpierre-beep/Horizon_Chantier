#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import subprocess
import shutil
import re
import time
from pathlib import Path
import tkinter as tk
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

from tkinter import ttk, filedialog, messagebox

from openpyxl import load_workbook

APP_NAME = "Horizon Chantier"

# =========================
# FICHIERS (STRICT / SANS REFONTE)
# =========================
PV_SOURCE = "soumission_chapelette__MAJ_PV_2026-01-29_13-52-57.xlsx"
PV_BACKUP = "soumission_chapelette__backup_avant_injectionPV_2026-01-29_13-52-57.xlsx"
PV_CORRIGE = "soumission_chapelette__MAJ_injection_CORRIGE_PV_2026-01-29_13-52-57.xlsx"
FICHIER_PV = PV_SOURCE


# =========================
# AppleScript helpers (Excel ouvert)
# =========================
def _osascript(script: str) -> str:
    p = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
    if p.returncode != 0:
        raise RuntimeError(p.stderr.strip() or "Erreur AppleScript")
    return (p.stdout or "").strip()


def _a1_to_col_row(a1: str) -> tuple[str, int]:
    """
    "C12" -> ("C", 12)
    """
    m = re.fullmatch(r"([A-Z]+)(\d+)", a1.strip().upper().replace("$", ""))
    if not m:
        raise ValueError(f"Adresse cellule invalide : {a1}")
    return m.group(1), int(m.group(2))


def _is_allowed_target(cell_a1: str) -> bool:
    """
    Sécurité absolue : n'autorise que
    - P3:P12  (matière)
    - P17:P26 (main-d'œuvre)
    """
    col, row = _a1_to_col_row(cell_a1)
    if col == "P" and 3 <= row <= 12:
        return True
    if col == "P" and 17 <= row <= 26:
        return True
    return False


def _normalize_excel_addr(addr: str) -> str:
    """
    Normalise:
    - "$P$3" -> "P3"
    - "P3"   -> "P3"
    """
    if not addr:
        return ""
    a = addr.strip().replace("$", "").upper()
    if re.fullmatch(r"[A-Z]+[0-9]+", a):
        return a
    return ""


def excel_find_first_empty_cell(workbook_hint: str, sheet_name: str, cell_range: str) -> str:
    """
    1ère cellule vide dans une plage (colonne unique) ex: P3:P12
    Retour "P3" ou "" si aucune.
    Robuste: retrouve le classeur par "contient" (Excel Mac change parfois le nom).
    """
    hint = (workbook_hint or "").strip()
    if not hint:
        return ""

    for _ in range(8):
        script = f'''
        tell application "Microsoft Excel"
            set wbRef to missing value
            repeat with w in workbooks
                try
                    set wn to (name of w as string)
                    if wn contains "{hint}" then
                        set wbRef to w
                        exit repeat
                    end if
                end try
            end repeat

            if wbRef is missing value then return ""

            tell wbRef
                activate
                try
                    tell worksheet "{sheet_name}"
                        activate
                        set rng to range "{cell_range}"
                        repeat with c in cells of rng
                            try
                                set v to value of c
                                set f to formula of c

                                if v is missing value then return (address of c) as string

                                try
                                    if (v as string) is "" then return (address of c) as string
                                end try

                                -- si formule et résultat vide
                                try
                                    if (f as string) is not "" then
                                        if (v as string) is "" then return (address of c) as string
                                    end if
                                end try

                            on error
                                return (address of c) as string
                            end try
                        end repeat
                    end tell
                end try
            end tell
        end tell
        return ""
        '''
        try:
            addr_raw = _osascript(script)
            addr = _normalize_excel_addr(addr_raw)
            if addr:
                return addr
        except Exception:
            pass

        time.sleep(0.2)

    return ""


def excel_set_cell_value(workbook_hint: str, sheet_name: str, cell_a1: str, value: str) -> None:
    """
    Écrit une valeur dans Excel (classeur déjà ouvert) sans fermer.
    Sécurité : n'écrit QUE dans P3:P12 ou P17:P26.
    Robuste: retrouve le classeur par "contient".
    """
    cell_a1 = (cell_a1 or "").strip().replace("$", "").upper()
    if not _is_allowed_target(cell_a1):
        raise ValueError(f"Sécurité: écriture interdite hors zones autorisées : {cell_a1}")

    hint = (workbook_hint or "").strip()
    if not hint:
        raise ValueError("Nom de classeur vide")

    safe_val = (value or "").replace('"', '\\"')

    script = f'''
    tell application "Microsoft Excel"
        set wbRef to missing value
        repeat with w in workbooks
            try
                set wn to (name of w as string)
                if wn contains "{hint}" then
                    set wbRef to w
                    exit repeat
                end if
            end try
        end repeat

        if wbRef is missing value then return

        tell wbRef
            if not (exists worksheet "{sheet_name}") then return
            tell worksheet "{sheet_name}"
                set value of range "{cell_a1}" to "{safe_val}"
            end tell
        end tell
        activate
    end tell
    '''
    _osascript(script)


# =========================
# Lecture bibliothèques (colonne A) - lecture seule
# =========================
def read_biblio_colA_openpyxl(pr_path: Path, sheet_name: str) -> list[str]:
    wb = load_workbook(pr_path, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Onglet introuvable : {sheet_name}")
        ws = wb[sheet_name]
        out: list[str] = []
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v is None:
                continue
            s = str(v).strip()
            if s:
                out.append(s)
        return out
    finally:
        wb.close()


# =========================
# Popup recherche (20 résultats)
# =========================
def popup_search_20(parent: tk.Tk, titre: str, items: list[str], on_pick) -> None:
    win = tk.Toplevel(parent)
    win.title(titre)
    win.geometry("780x520")
    win.transient(parent)
    win.grab_set()

    frm = ttk.Frame(win, padding=12)
    frm.pack(fill="both", expand=True)

    ttk.Label(frm, text=titre, font=("Helvetica", 14, "bold")).pack(anchor="w")

    q = tk.StringVar()
    entry = ttk.Entry(frm, textvariable=q)
    entry.pack(fill="x", pady=(10, 8))
    entry.focus_set()

    lb = tk.Listbox(frm, height=18)
    lb.pack(fill="both", expand=True)

    ttk.Label(frm, text="Tape 2 lettres. Double-clic ou Entrée pour valider.", foreground="#555").pack(
        anchor="w", pady=(8, 0)
    )

    current: list[str] = []

    def refresh(*_):
        text = q.get().strip().lower()
        lb.delete(0, tk.END)

        if not text:
            current[:] = items[:20]
        else:
            filtered = [s for s in items if text in s.lower()]
            current[:] = filtered[:20]

        for s in current:
            lb.insert(tk.END, s)

        if current:
            lb.selection_set(0)

    def choose(_event=None):
        sel = lb.curselection()
        if not sel:
            return
        picked = current[sel[0]]
        try:
            on_pick(picked)
        finally:
            win.destroy()

    q.trace_add("write", refresh)
    lb.bind("<Double-Button-1>", choose)
    win.bind("<Return>", choose)

    refresh()


# ============================================================
# Montant en lettres FR (euros + centimes) - inchangé
# ============================================================
_UNITS = ["zéro", "un", "deux", "trois", "quatre", "cinq", "six", "sept", "huit", "neuf"]
_TEENS = ["dix", "onze", "douze", "treize", "quatorze", "quinze", "seize", "dix-sept", "dix-huit", "dix-neuf"]
_TENS = ["", "", "vingt", "trente", "quarante", "cinquante", "soixante", "soixante", "quatre-vingt", "quatre-vingt"]


def _below_100(n: int) -> str:
    assert 0 <= n < 100
    if n < 10:
        return _UNITS[n]
    if 10 <= n < 20:
        return _TEENS[n - 10]

    ten = n // 10
    unit = n % 10

    if ten == 7:
        return "soixante" + ("-" + _below_100(10 + unit) if unit else "-dix")
    if ten == 9:
        return "quatre-vingt" + ("-" + _below_100(10 + unit) if unit else "-dix")

    base = _TENS[ten]

    if ten == 8 and unit == 0:
        return "quatre-vingts"
    if unit == 0:
        return base

    if unit == 1 and ten in (2, 3, 4, 5, 6):
        return f"{base} et un"

    return f"{base}-{_UNITS[unit]}"


def _below_1000(n: int) -> str:
    assert 0 <= n < 1000
    if n < 100:
        return _below_100(n)

    hundred = n // 100
    rest = n % 100

    if hundred == 1:
        head = "cent"
    else:
        head = f"{_UNITS[hundred]} cent"

    if rest == 0 and hundred > 1:
        return head + "s"
    if rest == 0:
        return head

    return head + " " + _below_100(rest)


def _int_to_words_fr(n: int) -> str:
    if n == 0:
        return "zéro"

    parts: list[str] = []

    billions = n // 1_000_000_000
    n %= 1_000_000_000
    millions = n // 1_000_000
    n %= 1_000_000
    thousands = n // 1000
    n %= 1000
    rest = n

    if billions:
        parts.append("un milliard" if billions == 1 else _below_1000(billions) + " milliards")

    if millions:
        parts.append("un million" if millions == 1 else _below_1000(millions) + " millions")

    if thousands:
        parts.append("mille" if thousands == 1 else _below_1000(thousands) + " mille")

    if rest:
        parts.append(_below_1000(rest))

    return " ".join(parts)


def montant_en_lettres_fr(valeur: float) -> str:
    v = round(float(valeur), 2)
    euros = int(v)
    centimes = int(round((v - euros) * 100))
    if centimes == 100:
        euros += 1
        centimes = 0

    euros_txt = _int_to_words_fr(euros)
    euro_label = "euro" if euros == 1 else "euros"

    if centimes == 0:
        return f"{euros_txt} {euro_label}"

    cent_txt = _int_to_words_fr(centimes)
    cent_label = "centime" if centimes == 1 else "centimes"
    return f"{euros_txt} {euro_label} et {cent_txt} {cent_label}"


# ============================================================
# PR / PV (inchangés)
# ============================================================
def open_pr(chantier_dir: str | Path):
    chantier = Path(chantier_dir)
    pr = chantier / "data" / "prix_de_revient.xlsx"
    if not pr.exists():
        raise FileNotFoundError(f"PR introuvable : {pr}")
    subprocess.run(["open", str(pr)], check=False)


def inject_pv(chantier_dir: str | Path) -> int:
    chantier = Path(chantier_dir)
    pr_path = chantier / "data" / "prix_de_revient.xlsx"
    pv_path = chantier / PV_SOURCE

    if not pr_path.exists():
        raise FileNotFoundError(f"PR introuvable : {pr_path}")
    if not pv_path.exists():
        raise FileNotFoundError(f"PV source introuvable : {pv_path}")

    pr_wb = load_workbook(pr_path, data_only=True)
    pr_ws = pr_wb["Chiffrage"] if "Chiffrage" in pr_wb.sheetnames else pr_wb.active


    # trouver la 1ère ligne d'article à partir de A3
    row_article = None
    for r in range(3, pr_ws.max_row + 1):
        v = pr_ws.cell(row=r, column=1).value  # colonne A
        if v is not None and str(v).strip() != "":
            row_article = r
            break

    if row_article is None:
        pr_wb.close()
        raise ValueError("Aucun ARTICLE trouvé dans la colonne A (à partir de A3).")

    article = pr_ws.cell(row=row_article, column=1).value
    designation = pr_ws.cell(row=row_article, column=2).value

    # on garde ton prix comme avant
    price = pr_ws["E32"].value

    pr_wb.close()



    if article is None or str(article) == "":
        raise ValueError("ARTICLE vide dans PR (A3).")
    if price is None or str(price) == "":
        raise ValueError("PRIX vide dans PR (E32).")

    article = str(article)
    designation = "" if designation is None else str(designation)

    price_f = float(price)
    price_2d = round(price_f, 2)
    letters = montant_en_lettres_fr(price_2d)

    wb = load_workbook(pv_path)
    ws = wb["PV"] if "PV" in wb.sheetnames else wb.active

    row_found = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        if str(v) == article:
            row_found = r
            break

    if row_found is None:
        wb.close()
        raise ValueError(f"Article introuvable dans PV source : '{article}'")

    ws.cell(row=row_found, column=2).value = designation

    g = ws.cell(row=row_found, column=7)
    g.value = price_2d
    g.number_format = "#,##0.00"

    ws.cell(row=row_found, column=8).value = letters

    wb.save(pv_path)
    wb.close()
    return 1


def sync_pv_to_copies(chantier_dir: str | Path) -> None:
    chantier = Path(chantier_dir)

    src = chantier / PV_SOURCE
    tgt_backup = chantier / PV_BACKUP
    tgt_corrige = chantier / PV_CORRIGE

    for p in (src, tgt_backup, tgt_corrige):
        if not p.exists():
            raise FileNotFoundError(f"Fichier introuvable : {p}")

    src_wb = load_workbook(src, data_only=True)
    src_ws = src_wb["PV"] if "PV" in src_wb.sheetnames else src_wb.active
    data = [row for row in src_ws.iter_rows(values_only=True)]
    src_wb.close()

    def _write(target: Path):
        wb = load_workbook(target)
        ws = wb["PV"] if "PV" in wb.sheetnames else wb.active

        for r_idx, row in enumerate(data, start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if cell.__class__.__name__ == "MergedCell":
                    continue
                cell.value = val

        wb.save(target)
        wb.close()

    _write(tgt_backup)
    _write(tgt_corrige)


# ---------------------------
# Dossiers (auto Chantier/Chantiers)
# ---------------------------
def dossier_base() -> Path:
    return Path.home() / "Desktop" / "Horizon_Chantier_Data"


def dossier_chantiers() -> Path:
    base = dossier_base()
    d1 = base / "Chantier"
    d2 = base / "Chantiers"
    if d1.exists():
        return d1
    if d2.exists():
        return d2
    d2.mkdir(parents=True, exist_ok=True)
    return d2
def ouvrir_doc(self, nom_fichier):
    sel = self.tree.selection()
    if not sel:
        messagebox.showwarning("Sélection", "Sélectionne un chantier.")
        return

    item = self.tree.item(sel[0])

    # 1) Nom sélectionné (text si présent, sinon values[0])
    chantier_sel = (item.get("text") or "").strip()
    if not chantier_sel:
        vals = item.get("values", [])
        chantier_sel = str(vals[0]).strip() if vals else ""

    base = dossier_chantiers()

    # 2) On cherche le dossier chantier qui existe vraiment
    dossier = base / chantier_sel
    if not dossier.exists():
        def norm(s):
            return s.lower().strip().replace(" ", "_").replace("-", "_")

        cible = norm(chantier_sel)
        dossier = next(
            (p for p in base.iterdir() if p.is_dir() and norm(p.name) == cible),
            None
        )

    if not dossier or not Path(dossier).exists():
        messagebox.showerror(
            "Introuvable",
            f"Dossier chantier introuvable pour :\n{chantier_sel}\n\nBase : {base}"
        )
        return

    chemin = Path(dossier) / nom_fichier

    if not chemin.exists():
        # Si on n'a pas trouvé le fichier, on tente l'autre dossier chantier
        # (ex: "soumission_chapelette" vs "Soumission Chapelette")
        def norm(s):
            return str(s).lower().strip().replace(" ", "_").replace("-", "_")

        cible = norm(chantier_sel)
        for p in base.iterdir():
            if p.is_dir() and norm(p.name) == cible:
                alt = p / nom_fichier
                if alt.exists():
                   chemin = alt
                   break

    if not chemin.exists():
        messagebox.showerror("Introuvable", f"Fichier absent :\n{chemin}")
        return

    if sys.platform == "darwin":
        subprocess.run(["open", str(chemin)], check=False)
    elif os.name == "nt":
        os.startfile(str(chemin))

def ouvrir_dossier(path: Path) -> None:
    try:
        subprocess.run(["open", str(path)], check=False)
    except Exception as e:
        messagebox.showerror("Erreur", f"Impossible d'ouvrir le dossier.\n{e}")


# ---------------------------
# Lecture chantier JSON
# ---------------------------
def lire_json(path: Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def ecrire_json(path: Path, data: dict) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def infos_chantier(chantier: dict) -> dict:
    return {
        "nom": chantier.get("nom", chantier.get("chantier", "")) or "",
        "client": chantier.get("client", "") or "",
        "etat": chantier.get("etat", chantier.get("type_travaux", "")) or "",
        "avancement": chantier.get("avancement", chantier.get("avancement_pct", 0)) or 0,
        "debut": chantier.get("date_debut", chantier.get("debut_travaux", "")) or "",
        "fin": chantier.get("date_fin", chantier.get("fin_prevue", chantier.get("date_fin_prevue", ""))) or "",
    }


# ---------------------------
# Fenêtre Bordereau (visu JSON)
# ---------------------------
def afficher_bordereau(parent: tk.Tk, chantier: dict, titre: str = "Bordereau") -> None:
    bord = chantier.get("bordereau", {})
    articles = bord.get("articles", {})

    if not articles:
        messagebox.showwarning("Bordereau", "Aucun bordereau dans ce chantier.")
        return

    win = tk.Toplevel(parent)
    win.title(titre)
    win.geometry("1200x650")

    cols = ("Article", "Libellé", "Unité", "Qté", "PU", "Total")

    tree = ttk.Treeview(win, columns=cols, show="headings")
    vsb = ttk.Scrollbar(win, orient="vertical", command=tree.yview)
    hsb = ttk.Scrollbar(win, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

    for c in cols:
        tree.heading(c, text=c)
        if c == "Libellé":
            tree.column(c, width=700, anchor="w")
        elif c == "Article":
            tree.column(c, width=160, anchor="w")
        elif c == "Unité":
            tree.column(c, width=80, anchor="w")
        else:
            tree.column(c, width=120, anchor="e")

    def v(x):
        return "" if x is None else x

    for article_id in sorted(articles.keys()):
        a = articles[article_id] or {}
        tree.insert(
            "",
            "end",
            values=(
                v(a.get("article_id", article_id)),
                v(a.get("libelle", "")),
                v(a.get("unite", "")),
                v(a.get("quantite", 0)),
                v(a.get("pu_vente", 0)),
                v(a.get("total_vente", 0)),
            ),
        )

    tree.grid(row=0, column=0, sticky="nsew")
    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")

    win.grid_rowconfigure(0, weight=1)
    win.grid_columnconfigure(0, weight=1)


# ---------------------------
# App
# ---------------------------
class HorizonChantierApp(tk.Tk):


    def __init__(self) -> None:
        super().__init__()
        self.title(APP_NAME)
        self.geometry("1400x800")
        self.minsize(1300, 950)

        self._build_ui()
        self.refresh_liste()

    def _build_ui(self) -> None:
        root = ttk.Frame(self, padding=18)
        root.pack(fill="both", expand=True)
        style = ttk.Style()
        style.configure("TitreBloc.TLabel",
                font=("Helvetica", 12, "bold"),
                foreground="#1f4e79")
                
        title = ttk.Label(root, text=APP_NAME, font=("Helvetica", 22, "bold"))
        title.pack(anchor="w")

        self.lbl_path = ttk.Label(root, text=f"Emplacement: {dossier_chantiers()}")
        self.lbl_path.pack(fill="x", pady=(4, 8))

        cols = ("Chantier", "Client", "État", "Avancement", "Début", "Fin prévue")

        zone = ttk.Frame(root)
        zone.pack(fill="both", expand=True)

        zone.columnconfigure(0, weight=0)
        zone.columnconfigure(1, weight=1)
        zone.columnconfigure(2, weight=0)
        zone.rowconfigure(0, weight=1)

        frame_gauche = ttk.Frame(zone)
        frame_gauche.grid(row=0, column=0, sticky="ns", padx=(0, 12))

        frame_centre = ttk.Frame(zone)
        frame_centre.grid(row=0, column=1, sticky="nsew")

        frame_droite = ttk.Frame(zone)
        frame_droite.grid(row=0, column=2, sticky="ns", padx=(12, 0))

        self.tree = ttk.Treeview(frame_centre, columns=cols, show="headings")
        for c in cols:
            self.tree.heading(c, text=c)
            if c == "Chantier":
                self.tree.column(c, width=260, anchor="w")
            elif c == "Client":
                self.tree.column(c, width=180, anchor="w")
            elif c == "État":
                self.tree.column(c, width=120, anchor="w")
            elif c == "Avancement":
                self.tree.column(c, width=110, anchor="e")
            else:
                self.tree.column(c, width=120, anchor="w")

        self.tree.pack(fill="both", expand=True)

        # -------------------------
        # GAUCHE
        # -------------------------

        ttk.Separator(frame_gauche).pack(fill="x", pady=10)
        tk.Label(frame_gauche, text="AIDE",
            font=("Helvetica", 15, "bold"),
            fg="#C62828").pack(anchor="w", pady=(2, 4))
        ttk.Separator(frame_gauche).pack(fill="x", pady=(0, 8))
        ttk.Button(frame_gauche, text="Aide / Dépannage", command=self.ouvrir_aide).pack(fill="x", pady=3)

        ttk.Separator(frame_gauche).pack(fill="x", pady=10)
        tk.Label(frame_gauche, text="GESTION",
            font=("Helvetica", 14, "bold"),
            fg="#0052cc").pack(anchor="w", pady=(0, 2))
        ttk.Separator(frame_gauche).pack(fill="x", pady=(0, 8))

        ttk.Button(frame_gauche, text="📄 Ouvrir dans le logiciel", command=self.ouvrir_chantier).pack(fill="x", pady=3)
        ttk.Button(frame_gauche, text="➕ Nouveau chantier", command=self.nouveau_chantier).pack(fill="x", pady=3)
        ttk.Button(frame_gauche, text="✏️ Modifier", command=self.modifier_chantier).pack(fill="x", pady=3)
        ttk.Button(frame_gauche, text="🗑️ Supprimer", command=self.supprimer_chantier).pack(fill="x", pady=3)
        ttk.Button(frame_gauche, text="🔄 Rafraîchir", command=self.refresh_liste).pack(fill="x", pady=3)

        ttk.Separator(frame_gauche).pack(fill="x", pady=10)

        tk.Label(frame_gauche, text="DOCUMENTS",
         font=("Helvetica", 15, "bold"),
         fg="#0052cc").pack(anchor="w", pady=(2, 4))
        ttk.Separator(frame_gauche).pack(fill="x", pady=(0, 8))

        ttk.Button(frame_gauche, text="📁 Ce chantier", command=self.dossier_du_chantier).pack(fill="x", pady=3)

        ttk.Button(
            frame_gauche,
            text="🔎 Dossier administratif",
            command=lambda: __import__("subprocess").run([
                "open",
                str(dossier_chantiers() / self.tree.item(self.tree.selection()[0])["values"][0] / "Administratif")
            ])
        ).pack(fill="x", pady=3)

        ttk.Button(
            frame_gauche,
            text="📄 Cahier des charges administratif",
            command=lambda: __import__("subprocess").Popen([
                "open",
                str(dossier_chantiers() / self.tree.item(self.tree.selection()[0])["values"][0] / "Cahier_administratif.pdf")
            ])
        ).pack(fill="x", pady=3)

        ttk.Button(
           frame_gauche,
            text="🛠 Cahier des charges technique",
            command=lambda: __import__("subprocess").Popen([
                "open",
                str(dossier_chantiers() / self.tree.item(self.tree.selection()[0])["values"][0] / "Cahier_technique.pdf")
            ])
        ).pack(fill="x", pady=3)

        ttk.Button(
            frame_gauche,
            text="📐 Postes / Métré",
            command=lambda: __import__("subprocess").run([
                "open",
                str(dossier_chantiers() / self.tree.item(self.tree.selection()[0])["values"][0] / "Metre_detaille.pdf")
            ])
        ).pack(fill="x", pady=3)

        ttk.Button(
            frame_gauche,
            text="🦺 Plan de sécurité (PSS)",
            command=lambda: __import__("subprocess").Popen([
                "open",
                str(dossier_chantiers() / self.tree.item(self.tree.selection()[0])["values"][0] / "PSS.pdf")
            ])
        ).pack(fill="x", pady=3)

        ttk.Button(
            frame_gauche,
            text="📄 Décompte intempéries",
            command=lambda: __import__("subprocess").run([
                "open",
                str(dossier_chantiers() / self.tree.item(self.tree.selection()[0])["values"][0] / "Décompte_intempéries.doc")
            ])
        ).pack(fill="x", pady=3)

        # -------------------------
        # DROITE
        # -------------------------

        

        ttk.Separator(frame_droite).pack(fill="x", pady=10)

        tk.Label(frame_droite, text="EXECUTION",
         font=("Helvetica", 15, "bold"),
         fg="#0052cc").pack(anchor="w", pady=(2, 4))
        ttk.Separator(frame_droite).pack(fill="x", pady=(0, 8))

        ttk.Button(
            frame_droite,
            text="📅 Planning d’exécution",
            command=lambda: __import__("subprocess").Popen([
                "open",
                str(dossier_chantiers() / self.tree.item(self.tree.selection()[0])["values"][0] / "Planning_execution.xlsx")
            ])
        ).pack(fill="x", pady=3)

        ttk.Separator(frame_droite).pack(fill="x", pady=10)

        tk.Label(frame_droite, text="CALCULS",
         font=("Helvetica", 15, "bold"),
         fg="#0052cc").pack(anchor="w", pady=(2, 4))
        ttk.Separator(frame_droite).pack(fill="x", pady=(0, 8))

        ttk.Button(
            frame_droite,
            text="🔎 Coût de la sécurité",
            command=lambda: __import__("subprocess").Popen([
                "open",
                str(dossier_chantiers() / self.tree.item(self.tree.selection()[0])["values"][0] / "Cout_securite.xlsx")
            ])
        ).pack(fill="x", pady=3)

        ttk.Button(
            frame_droite,
            text="📐 Formule de révision",
            command=lambda: __import__("subprocess").run([
                "open",
                str(dossier_chantiers() / self.tree.item(self.tree.selection()[0])["values"][0] / "Formule_révision.xlsm")
            ], check=True)
        ).pack(fill="x", pady=3)

        ttk.Button(
            frame_droite,
            text="💰 Prix de revient",
            command=lambda: __import__("subprocess").run([
                "open",
                str(dossier_chantiers() / self.tree.item(self.tree.selection()[0])["values"][0] / "data" / "prix_de_revient.xlsx")
            ])
        ).pack(fill="x", pady=3)

        ttk.Button(frame_droite, text="🔍 Vérifier PR", command=self.verifier_pr).pack(fill="x", pady=3)
        ttk.Button(frame_droite, text="⚙️ Calcul / Reca", command=self.calcul_reca).pack(fill="x", pady=3)

        ttk.Separator(frame_droite).pack(fill="x", pady=10)

        tk.Label(frame_droite, text="ETAT D'AVANCEMENT",
         font=("Helvetica", 15, "bold"),
         fg="#0052cc").pack(anchor="w", pady=(2, 4))
        ttk.Separator(frame_droite).pack(fill="x", pady=(0, 8))

        ttk.Button(
            frame_droite,
            text="📊 État d'avancement",
            command=lambda: subprocess.Popen([
                "open",
                next(Path(
                    dossier_chantiers() / self.tree.item(self.tree.selection()[0])["values"][0]
                ).glob("Etat_avancement*.xlsm")).as_posix()
            ])
        ).pack(fill="x", pady=3)

        ttk.Button(frame_droite, text="📊 Calcul état", command=self.calcul_etat_avancement).pack(fill="x", pady=3)
        ttk.Button(frame_droite, text="♻️ Clôturer état", command=self.mise_a_zero_etat).pack(fill="x", pady=3)

        ttk.Separator(frame_droite).pack(fill="x", pady=10)

        tk.Label(frame_droite, text="PILOTAGE",
         font=("Helvetica", 15, "bold"),
         fg="#0052cc").pack(anchor="w", pady=(2, 4))
        ttk.Separator(frame_droite).pack(fill="x", pady=(0, 8))

        ttk.Button(frame_droite, text="🧭 Pilotage", command=self.pilotage_chantier).pack(fill="x", pady=3)
        ttk.Button(frame_droite, text="📊 Pilotage délai", command=self.pilotage_delai).pack(fill="x", pady=3)
        ttk.Button(frame_droite, text="⏱ Rendements", command=self.rendements_chantier).pack(fill="x", pady=3)
       
       

    def ouvrir_aide(self) -> None:
        existing = getattr(self, "_help_window", None)
        if existing and existing.winfo_exists():
            existing.deiconify()
            existing.lift()
            existing.focus_force()
            return

        win = tk.Toplevel(self)
        self._help_window = win
        win.title("Aide / Mode d’emploi - Horizon Chantier")
        win.geometry("900x600")
        win.minsize(820, 520)
        win.transient(self)
        topmost_var = tk.BooleanVar(value=True)

        def appliquer_toujours_visible() -> None:
            try:
                win.attributes("-topmost", topmost_var.get())
            except Exception:
                pass

        def fermer() -> None:
            self._help_window = None
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", fermer)
        appliquer_toujours_visible()

        header = ttk.Frame(win, padding=(18, 16, 18, 10))
        header.pack(fill="x")

        ttk.Label(
            header,
            text="Aide / Mode d’emploi - Horizon Chantier",
            font=("Helvetica", 20, "bold"),
            foreground="#163A59",
        ).pack(anchor="w")

        ttk.Label(
            header,
            text="Choisir le guide à afficher puis suivre les consignes sans modifier les formules Excel.",
            font=("Helvetica", 10),
            foreground="#4F6475",
        ).pack(anchor="w", pady=(4, 0))

        zone_choix = ttk.Frame(header)
        zone_choix.pack(anchor="w", pady=(10, 0))
        ttk.Label(zone_choix, text="Aide :", font=("Helvetica", 10, "bold")).pack(side="left", padx=(0, 8))
        aide_var = tk.StringVar(value="Mode d’emploi")
        choix_aide = ttk.Combobox(
            zone_choix,
            textvariable=aide_var,
            values=["Mode d’emploi", "Dépannage PR"],
            state="readonly",
            width=18,
        )
        choix_aide.pack(side="left")

        ttk.Separator(win).pack(fill="x")

        corps = ttk.Frame(win, padding=(14, 12, 14, 8))
        corps.pack(fill="both", expand=True)
        corps.columnconfigure(0, weight=1)
        corps.rowconfigure(0, weight=1)

        texte = tk.Text(
            corps,
            wrap="word",
            font=("Helvetica", 11),
            relief="flat",
            bd=0,
            padx=22,
            pady=18,
            spacing1=2,
            spacing3=7,
            background="#FFFFFF",
            foreground="#1E1E1E",
            insertwidth=0,
        )
        texte.grid(row=0, column=0, sticky="nsew")

        scroll = ttk.Scrollbar(corps, orient="vertical", command=texte.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        texte.configure(yscrollcommand=scroll.set)

        texte.tag_configure("titre", font=("Helvetica", 19, "bold"), foreground="#163A59", spacing3=14)
        texte.tag_configure("intro", font=("Helvetica", 11), foreground="#405261", spacing3=10)
        texte.tag_configure("section", font=("Helvetica", 14, "bold"), foreground="#1F4E79", spacing1=14, spacing3=8)
        texte.tag_configure("texte", font=("Helvetica", 11), foreground="#1E1E1E", spacing3=6)
        texte.tag_configure("liste", font=("Helvetica", 11), foreground="#1E1E1E", lmargin1=22, lmargin2=42, spacing3=4)
        texte.tag_configure("regle", font=("Helvetica", 12, "bold"), foreground="#8A1C1C", background="#FFF3CD", spacing1=8, spacing3=10, lmargin1=10, lmargin2=10)
        texte.tag_configure("formule", font=("Courier", 11, "bold"), foreground="#163A59", spacing3=4, lmargin1=22, lmargin2=42)
        texte.tag_configure("ok", font=("Helvetica", 11, "bold"), foreground="#1E6B34", spacing3=4)
        texte.tag_configure("attention", font=("Helvetica", 11, "bold"), foreground="#9C4F00", spacing3=4)
        texte.tag_configure("danger", font=("Helvetica", 11, "bold"), foreground="#B42318", spacing3=4)
        texte.tag_configure("aide", font=("Helvetica", 11, "bold"), foreground="#5B3F99", spacing3=4)
        texte.tag_configure("fin", font=("Helvetica", 10, "italic"), foreground="#4F6475", spacing1=8, spacing3=10)

        contenu_mode_emploi = [
            ("titre", "Mode d’emploi complet\n"),
            ("intro", "Cette fenêtre est prévue pour qu’un collègue perdu clique sur AIDE et sache quoi faire, dans quel ordre, et surtout ce qu’il ne faut jamais modifier.\n"),
            ("regle", "🎯 RÈGLE PRINCIPALE : si la cellule n’est pas gris clair, on ne touche pas.\n"),
            ("liste", "⚠️ Une cellule blanche, calculée ou décorative ne doit pas être modifiée.\n"),
            ("liste", "❌ Ne jamais corriger un résultat en touchant une formule.\n"),
            ("liste", "✅ En cas de doute: arrêter, enregistrer, fermer Excel si nécessaire, puis relire ce guide.\n"),

            ("section", "\n🧱 1. Créer un chantier\n"),
            ("liste", "✅ Sélectionner la fonction Nouveau chantier.\n"),
            ("liste", "✅ Renseigner les informations de base du chantier.\n"),
            ("liste", "✅ Vérifier que le dossier du chantier a bien été créé.\n"),
            ("liste", "🔍 Le logiciel a besoin de ce dossier pour retrouver les fichiers Excel et produire les documents ensuite.\n"),

            ("section", "\n🔍 2. Importer manuellement les fichiers Excel dans le dossier chantier\n"),
            ("liste", "✅ Une fois le chantier créé, copier manuellement dans le dossier chantier les fichiers Excel nécessaires au chantier.\n"),
            ("liste", "✅ Les fichiers courants sont notamment: prix de revient, état d’avancement, délai, révision, avenants et documents de calcul associés.\n"),
            ("liste", "⚠️ Les noms de fichiers et la structure du dossier doivent rester cohérents avec ce que le logiciel attend.\n"),
            ("liste", "❌ Ne pas renommer ou déplacer un fichier utilisé par le logiciel sans raison précise.\n"),

            ("section", "\n🟨 3. Fonctionnement du PR\n"),
            ("texte", "Le PR est le fichier Prix de revient. Il sert de base pour les prix, les vérifications et certaines injections dans l’état d’avancement.\n"),
            ("liste", "✅ Ouvrir le PR avec le bouton Prix de revient.\n"),
            ("liste", "✅ Encoder uniquement dans les cellules gris clair prévues à cet effet.\n"),
            ("liste", "⚠️ Enregistrer Excel après toute modification.\n"),
            ("liste", "❌ Ne jamais modifier une formule pour forcer un résultat.\n"),

            ("section", "\n✅ 4. Vérifier PR\n"),
            ("texte", "Le bouton Vérifier PR sert à contrôler la cohérence des calculs du fichier PR.\n"),
            ("liste", "✅ À utiliser après encodage et après enregistrement du fichier Excel.\n"),
            ("liste", "🔍 Si une erreur apparaît, il faut corriger la donnée d’entrée dans le PR, pas le résultat calculé ailleurs.\n"),
            ("liste", "⚠️ Si Excel est encore ouvert et que les données n’ont pas été enregistrées, la vérification peut porter sur une ancienne version du fichier.\n"),

            ("section", "\n🟩 5. Importer PR\n"),
            ("texte", "Dans Horizon Chantier, l’import PR se fait via le bouton Calcul / Reca.\n"),
            ("liste", "✅ Ouvrir le PR, encoder ce qui doit l’être, puis enregistrer.\n"),
            ("liste", "✅ Lancer Vérifier PR si nécessaire.\n"),
            ("liste", "✅ Ensuite lancer Calcul / Reca pour importer les données utiles du PR et recalculer ce qui doit l’être.\n"),
            ("liste", "⚠️ Si un fichier est verrouillé ou occupé, fermer Excel puis relancer l’action.\n"),

            ("section", "\n🟨 6. Encoder l’état d’avancement\n"),
            ("liste", "✅ Ouvrir le fichier État d’avancement depuis le bouton prévu.\n"),
            ("liste", "✅ Encoder uniquement les cellules gris clair prévues pour la période en cours.\n"),
            ("liste", "✅ Enregistrer le fichier avant de revenir dans Horizon Chantier.\n"),
            ("liste", "❌ Ne pas écrire dans une cellule calculée pour corriger un total manuellement.\n"),

            ("section", "\n✅ 7. Recalculer\n"),
            ("texte", "Après encodage, le recalcul permet de mettre à jour les colonnes et montants calculés.\n"),
            ("liste", "✅ Utiliser Calcul état pour recalculer l’état d’avancement.\n"),
            ("liste", "✅ Utiliser Clôturer état uniquement quand la période doit être clôturée et reportée dans le cumulé.\n"),
            ("liste", "⚠️ Si le calcul ne prend pas, vérifier que le fichier Excel est enregistré et, si nécessaire, fermé.\n"),

            ("section", "\n🟦 8. Pilotage chantier\n"),
            ("texte", "Le pilotage chantier génère une lecture synthétique du marché, du réalisé, de la production cumulée, du reste à facturer et de l’avancement.\n"),
            ("liste", "✅ À lancer uniquement après encodage correct et recalcul.\n"),
            ("liste", "🔍 Le PDF de pilotage est une sortie de lecture. La source de vérité reste Excel.\n"),

            ("section", "\n🟪 9. Pilotage délai\n"),
            ("texte", "Le pilotage délai sert à suivre le délai corrigé, le délai consommé, les jours restants et les écarts.\n"),
            ("liste", "✅ Utiliser ce bouton lorsque le fichier délai est à jour et enregistré.\n"),
            ("liste", "⚠️ Si le résultat paraît faux, vérifier d’abord les données du fichier délai.\n"),

            ("section", "\n✅ 10. Rendement\n"),
            ("texte", "Le bouton Rendements donne une lecture synthétique du rendement chantier.\n"),
            ("liste", "✅ Il permet d’évaluer les heures consommées, corrigées, restantes et les écarts utiles au suivi.\n"),
            ("liste", "🔍 Comme pour les autres documents, le résultat dépend entièrement de la qualité des données Excel de départ.\n"),

            ("section", "\n❌ 11. Erreurs fréquentes\n"),
            ("liste", "❌ Modifier une formule pour corriger rapidement un chiffre.\n"),
            ("liste", "❌ Encoder dans une cellule qui n’est pas gris clair.\n"),
            ("liste", "❌ Oublier d’enregistrer Excel avant une vérification, un import ou un calcul.\n"),
            ("liste", "❌ Lancer plusieurs fois une action sans contrôler les données d’entrée.\n"),
            ("liste", "❌ Travailler sur le mauvais chantier sélectionné dans la liste.\n"),
            ("liste", "❌ Croire que le PDF doit être corrigé manuellement alors que la vraie correction doit se faire dans Excel.\n"),

            ("section", "\n🆘 12. Quoi faire en cas de problème\n"),
            ("liste", "🆘 Arrêter les manipulations manuelles hasardeuses.\n"),
            ("liste", "🔍 Identifier le chantier concerné et le bouton utilisé.\n"),
            ("liste", "⚠️ Vérifier immédiatement si Excel est ouvert, non enregistré ou verrouillé.\n"),
            ("liste", "✅ Enregistrer les fichiers Excel puis fermer Excel si le logiciel doit reprendre la main.\n"),
            ("liste", "✅ Relancer une seule fois l’action.\n"),
            ("liste", "❌ Si le problème persiste, ne pas toucher aux formules pour forcer le résultat.\n"),
            ("liste", "🆘 Remonter le problème avec le nom du chantier, l’étape bloquante et le message vu à l’écran.\n"),

            ("section", "\n🎯 Rappel final\n"),
            ("ok", "✅ Bon réflexe: encoder dans les cellules gris clair, enregistrer Excel, puis utiliser Horizon Chantier.\n"),
            ("danger", "❌ Mauvais réflexe: corriger un total à la main dans une cellule calculée.\n"),
            ("aide", "🆘 En cas de doute: revenir dans cette aide avant de continuer.\n"),
            ("fin", "\nFin du mode d’emploi.\n"),
        ]

        contenu_depannage_pr = [
            ("titre", "PR (PRIX DE REVIENT) - FORMULES & DÉPANNAGE\n"),
            ("intro", "Cette aide rappelle les formules clés du PR et les points de contrôle à ne jamais casser.\n"),
            ("regle", "RÈGLE : Si la cellule n’est pas gris clair → TU NE TOUCHES PAS\n"),

            ("section", "\nMATIÈRE (P13)\n"),
            ("formule", "=SIERREUR(SOMME(P3:P11)/P12;0)\n"),

            ("section", "\nMAIN-D’ŒUVRE (P26)\n"),
            ("formule", "=SIERREUR(SOMME(P16:P24)/P25;0)\n"),

            ("section", "\nRÉCAPITULATIF\n"),
            ("formule", "C28 = P13\n"),
            ("formule", "C29 = P26\n"),
            ("formule", "C30 = C28 + C29\n"),

            ("section", "\nPRIX DE VENTE\n"),
            ("formule", "J28 = C30 * F28\n"),
            ("formule", "J29 = (C30 + J28) * F29\n"),
            ("formule", "F30 = C30 + J28 + J29\n"),
            ("formule", "F32 = F30 * (1 + F31)\n"),

            ("section", "\nLISTES DÉROULANTES\n"),
            ("texte", "Si la flèche disparaît :\n"),
            ("liste", "1. Données\n"),
            ("liste", "2. Validation des données\n"),
            ("liste", "3. Autoriser : Liste\n"),
            ("texte", "\nSource :\n"),
            ("formule", "Matière -> Bibliothèque_Matière!A2:A500\n"),
            ("formule", "Main d’œuvre -> Bibliothèque_MainOeuvre!A2:A500\n"),

            ("section", "\nIMPORTANT\n"),
            ("danger", "Ne jamais modifier les formules\n"),
            ("fin", "\nFin de l’aide PR.\n"),
        ]

        def afficher_aide(*_):
            texte.configure(state="normal")
            texte.delete("1.0", "end")

            if aide_var.get() == "Dépannage PR":
                contenu = contenu_depannage_pr
            else:
                contenu = contenu_mode_emploi

            for tag, bloc in contenu:
                texte.insert("end", bloc, tag)

            texte.configure(state="disabled")
            texte.yview_moveto(0)

        choix_aide.bind("<<ComboboxSelected>>", afficher_aide)
        afficher_aide()

        footer = ttk.Frame(win, padding=(14, 0, 14, 14))
        footer.pack(fill="x")
        ttk.Checkbutton(
            footer,
            text="Toujours visible",
            variable=topmost_var,
            command=appliquer_toujours_visible,
        ).pack(side="left")
        ttk.Button(footer, text="Fermer", command=fermer).pack(side="right")

    def ouvrir_aide_mode_emploi(self) -> None:
        self.ouvrir_aide()

    def calcul_etat_avancement(self):
        chemin = next(
            Path(
                dossier_chantiers() / self.tree.item(self.tree.selection()[0])["values"][0]
            ).glob("Etat_avancement*.xlsm")
        ).as_posix()

        feuille = "Bordereau"

        wb = load_workbook(chemin, keep_vba=True)
        ws = wb[feuille]

        def normaliser_libelle(valeur):
            texte = str(valeur or "").strip().lower()
            texte = texte.replace("é", "e").replace("è", "e").replace("ê", "e").replace("ë", "e")
            texte = texte.replace("à", "a").replace("â", "a").replace("ä", "a")
            texte = texte.replace("ù", "u").replace("û", "u").replace("ü", "u")
            texte = texte.replace("î", "i").replace("ï", "i")
            texte = texte.replace("ô", "o").replace("ö", "o")
            texte = texte.replace("ç", "c")
            texte = texte.replace("\n", " ")
            texte = re.sub(r"\s+", " ", texte)
            return texte

        def trouver_colonne_bordereau(alias):
            alias_normalises = {normaliser_libelle(a) for a in alias}
            max_ligne_entete = min(15, ws.max_row)
            for ligne in range(1, max_ligne_entete + 1):
                for col in range(1, ws.max_column + 1):
                    libelle = normaliser_libelle(ws.cell(ligne, col).value)
                    if libelle in alias_normalises or any(a in libelle for a in alias_normalises):
                        return ws.cell(ligne, col).column_letter
            return None

        col_quantite = trouver_colonne_bordereau({"quantite", "qte", "qte.", "qte :", "qte/", "qté"})
        col_prix_unitaire = trouver_colonne_bordereau({"prix unitaire", "pu", "p.u.", "prix unit"})

        if not col_quantite or not col_prix_unitaire:
            erreurs = []
            if not col_quantite:
                erreurs.append("colonne quantité introuvable (attendu : quantité, quantite, qté, qte)")
            if not col_prix_unitaire:
                erreurs.append("colonne prix unitaire introuvable (attendu : prix unitaire, pu, p.u., prix unit)")
            wb.close()
            messagebox.showerror("Calcul état", "Impossible de recalculer la feuille Bordereau.\n" + "\n".join(erreurs))
            return

        def nombre(valeur):
            if valeur in (None, "", "-", "—"):
                return 0
            try:
                return float(str(valeur).replace("€", "").replace(" ", "").replace(",", "."))
            except:
                return 0

        def ligne_bordereau_valide(ligne):
            c = ws[f"C{ligne}"].value
            j = ws[f"{col_quantite}{ligne}"].value
            l = ws[f"{col_prix_unitaire}{ligne}"].value
            p = ws[f"P{ligne}"].value
            q = ws[f"Q{ligne}"].value

            return any(v not in (None, "", "-", "—") for v in [c, j, l, p, q])

        def derniere_ligne_bordereau():
            derniere = 24
            lignes_vides = 0

            for ligne in range(24, ws.max_row + 1):
                if ligne_bordereau_valide(ligne):
                    derniere = ligne
                    lignes_vides = 0
                else:
                    lignes_vides += 1

                    # dès qu'on a plusieurs lignes vides d'affilée,
                    # on considère que le bordereau est terminé
                    if lignes_vides >= 5:
                        break

            return derniere

        fin = derniere_ligne_bordereau()

        for ligne in range(24, fin + 1):
            if not ligne_bordereau_valide(ligne):
                continue

            j = nombre(ws[f"{col_quantite}{ligne}"].value)
            l = nombre(ws[f"{col_prix_unitaire}{ligne}"].value)
            p = nombre(ws[f"P{ligne}"].value)
            q = nombre(ws[f"Q{ligne}"].value)

            r = p + q
            s = 0 if j == 0 else r / j
            t = r * l
            u = q
            v = l
            w = u * v

            ws[f"R{ligne}"] = r
            ws[f"S{ligne}"] = s
            ws[f"T{ligne}"] = t
            ws[f"U{ligne}"] = u
            ws[f"V{ligne}"] = v
            ws[f"W{ligne}"] = w

        

        wb.save(chemin)
        print("Etat recalculé")


       

    def mise_a_zero_etat(self):
        chemin = next(
            Path(
                dossier_chantiers() / self.tree.item(self.tree.selection()[0])["values"][0]
            ).glob("Etat_avancement*.xlsm")
        ).as_posix()

        feuille = "Bordereau"

        wb = load_workbook(chemin, keep_vba=True)
        ws = wb[feuille]

        ligne = 24

        while ligne <= ws.max_row:
            ws[f"P{ligne}"] = ws[f"R{ligne}"].value
            ws[f"Q{ligne}"] = 0
            ligne += 1

        wb.save(chemin)
        self.calcul_etat_avancement()
        print("Clôture état effectuée")

    def verifier_pr(self):
        p = self._chemin_chantier_selectionne()
        if not p:
            messagebox.showwarning("PR", "Sélectionne un chantier dans la liste.")
            return

        dossier_chantier = self._dossier_chantier_selectionne()
        chemin = dossier_chantier / "data" / "prix_de_revient.xlsx"

        import os
        from datetime import datetime

        if not chemin.exists():
            messagebox.showwarning(
                "Vérifier PR",
                "Le fichier prix_de_revient.xlsx est introuvable."
            )
            return

        timestamp = os.path.getmtime(chemin)
        date_modif = datetime.fromtimestamp(timestamp).strftime("%d/%m/%Y %H:%M:%S")

        messagebox.showinfo(
            "Date du PR",
            f"Dernier enregistrement du PR :\n{date_modif}\n\n"
            "⚠️ Si tu viens de modifier le PR :\n"
            "➡️ fais CMD + S dans Excel\n"
            "➡️ puis clique sur Vérifier PR"
        )
        
        wb = load_workbook(chemin, data_only=True)
        ws = wb["Chiffrage"]

        def nombre(v):
            if v in (None, "", "-", "—"):
                return 0
            try:
                return float(str(v).replace("€", "").replace(" ", "").replace(",", "."))
            except:
                return 0

        erreurs = []

        # MATIÈRE
        for ligne in range(3, 13):
            i = nombre(ws[f"I{ligne}"].value)
            j = nombre(ws[f"J{ligne}"].value)
            h = nombre(ws[f"H{ligne}"].value)

            attendu_k = i * (1 + j)
            attendu_l = h * attendu_k

            k = nombre(ws[f"K{ligne}"].value)
            l = nombre(ws[f"L{ligne}"].value)

            if abs(k - attendu_k) > 0.01:
                erreurs.append(f"Ligne {ligne} PU corrigé faux")

            if abs(l - attendu_l) > 0.01:
                erreurs.append(f"Ligne {ligne} Total matière faux")

        # MAIN-D’ŒUVRE
        for ligne in range(16, 26):
            c = nombre(ws[f"C{ligne}"].value)
            h = nombre(ws[f"H{ligne}"].value)
            j = nombre(ws[f"J{ligne}"].value)

            attendu_i = c * h
            attendu_k = attendu_i * j

            i = nombre(ws[f"I{ligne}"].value)
            k = nombre(ws[f"K{ligne}"].value)

            if abs(i - attendu_i) > 0.01:
                erreurs.append(f"Ligne {ligne} Heures fausses")

            if abs(k - attendu_k) > 0.01:
                erreurs.append(f"Ligne {ligne} Coût MO faux")

        if erreurs:
            messagebox.showwarning("PR", "\n".join(erreurs))
        else:
            messagebox.showinfo("PR", "✔ Tous les calculs sont corrects")

        # 🔒 Sécurisation automatique du PR
        wb = load_workbook(chemin)

        def est_cellule_gris_clair(cell):
            fill = cell.fill
            if fill is None or fill.fill_type != "solid":
                return False

            couleurs_gris_clair = {
                "FFD9D9D9", "00D9D9D9",
                "FFDDDDDD", "00DDDDDD",
                "FFEDEDED", "00EDEDED",
                "FFF2F2F2", "00F2F2F2",
                "FFF3F3F3", "00F3F3F3",
            }

            for couleur in (fill.fgColor, fill.start_color):
                if getattr(couleur, "type", None) == "rgb":
                    rgb = (getattr(couleur, "rgb", None) or "").upper()
                    if rgb in couleurs_gris_clair:
                        return True
            return False

        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.__class__.__name__ == "MergedCell":
                        continue

                    est_formule = (
                        cell.value is not None
                        and isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    )
                    est_gris_clair = est_cellule_gris_clair(cell)

                    cell.protection = cell.protection.copy(
                        locked=not est_gris_clair or est_formule
                    )

            ws.protection.sheet = True
            ws.protection.password = "1234"

        wb.save(chemin)


            
    def pilotage_chantier(self):
        import os
        import subprocess
        import tempfile
        from datetime import datetime
        import matplotlib.pyplot as plt
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_RIGHT, TA_CENTER

        chemin = next(
            Path(
                dossier_chantiers() / self.tree.item(self.tree.selection()[0])["values"][0]
            ).glob("Etat_avancement*.xlsm")
        ).as_posix()

        wb = load_workbook(chemin, data_only=True)
        ws = wb["Bordereau"]
        

        def nombre(val):
            if val in (None, "", "-", "—"):
                return 0
            try:
                return float(str(val).replace("€", "").replace(" ", "").replace(",", "."))
            except:
                return 0

        def texte(val):
            return str(val or "").strip().lower()

        def trouver_ligne_libelle(mots):
            for row in range(ws.max_row, 1, -1):
                for col in range(1, ws.max_column + 1):
                    val = texte(ws.cell(row=row, column=col).value)
                    if not val:
                        continue
                    if all(mot in val for mot in mots):
                        return row, col
            return None, None

        def premiere_valeur_a_droite(row, col_depart, max_ecart=3):
            if not row or not col_depart:
                return 0

            col_fin = min(col_depart + max_ecart, ws.max_column)

            for col in range(col_depart + 1, col_fin + 1):
                val = ws.cell(row=row, column=col).value

                # Ignore les cellules vides ou décoratives
                if val in (None, "", "-", "—"):
                    continue

                # Si Excel renvoie déjà un nombre, on le prend tel quel
                if isinstance(val, (int, float)):
                    return float(val)

                # Sinon on passe par le parseur existant
                num = nombre(val)
                txt = str(val).strip()

                # Conserve aussi un vrai zéro explicite
                if num != 0 or txt in {"0", "0,0", "0.0", "0,00", "0.00"}:
                    return num

            return 0

        def valeur_soumission_publique(libelles):
            for libelle in libelles:
                row, col = trouver_ligne_libelle([libelle])
                if row and col:
                    return premiere_valeur_a_droite(row, col), row, col
            return 0, None, None

        nom_chantier = self.tree.item(self.tree.selection()[0])['values'][0]
        fichier_prive = (
            "Avenants" in wb.sheetnames
            and "Revision_global" in wb.sheetnames
            and "Révision" in wb.sheetnames
        )

        if fichier_prive:
            ws_rev = wb["Révision"]
            ws_glob = wb["Revision_global"]
            ws_avenants = wb["Avenants"]

            revision = nombre(ws_rev["F32"].value)
            revision_cumulee = nombre(ws_glob["D62"].value)

            avenants_plus = nombre(ws_avenants["D62"].value)
            avenants_moins = nombre(ws_avenants["E62"].value)

            ligne_soumission, col_soumission = trouver_ligne_libelle(["soumission", "tva"])
            ligne_mois, col_mois = trouver_ligne_libelle(["total", "mois", "tva"])
            ligne_execute, col_execute = trouver_ligne_libelle(["exécuté"])

            if not ligne_execute:
                ligne_execute, col_execute = trouver_ligne_libelle(["execute"])

            montant_soumission = premiere_valeur_a_droite(ligne_soumission, col_soumission)
            total_mois = premiere_valeur_a_droite(ligne_mois, col_mois)
            total_realise = premiere_valeur_a_droite(ligne_execute, col_execute)
            total_marche = nombre(ws_avenants["E65"].value)
        else:
            montant_soumission, ligne_soumission, col_soumission = valeur_soumission_publique(
                ["total soumission hors tva"]
            )
            _total_etat_cumule, _ligne_etat_cumule, _col_etat_cumule = valeur_soumission_publique(
                ["total état cumulé hors tva", "total etat cumulé hors tva"]
            )
            total_mois, ligne_mois, col_mois = valeur_soumission_publique(
                ["total du mois hors tva", "total du mois hors tva"]
            )
            avenants_plus, _ligne_avenants, _col_avenants = valeur_soumission_publique(
                ["total des avenants cumulé", "total des avenants cumule"]
            )
            total_realise, ligne_execute, col_execute = valeur_soumission_publique(
                ["total exécuté", "total execute"]
            )
            revision_cumulee, _ligne_revision, _col_revision = valeur_soumission_publique(
                ["total des révision cumulé", "total des revision cumule"]
            )
            _montant_global, _ligne_global, _col_global = valeur_soumission_publique(
                ["montant global à facturer", "montant global a facturer"]
            )

            revision = 0
            avenants_moins = 0
            total_marche = montant_soumission + avenants_plus - avenants_moins

        print("revision =", revision)
        print("revision_cumulee =", revision_cumulee)
        print("avenants_plus =", avenants_plus)
        print("avenants_moins =", avenants_moins)

        print("ligne_mois =", ligne_mois, "col_mois =", col_mois)
        print("total_mois =", total_mois)
       

    
        date_du_jour = datetime.now().strftime("%d/%m/%Y")
        realise_mois = total_mois
        realise_cumule = total_realise
        production_mois = realise_mois + revision
        production_cumulee = realise_cumule + revision_cumulee
        reste_a_facturer = total_marche - production_cumulee
        avancement = 0 if total_marche == 0 else (production_cumulee / total_marche) * 100

        texte_popup = f"""PILOTAGE CHANTIER

       
    Chantier : {nom_chantier}
    Date     : {date_du_jour}

    Montant soumission              : {montant_soumission:,.2f} €
    Avenants cumulés en plus        : {avenants_plus:,.2f} €
    Avenants cumulés en moins       : {avenants_moins:,.2f} €
    Marché total                    : {total_marche:,.2f} €
    Réalisé du mois                 : {realise_mois:,.2f} €
    Révision du mois                : {revision:,.2f} €
    Production du mois              : {production_mois:,.2f} €
    Réalisé cumulé                  : {realise_cumule:,.2f} €
    Révision cumulée                : {revision_cumulee:,.2f} €
    Production cumulée              : {production_cumulee:,.2f} €
    Reste à facturer                : {reste_a_facturer:,.2f} €
    Avancement                      : {avancement:.1f} %
    """

        pdf_path = str(Path(chemin).with_name("pilotage_chantier.pdf"))
        wb.close()

        doc = SimpleDocTemplate(
            pdf_path,
            pagesize=landscape(A4),
            leftMargin=8 * mm,
            rightMargin=8 * mm,
            topMargin=5 * mm,
            bottomMargin=5 * mm,
        )

        elements = []
        graphique_tmp = None
        logo_path = str(dossier_base() / "logo_jt_bati.png")

        style_titre = ParagraphStyle(
            "PilotageTitre",
            fontName="Helvetica-Bold",
            fontSize=18,
            leading=18,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#163A59"),
        )
        style_meta_label = ParagraphStyle(
            "PilotageMetaLabel",
            fontName="Helvetica-Bold",
            fontSize=8.5,
            leading=9,
            textColor=colors.HexColor("#36556F"),
        )
        style_meta_valeur = ParagraphStyle(
            "PilotageMetaValeur",
            fontName="Helvetica",
            fontSize=8.8,
            leading=9.2,
            textColor=colors.black,
        )
        style_bloc_titre = ParagraphStyle(
            "PilotageBlocTitre",
            fontName="Helvetica-Bold",
            fontSize=9.6,
            leading=10,
            alignment=TA_CENTER,
            textColor=colors.white,
        )
        style_libelle = ParagraphStyle(
            "PilotageLibelle",
            fontName="Helvetica",
            fontSize=8.6,
            leading=9,
            textColor=colors.black,
        )
        style_valeur = ParagraphStyle(
            "PilotageValeur",
            fontName="Helvetica-Bold",
            fontSize=8.8,
            leading=9,
            alignment=TA_RIGHT,
            textColor=colors.black,
        )
        style_carte_label = ParagraphStyle(
            "PilotageCarteLabel",
            fontName="Helvetica-Bold",
            fontSize=7.6,
            leading=8,
            textColor=colors.HexColor("#5A7086"),
        )
        style_carte_valeur = ParagraphStyle(
            "PilotageCarteValeur",
            fontName="Helvetica-Bold",
            fontSize=11.2,
            leading=11.4,
            alignment=TA_RIGHT,
            textColor=colors.HexColor("#173C5A"),
        )
        style_carte_valeur_finale = ParagraphStyle(
            "PilotageCarteValeurFinale",
            fontName="Helvetica-Bold",
            fontSize=12.2,
            leading=12.2,
            alignment=TA_RIGHT,
            textColor=colors.HexColor("#2E7D32"),
        )
        style_bandeau = ParagraphStyle(
            "PilotageBandeau",
            fontName="Helvetica-Bold",
            fontSize=9.3,
            leading=9.8,
            textColor=colors.HexColor("#173C5A"),
        )
        style_avancement_titre = ParagraphStyle(
            "PilotageAvancementTitre",
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=8.2,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#5A7086"),
        )
        style_avancement_valeur = ParagraphStyle(
            "PilotageAvancementValeur",
            fontName="Helvetica-Bold",
            fontSize=10.5,
            leading=10.8,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#163A59"),
        )

        logo_cell = Spacer(1, 1)
        if os.path.exists(logo_path):
            try:
                logo_cell = Image(logo_path, width=27 * mm, height=10.8 * mm)
                logo_cell.hAlign = "LEFT"
            except:
                logo_cell = Spacer(1, 1)

        meta_table = Table([
            [
                Paragraph("CHANTIER", style_meta_label),
                Paragraph(nom_chantier, style_meta_valeur),
                Paragraph("DATE", style_meta_label),
                Paragraph(date_du_jour, style_meta_valeur),
            ]
        ], colWidths=[18 * mm, 47 * mm, 12 * mm, 22 * mm])
        meta_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))

        header_table = Table([
            [logo_cell, Paragraph("PILOTAGE CHANTIER", style_titre), meta_table]
        ], colWidths=[34 * mm, 158 * mm, 74 * mm])
        header_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F4F8FB")),
            ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#D9E4EE")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (1, 0), (1, 0), "CENTER"),
            ("ALIGN", (2, 0), (2, 0), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 3))

        valeurs_graphique = [total_marche, production_cumulee, reste_a_facturer]
        valeurs_plot = [max(total_marche, 0), max(production_cumulee, 0), max(reste_a_facturer, 0)]
        etiquettes_graphique = ["Marché\ntotal", "Production\ncumulée", "Reste à\nfacturer"]
        couleurs_graphique = ["#1F4E79", "#E69138", "#70AD47"]

        fig, ax = plt.subplots(figsize=(3.1, 3.7), facecolor="white")
        barres = ax.bar(
            range(3),
            valeurs_plot,
            width=0.55,
            color=couleurs_graphique,
            edgecolor="#FFFFFF",
            linewidth=0.8,
            zorder=3,
        )

        max_valeur = max([abs(v) for v in valeurs_plot] + [1])
        marge = max_valeur * 0.18
        ax.set_ylim(0, max(max(valeurs_plot), 1) + marge)
        ax.set_xticks(range(3), etiquettes_graphique)
        ax.tick_params(axis="x", labelsize=7.8, colors="#173C5A", length=0, pad=6)
        ax.tick_params(axis="y", labelsize=7.5, colors="#6A7E90")
        ax.grid(axis="y", color="#D9E4EE", linewidth=0.8, alpha=0.8, zorder=0)
        ax.set_axisbelow(True)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.spines["left"].set_color("#D9E4EE")
        ax.spines["bottom"].set_color("#D9E4EE")

        for barre, valeur in zip(barres, valeurs_graphique):
            ax.text(
                barre.get_x() + barre.get_width() / 2,
                barre.get_height() + max(marge * 0.08, 0.6),
                f"{valeur:,.2f} €",
                ha="center",
                va="bottom",
                fontsize=7.8,
                fontweight="bold",
                color="#173C5A",
            )

        with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
            graphique_tmp = tmp.name

        plt.tight_layout()
        plt.savefig(graphique_tmp, dpi=220, bbox_inches="tight", facecolor="white")
        plt.close(fig)

        graphique = Image(graphique_tmp, width=66 * mm, height=66 * mm)
        graphique.hAlign = "CENTER"

        largeur_blocs_gauche = [95 * mm, 30 * mm]
        largeur_blocs_droite = [38 * mm, 24 * mm]

        bloc_1 = Table([
            ["CONSTRUCTION", ""],
            [Paragraph("Montant soumission", style_libelle), Paragraph(f"{montant_soumission:,.2f} €", style_valeur)],
            [Paragraph("+ Avenants cumulés en plus", style_libelle), Paragraph(f"{avenants_plus:,.2f} €", style_valeur)],
            [Paragraph("- Avenants cumulés en moins", style_libelle), Paragraph(f"{avenants_moins:,.2f} €", style_valeur)],
            [Paragraph("= Marché total", style_libelle), Paragraph(f"{total_marche:,.2f} €", style_valeur)],
        ], colWidths=largeur_blocs_gauche)
        bloc_1.hAlign = "CENTER"
        bloc_1.setStyle(TableStyle([
            ("SPAN", (0, 0), (1, 0)),
            ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
            ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (1, 0), 12),
            ("BACKGROUND", (0, 1), (1, -1), colors.white),
            ("ROWBACKGROUNDS", (0, 1), (1, -1), [colors.white, colors.HexColor("#F8FBFF")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B7C5D3")),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
            ("BACKGROUND", (0, 4), (1, 4), colors.HexColor("#FFF2CC")),
            ("FONTNAME", (0, 4), (1, 4), "Helvetica-Bold"),
        ]))

        bloc_2 = Table([
            ["RÉALISÉ", ""],
            [Paragraph("Réalisé du mois", style_libelle), Paragraph(f"{realise_mois:,.2f} €", style_valeur)],
            [Paragraph("+ Révision du mois", style_libelle), Paragraph(f"{revision:,.2f} €", style_valeur)],
            [Paragraph("= Production du mois", style_libelle), Paragraph(f"{production_mois:,.2f} €", style_valeur)],
            [Paragraph("Réalisé cumulé", style_libelle), Paragraph(f"{realise_cumule:,.2f} €", style_valeur)],
            [Paragraph("+ Révision cumulée", style_libelle), Paragraph(f"{revision_cumulee:,.2f} €", style_valeur)],
            [Paragraph("= Production cumulée", style_libelle), Paragraph(f"{production_cumulee:,.2f} €", style_valeur)],
        ], colWidths=largeur_blocs_gauche)
        bloc_2.hAlign = "CENTER"
        bloc_2.setStyle(TableStyle([
            ("SPAN", (0, 0), (1, 0)),
            ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#4F81BD")),
            ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
            ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (1, 0), 12),
            ("BACKGROUND", (0, 1), (1, -1), colors.white),
            ("ROWBACKGROUNDS", (0, 1), (1, -1), [colors.white, colors.HexColor("#F7FAFD")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B7C5D3")),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
            ("BACKGROUND", (0, 3), (1, 3), colors.HexColor("#FFF2CC")),
            ("FONTNAME", (0, 3), (1, 3), "Helvetica-Bold"),
            ("BACKGROUND", (0, 6), (1, 6), colors.HexColor("#FFF2CC")),
            ("FONTNAME", (0, 6), (1, 6), "Helvetica-Bold"),
        ]))

        bloc_3 = Table([
            ["RÉSULTAT", ""],
            [Paragraph("Reste à facturer", style_libelle), Paragraph(f"{reste_a_facturer:,.2f} €", style_valeur)],
            [Paragraph("Avancement", style_libelle), Paragraph(f"{avancement:.1f} %", style_valeur)],
            [Paragraph("Écart financier final", style_libelle), Paragraph(f"{reste_a_facturer:,.2f} €", style_valeur)],
        ], colWidths=largeur_blocs_gauche)
        bloc_3.hAlign = "CENTER"
        bloc_3.setStyle(TableStyle([
            ("SPAN", (0, 0), (1, 0)),
            ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#C0504D")),
            ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
            ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (1, 0), 12),
            ("BACKGROUND", (0, 1), (1, -1), colors.white),
            ("ROWBACKGROUNDS", (0, 1), (1, -1), [colors.white, colors.HexColor("#FCF7F7")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#C9B1B1")),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
            ("BACKGROUND", (0, 1), (1, 1), colors.HexColor("#D9EAD3")),
            ("FONTNAME", (0, 1), (1, 1), "Helvetica-Bold"),
            ("BACKGROUND", (0, 3), (1, 3), colors.HexColor("#D9EAD3")),
            ("FONTNAME", (0, 3), (1, 3), "Helvetica-Bold"),
        ]))

        colonne_gauche = Table([
            [bloc_1],
            [bloc_2],
            [bloc_3],
        ], colWidths=[125 * mm])
        colonne_gauche.setStyle(TableStyle([
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))

        bloc_graphique = Table([
            [Paragraph("RÉSUMÉ VISUEL", style_bloc_titre)],
            [graphique],
        ], colWidths=[70 * mm])
        bloc_graphique.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#173C5A")),
            ("BACKGROUND", (0, 1), (0, 1), colors.white),
            ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#B7C5D3")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ]))

        indicateurs_table = Table([
            [Paragraph("Marché total", style_carte_label), Paragraph(f"{total_marche:,.2f} €", style_carte_valeur)],
            [Paragraph("Production du mois", style_carte_label), Paragraph(f"{production_mois:,.2f} €", style_carte_valeur)],
            [Paragraph("Production cumulée", style_carte_label), Paragraph(f"{production_cumulee:,.2f} €", style_carte_valeur)],
            [Paragraph("Avancement %", style_carte_label), Paragraph(f"{avancement:.1f} %", style_carte_valeur)],
            [Paragraph("Reste à facturer", style_carte_label), Paragraph(f"{reste_a_facturer:,.2f} €", style_carte_valeur_finale)],
            [Paragraph("Écart financier final", style_carte_label), Paragraph(f"{reste_a_facturer:,.2f} €", style_carte_valeur_finale)],
        ], colWidths=largeur_blocs_droite)
        indicateurs_table.setStyle(TableStyle([
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#F6F9FC")]),
            ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#D6E0EA")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
            ("BACKGROUND", (0, 4), (1, 4), colors.HexColor("#D9EAD3")),
            ("BACKGROUND", (0, 5), (1, 5), colors.HexColor("#D9EAD3")),
        ]))

        indicateurs_cles = Table([
            [Paragraph("INDICATEURS CLÉS", style_bloc_titre)],
            [indicateurs_table],
        ], colWidths=[70 * mm])
        indicateurs_cles.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#365F91")),
            ("BACKGROUND", (0, 1), (0, 1), colors.white),
            ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#B7C5D3")),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ]))

        colonne_droite = Table([
            [bloc_graphique],
            [indicateurs_cles],
        ], colWidths=[70 * mm])
        colonne_droite.setStyle(TableStyle([
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))

        corps = Table([
            [colonne_gauche, colonne_droite]
        ], colWidths=[128 * mm, 70 * mm])
        corps.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        elements.append(corps)
        elements.append(Spacer(1, 2))

        footer_row_height = 14.5 * mm
        footer_split_height = 5.2 * mm

        avancement_table = Table([
            [Paragraph("AVANCEMENT", style_avancement_titre)],
            [Paragraph(f"{avancement:.1f} %", style_avancement_valeur)],
        ], colWidths=[30 * mm], rowHeights=[footer_split_height, footer_row_height - footer_split_height])
        avancement_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EAF1F8")),
            ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#B7C5D3")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))

        footer_explication = Table([
            [
                Paragraph(
                    f"Reste à facturer = Marché total - Production cumulée<br/>{total_marche:,.2f} € - {production_cumulee:,.2f} € = {reste_a_facturer:,.2f} €",
                    style_bandeau,
                )
            ]
        ], colWidths=[168 * mm], rowHeights=[footer_row_height])
        footer_explication.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#FFF2CC")),
            ("BOX", (0, 0), (0, 0), 0.7, colors.HexColor("#D6B656")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))

        footer_table = Table([
            [footer_explication, avancement_table]
        ], colWidths=[168 * mm, 30 * mm], rowHeights=[footer_row_height])
        footer_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        elements.append(footer_table)

        doc.build(elements)

        if graphique_tmp and os.path.exists(graphique_tmp):
            try:
                os.unlink(graphique_tmp)
            except:
                pass

        subprocess.run(["open", pdf_path])

        
        messagebox.showinfo("Pilotage chantier", texte_popup)
   
   
    def pilotage_delai(self):
        import os
        import subprocess
        import tempfile
        from datetime import datetime
        import matplotlib.pyplot as plt
        from openpyxl import load_workbook
        from tkinter import messagebox
        from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Image, Paragraph
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.enums import TA_RIGHT, TA_CENTER

        def nombre(valeur):
            try:
                if valeur is None:
                    return 0
                if isinstance(valeur, str):
                    valeur = valeur.replace("%", "").replace("€", "").replace(" ", "").replace(",", ".").strip()
                    if valeur in ("", "-", "—"):
                        return 0
                return float(valeur)
            except:
                return 0

        try:
            if not self.tree.selection():
                messagebox.showwarning("Attention", "Veuillez sélectionner un chantier.")
                return

            nom_chantier = self.tree.item(self.tree.selection()[0])["values"][0]
            dossier_chantier = dossier_chantiers() / nom_chantier
            fichier_delai = os.path.join(dossier_chantier, "Delai.xlsx")

            if not os.path.exists(fichier_delai):
                messagebox.showerror("Erreur", f"Fichier introuvable :\n{fichier_delai}")
                return

            wb = load_workbook(fichier_delai, data_only=True)
            ws = wb.active

            def lire_valeur_delai(libelles):
                if isinstance(libelles, str):
                    libelles = [libelles]
                for row in range(1, ws.max_row + 1):
                    libelle = str(ws[f"M{row}"].value or "").strip().lower()
                    for recherche in libelles:
                        if recherche in libelle:
                            return nombre(ws[f"O{row}"].value)
                return 0

            # Bloc haut : délai corrigé
            delai_soumission_jour = lire_valeur_delai("délai soumission jour")
            jours_avenant_plus = lire_valeur_delai("jour accordes avenant en plus")
            jours_attente_technique = lire_valeur_delai("jour accordes dû à une attente technique")
            jours_imprevu = lire_valeur_delai("jour accordes dû à imprévu, prévisible non visible")
            delai_corrige_jour = lire_valeur_delai("délai corrigé en jour")

            # Bloc bas : délai consommé
            jours_consommes = lire_valeur_delai(["jours actif consommé", "jours consommé"])
            jours_avenant_moins = lire_valeur_delai("jour accordes avenant en moins")
            jours_intemperies = lire_valeur_delai("jour intempéries")
            conges = lire_valeur_delai(["congé/ferrier/compensatoire jour", "congé/ferrier/compensatoire"])
            delai_consomme = lire_valeur_delai(["délai consommé jour", "délai consommé"])

            # Résultat
            jours_restants = lire_valeur_delai("jours restant pour exécution")
            depassement = lire_valeur_delai("ecart en jour en moins")
            ecart_pct = lire_valeur_delai("ecart en %")

            if ecart_pct <= 1 and ecart_pct not in (0,):
                ecart_pct = ecart_pct * 100

            texte = (
                f"Chantier : {nom_chantier}\n\n"

                f"DÉLAI CORRIGÉ\n"
                f"Délai soumission jour : {delai_soumission_jour:.2f} j\n"
                f"Jour accordés avenant en plus : {jours_avenant_plus:.2f} j\n"
                f"Jour accordés dû à une attente Technique : {jours_attente_technique:.2f} j\n"
                f"Jour accordés dû à imprévu, prévisible non visible : {jours_imprevu:.2f} j\n"
                f"Délai corrigé en jour : {delai_corrige_jour:.2f} j\n\n"

                f"DÉLAI CONSOMMÉ\n"
                f"Jours actif consommé : {jours_consommes:.2f} j\n"
                f"Jour accordés avenant en moins : {jours_avenant_moins:.2f} j\n"
                f"Jour intempéries : {jours_intemperies:.2f} j\n"
                f"Congé/ferrier/compensatoire : {conges:.2f} j\n"
                f"Délai consommé : {delai_consomme:.2f} j\n\n"

                f"RÉSULTAT\n"
                f"Jours restant pour exécution : {jours_restants:.2f} j\n"
                f"Ecart en jour en moins : {depassement:.2f} j\n"
                f"Ecart en % : {ecart_pct:.1f} %"
            )

            # Popup
            messagebox.showinfo("📊 Pilotage délai", texte)

            # PDF
            wb.close()
            pdf_path = os.path.join(dossier_chantier, "pilotage_delai.pdf")
            doc = SimpleDocTemplate(
                pdf_path,
                pagesize=landscape(A4),
                leftMargin=10 * mm,
                rightMargin=10 * mm,
                topMargin=8 * mm,
                bottomMargin=8 * mm,
            )

            elements = []
            graphique_tmp = None
            logo_path = str(dossier_base() / "logo_jt_bati.png")
            date_du_jour = datetime.now().strftime("%d/%m/%Y")

            style_titre = ParagraphStyle(
                "DelaiTitre",
                fontName="Helvetica-Bold",
                fontSize=18,
                leading=20,
                alignment=TA_CENTER,
                textColor=colors.HexColor("#163A59"),
            )
            style_meta_label = ParagraphStyle(
                "DelaiMetaLabel",
                fontName="Helvetica-Bold",
                fontSize=8.5,
                leading=10,
                textColor=colors.HexColor("#36556F"),
            )
            style_meta_valeur = ParagraphStyle(
                "DelaiMetaValeur",
                fontName="Helvetica",
                fontSize=8.8,
                leading=10.5,
                textColor=colors.black,
            )
            style_bloc_titre = ParagraphStyle(
                "DelaiBlocTitre",
                fontName="Helvetica-Bold",
                fontSize=9.4,
                leading=10.5,
                alignment=TA_CENTER,
                textColor=colors.white,
            )
            style_libelle = ParagraphStyle(
                "DelaiLibelle",
                fontName="Helvetica",
                fontSize=7.8,
                leading=8.8,
                textColor=colors.black,
            )
            style_valeur = ParagraphStyle(
                "DelaiValeur",
                fontName="Helvetica-Bold",
                fontSize=8.1,
                leading=9.2,
                alignment=TA_RIGHT,
                textColor=colors.black,
            )
            style_carte_label = ParagraphStyle(
                "DelaiCarteLabel",
                fontName="Helvetica-Bold",
                fontSize=7.4,
                leading=8.4,
                textColor=colors.HexColor("#5A7086"),
            )
            style_carte_valeur = ParagraphStyle(
                "DelaiCarteValeur",
                fontName="Helvetica-Bold",
                fontSize=10.8,
                leading=12,
                alignment=TA_RIGHT,
                textColor=colors.HexColor("#173C5A"),
            )
            style_carte_valeur_finale = ParagraphStyle(
                "DelaiCarteValeurFinale",
                fontName="Helvetica-Bold",
                fontSize=11.6,
                leading=12.8,
                alignment=TA_RIGHT,
                textColor=colors.HexColor("#9C3B32"),
            )
            style_bandeau = ParagraphStyle(
                "DelaiBandeau",
                fontName="Helvetica-Bold",
                fontSize=9.1,
                leading=10.5,
                textColor=colors.HexColor("#173C5A"),
            )
            style_ecart_titre = ParagraphStyle(
                "DelaiEcartTitre",
                fontName="Helvetica-Bold",
                fontSize=8,
                leading=9,
                alignment=TA_CENTER,
                textColor=colors.HexColor("#5A7086"),
            )
            style_ecart_valeur = ParagraphStyle(
                "DelaiEcartValeur",
                fontName="Helvetica-Bold",
                fontSize=10.4,
                leading=12,
                alignment=TA_CENTER,
                textColor=colors.HexColor("#163A59"),
            )

            logo_cell = Spacer(1, 1)
            if os.path.exists(logo_path):
                try:
                    logo_cell = Image(logo_path, width=30 * mm, height=12 * mm)
                    logo_cell.hAlign = "LEFT"
                except:
                    logo_cell = Spacer(1, 1)

            meta_table = Table([
                [
                    Paragraph("CHANTIER", style_meta_label),
                    Paragraph(nom_chantier, style_meta_valeur),
                    Paragraph("DATE", style_meta_label),
                    Paragraph(date_du_jour, style_meta_valeur),
                ]
            ], colWidths=[18 * mm, 47 * mm, 12 * mm, 22 * mm])
            meta_table.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]))

            header_table = Table([
                [logo_cell, Paragraph("PILOTAGE DÉLAI", style_titre), meta_table]
            ], colWidths=[34 * mm, 158 * mm, 74 * mm])
            header_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F4F8FB")),
                ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#D9E4EE")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (1, 0), (1, 0), "CENTER"),
                ("ALIGN", (2, 0), (2, 0), "RIGHT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]))
            elements.append(header_table)
            elements.append(Spacer(1, 4))

            valeurs_graphique = [delai_corrige_jour, delai_consomme, jours_restants]
            valeurs_plot = [max(delai_corrige_jour, 0), max(delai_consomme, 0), max(jours_restants, 0)]
            etiquettes_graphique = ["Délai\ncorrigé", "Délai\nconsommé", "Jours\nrestants"]
            couleurs_graphique = ["#1F4E79", "#E69138", "#70AD47"]

            fig, ax = plt.subplots(figsize=(3.2, 4.0), facecolor="white")
            barres = ax.bar(
                range(3),
                valeurs_plot,
                width=0.55,
                color=couleurs_graphique,
                edgecolor="#FFFFFF",
                linewidth=0.8,
                zorder=3,
            )

            max_valeur = max([abs(v) for v in valeurs_plot] + [1])
            marge = max_valeur * 0.18
            ax.set_ylim(0, max(max(valeurs_plot), 1) + marge)
            ax.set_xticks(range(3), etiquettes_graphique)
            ax.tick_params(axis="x", labelsize=7.6, colors="#173C5A", length=0, pad=6)
            ax.tick_params(axis="y", labelsize=7.2, colors="#6A7E90")
            ax.grid(axis="y", color="#D9E4EE", linewidth=0.8, alpha=0.8, zorder=0)
            ax.set_axisbelow(True)
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            ax.spines["left"].set_color("#D9E4EE")
            ax.spines["bottom"].set_color("#D9E4EE")

            for barre, valeur in zip(barres, valeurs_graphique):
                ax.text(
                    barre.get_x() + barre.get_width() / 2,
                    barre.get_height() + max(marge * 0.08, 0.3),
                    f"{valeur:.2f} j",
                    ha="center",
                    va="bottom",
                    fontsize=7.6,
                    fontweight="bold",
                    color="#173C5A",
                )

            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
                graphique_tmp = tmp.name

            plt.tight_layout()
            plt.savefig(graphique_tmp, dpi=220, bbox_inches="tight", facecolor="white")
            plt.close(fig)

            graphique = Image(graphique_tmp, width=72 * mm, height=72 * mm)
            graphique.hAlign = "CENTER"

            largeur_blocs_gauche = [96 * mm, 29 * mm]
            largeur_blocs_droite = [38 * mm, 24 * mm]

            bloc_1 = Table([
                ["CONSTRUCTION", ""],
                [Paragraph("Délai soumission jour", style_libelle), Paragraph(f"{delai_soumission_jour:.2f} j", style_valeur)],
                [Paragraph("+ Jour accordés avenant en plus", style_libelle), Paragraph(f"{jours_avenant_plus:.2f} j", style_valeur)],
                [Paragraph("+ Jour accordés dû à une attente Technique", style_libelle), Paragraph(f"{jours_attente_technique:.2f} j", style_valeur)],
                [Paragraph("+ Jour accordés dû à imprévu, prévisible non visible", style_libelle), Paragraph(f"{jours_imprevu:.2f} j", style_valeur)],
                [Paragraph("= Délai corrigé en jour", style_libelle), Paragraph(f"{delai_corrige_jour:.2f} j", style_valeur)],
            ], colWidths=largeur_blocs_gauche)
            bloc_1.hAlign = "CENTER"
            bloc_1.setStyle(TableStyle([
                ("SPAN", (0, 0), (1, 0)),
                ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#1F4E79")),
                ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
                ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (1, 0), 11),
                ("BACKGROUND", (0, 1), (1, -1), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (1, -1), [colors.white, colors.HexColor("#F8FBFF")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B7C5D3")),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 2.5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
                ("BACKGROUND", (0, 5), (1, 5), colors.HexColor("#FFF2CC")),
                ("FONTNAME", (0, 5), (1, 5), "Helvetica-Bold"),
            ]))

            bloc_2 = Table([
                ["CONSOMMÉ", ""],
                [Paragraph("Jours actif consommé", style_libelle), Paragraph(f"{jours_consommes:.2f} j", style_valeur)],
                [Paragraph("- Jour accordés avenant en moins", style_libelle), Paragraph(f"{jours_avenant_moins:.2f} j", style_valeur)],
                [Paragraph("- Jour intempéries", style_libelle), Paragraph(f"{jours_intemperies:.2f} j", style_valeur)],
                [Paragraph("- Congé/ferrier/compensatoire", style_libelle), Paragraph(f"{conges:.2f} j", style_valeur)],
                [Paragraph("= Délai consommé", style_libelle), Paragraph(f"{delai_consomme:.2f} j", style_valeur)],
            ], colWidths=largeur_blocs_gauche)
            bloc_2.hAlign = "CENTER"
            bloc_2.setStyle(TableStyle([
                ("SPAN", (0, 0), (1, 0)),
                ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#4F81BD")),
                ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
                ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (1, 0), 11),
                ("BACKGROUND", (0, 1), (1, -1), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (1, -1), [colors.white, colors.HexColor("#F7FAFD")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B7C5D3")),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 2.5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
                ("BACKGROUND", (0, 5), (1, 5), colors.HexColor("#FFF2CC")),
                ("FONTNAME", (0, 5), (1, 5), "Helvetica-Bold"),
            ]))

            bloc_3 = Table([
                ["RÉSULTAT", ""],
                [Paragraph("Jours restant pour exécution", style_libelle), Paragraph(f"{jours_restants:.2f} j", style_valeur)],
                [Paragraph("Ecart en jour en moins", style_libelle), Paragraph(f"{depassement:.2f} j", style_valeur)],
                [Paragraph("Ecart en %", style_libelle), Paragraph(f"{ecart_pct:.1f} %", style_valeur)],
            ], colWidths=largeur_blocs_gauche)
            bloc_3.hAlign = "CENTER"
            bloc_3.setStyle(TableStyle([
                ("SPAN", (0, 0), (1, 0)),
                ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#C0504D")),
                ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
                ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (1, 0), 11),
                ("BACKGROUND", (0, 1), (1, -1), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (1, -1), [colors.white, colors.HexColor("#FCF7F7")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#C9B1B1")),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 2.5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
                ("BACKGROUND", (0, 1), (1, 1), colors.HexColor("#E7E7E7")),
                ("BACKGROUND", (0, 3), (1, 3), colors.HexColor("#D9EAD3")),
                ("FONTNAME", (0, 3), (1, 3), "Helvetica-Bold"),
            ]))

            colonne_gauche = Table([
                [bloc_1],
                [bloc_2],
                [bloc_3],
            ], colWidths=[125 * mm])
            colonne_gauche.setStyle(TableStyle([
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))

            bloc_graphique = Table([
                [Paragraph("RÉSUMÉ VISUEL", style_bloc_titre)],
                [graphique],
            ], colWidths=[70 * mm])
            bloc_graphique.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#173C5A")),
                ("BACKGROUND", (0, 1), (0, 1), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#B7C5D3")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]))

            indicateurs_table = Table([
                [Paragraph("Délai corrigé", style_carte_label), Paragraph(f"{delai_corrige_jour:.2f} j", style_carte_valeur)],
                [Paragraph("Jours actif consommé", style_carte_label), Paragraph(f"{jours_consommes:.2f} j", style_carte_valeur)],
                [Paragraph("Délai consommé", style_carte_label), Paragraph(f"{delai_consomme:.2f} j", style_carte_valeur)],
                [Paragraph("Jours restants", style_carte_label), Paragraph(f"{jours_restants:.2f} j", style_carte_valeur)],
                [Paragraph("Ecart en %", style_carte_label), Paragraph(f"{ecart_pct:.1f} %", style_carte_valeur_finale)],
            ], colWidths=largeur_blocs_droite)
            indicateurs_table.setStyle(TableStyle([
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#F6F9FC")]),
                ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#D6E0EA")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("BACKGROUND", (0, 4), (1, 4), colors.HexColor("#D9EAD3")),
            ]))

            indicateurs_cles = Table([
                [Paragraph("INDICATEURS CLÉS", style_bloc_titre)],
                [indicateurs_table],
            ], colWidths=[70 * mm])
            indicateurs_cles.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#365F91")),
                ("BACKGROUND", (0, 1), (0, 1), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#B7C5D3")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]))

            colonne_droite = Table([
                [bloc_graphique],
                [indicateurs_cles],
            ], colWidths=[70 * mm])
            colonne_droite.setStyle(TableStyle([
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))

            corps = Table([
                [colonne_gauche, colonne_droite]
            ], colWidths=[128 * mm, 70 * mm])
            corps.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]))
            elements.append(corps)
            elements.append(Spacer(1, 4))

            footer_row_height = 18 * mm
            footer_split_height = 6.5 * mm

            ecart_table = Table([
                [Paragraph("ECART %", style_ecart_titre)],
                [Paragraph(f"{ecart_pct:.1f} %", style_ecart_valeur)],
            ], colWidths=[30 * mm], rowHeights=[footer_split_height, footer_row_height - footer_split_height])
            ecart_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EAF1F8")),
                ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#B7C5D3")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))

            footer_explication = Table([
                [
                    Paragraph(
                        f"Jours restant pour exécution = Délai corrigé - Délai consommé<br/>{delai_corrige_jour:.2f} j - {delai_consomme:.2f} j = {jours_restants:.2f} j",
                        style_bandeau,
                    )
                ]
            ], colWidths=[168 * mm], rowHeights=[footer_row_height])
            footer_explication.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#FFF2CC")),
                ("BOX", (0, 0), (0, 0), 0.7, colors.HexColor("#D6B656")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))

            footer_table = Table([
                [footer_explication, ecart_table]
            ], colWidths=[168 * mm, 30 * mm], rowHeights=[footer_row_height])
            footer_table.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]))
            elements.append(footer_table)

            doc.build(elements)

            if graphique_tmp and os.path.exists(graphique_tmp):
                try:
                    os.unlink(graphique_tmp)
                except:
                    pass

            subprocess.run(["open", pdf_path])

        except Exception as e:
            if 'graphique_tmp' in locals() and graphique_tmp and os.path.exists(graphique_tmp):
                try:
                    os.unlink(graphique_tmp)
                except:
                    pass
            messagebox.showerror("Erreur pilotage délai", str(e))

    def rendements_chantier(self):
        import os
        import subprocess
        import tempfile
        from datetime import datetime
        import matplotlib.pyplot as plt
        from openpyxl import load_workbook
        from tkinter import messagebox
        from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Image, Paragraph
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_RIGHT, TA_CENTER
        from reportlab.lib.pagesizes import A4, landscape

        def nombre(valeur):
            try:
                if valeur is None:
                    return 0
                if isinstance(valeur, str):
                    valeur = valeur.replace("%", "").replace("€", "").replace(" ", "").replace(",", ".").strip()
                    if valeur in ("", "-", "—"):
                        return 0
                return float(valeur)
            except:
                return 0

        try:
            if not self.tree.selection():
                messagebox.showwarning("Attention", "Veuillez sélectionner un chantier.")
                return

            nom_chantier = self.tree.item(self.tree.selection()[0])["values"][0]
            dossier_chantier = dossier_chantiers() / nom_chantier
            fichier_delai = os.path.join(dossier_chantier, "Rendement.xlsx")

            if not os.path.exists(fichier_delai):
                messagebox.showerror("Erreur", f"Fichier introuvable :\n{fichier_delai}")
                return

            wb = load_workbook(fichier_delai, data_only=True)
            ws = wb.active

            heures_soumission = nombre(ws["O65"].value)
            heures_avenant_plus = nombre(ws["O69"].value)
            heures_attente_technique = nombre(ws["O73"].value)
            heures_imprevu = nombre(ws["O75"].value)
            heures_consommees = nombre(ws["O67"].value)
            heures_avenant_moins = nombre(ws["O71"].value)
            heures_corrigees = heures_soumission + heures_avenant_plus + heures_attente_technique + heures_imprevu - heures_avenant_moins
            heures_consommees_finales = heures_consommees + heures_avenant_moins
            heures_delai_consomme = heures_consommees_finales
            heures_restantes = heures_corrigees - heures_delai_consomme
            ecart_heures = heures_consommees_finales - heures_corrigees
            ecart_pct = nombre(ws["O82"].value)

            if ecart_pct <= 1 and ws["O82"].value not in (None, "", 0):
                ecart_pct = ecart_pct * 100

            valorisation_finale = ecart_heures * 55
            wb.close()

            texte = (
                f"Chantier : {nom_chantier}\n\n"

                f"CONSTRUCTION DU DÉLAI\n"
                f"Heures soumission : {heures_soumission:.2f} h\n"
                f"Heures accordées avenant en plus : {heures_avenant_plus:.2f} h\n"
                f"Heures accordées dû à une attente Technique : {heures_attente_technique:.2f} h\n"
                f"Heures accordées dû à imprévu, prévisible non visible : {heures_imprevu:.2f} h\n"
                f"Délai corrigé en heures : {heures_corrigees:.2f} h\n\n"

                f"CONSOMMÉ\n"
                f"Heures consommées : {heures_consommees:.2f} h\n"
                f"Heures accordées avenant en moins : {heures_avenant_moins:.2f} h\n"
                f"Heures consommées finales / délai consommé en heures : {heures_delai_consomme:.2f} h\n\n"

                f"RÉSULTAT\n"
                f"Heures restantes pour exécution : {heures_restantes:.2f} h\n"
                f"Écart en heures : {ecart_heures:.2f} h\n"
                f"Pourcentage : {ecart_pct:.1f} %\n"
                f"Valorisation financière finale : {valorisation_finale:,.2f} €"
            )

            messagebox.showinfo("⏱ Rendement chantier", texte)

            pdf_path = os.path.join(dossier_chantier, "rendement_chantier.pdf")
            doc = SimpleDocTemplate(
                pdf_path,
                pagesize=landscape(A4),
                leftMargin=10 * mm,
                rightMargin=10 * mm,
                topMargin=8 * mm,
                bottomMargin=8 * mm,
            )

            elements = []
            graphique_tmp = None

            logo_path = str(dossier_base() / "logo_jt_bati.png")
            date_du_jour = datetime.now().strftime("%d/%m/%Y")
            style_titre = ParagraphStyle(
                "RendementTitre",
                fontName="Helvetica-Bold",
                fontSize=18,
                leading=20,
                alignment=TA_CENTER,
                textColor=colors.HexColor("#163A59"),
            )
            style_meta_label = ParagraphStyle(
                "RendementMetaLabel",
                fontName="Helvetica-Bold",
                fontSize=8.5,
                leading=10,
                textColor=colors.HexColor("#36556F"),
            )
            style_meta_valeur = ParagraphStyle(
                "RendementMetaValeur",
                fontName="Helvetica",
                fontSize=8.8,
                leading=10.5,
                textColor=colors.black,
            )
            style_bloc_titre = ParagraphStyle(
                "RendementBlocTitre",
                fontName="Helvetica-Bold",
                fontSize=9.6,
                leading=11,
                alignment=TA_CENTER,
                textColor=colors.white,
            )
            style_libelle = ParagraphStyle(
                "RendementLibelle",
                fontName="Helvetica",
                fontSize=8.6,
                leading=9.8,
                textColor=colors.black,
            )
            style_valeur = ParagraphStyle(
                "RendementValeur",
                fontName="Helvetica-Bold",
                fontSize=8.8,
                leading=10,
                alignment=TA_RIGHT,
                textColor=colors.black,
            )
            style_carte_label = ParagraphStyle(
                "RendementCarteLabel",
                fontName="Helvetica-Bold",
                fontSize=7.6,
                leading=8.6,
                textColor=colors.HexColor("#5A7086"),
            )
            style_carte_valeur = ParagraphStyle(
                "RendementCarteValeur",
                fontName="Helvetica-Bold",
                fontSize=11.2,
                leading=12.5,
                alignment=TA_RIGHT,
                textColor=colors.HexColor("#173C5A"),
            )
            style_carte_valeur_finale = ParagraphStyle(
                "RendementCarteValeurFinale",
                fontName="Helvetica-Bold",
                fontSize=12.2,
                leading=13.5,
                alignment=TA_RIGHT,
                textColor=colors.HexColor("#9C3B32"),
            )
            style_bandeau = ParagraphStyle(
                "RendementBandeau",
                fontName="Helvetica-Bold",
                fontSize=9.3,
                leading=11,
                textColor=colors.HexColor("#173C5A"),
            )
            style_bandeau_valeur = ParagraphStyle(
                "RendementBandeauValeur",
                fontName="Helvetica-Bold",
                fontSize=10.3,
                leading=11.5,
                alignment=TA_RIGHT,
                textColor=colors.HexColor("#9C3B32"),
            )
            style_taux_titre = ParagraphStyle(
                "RendementTauxTitre",
                fontName="Helvetica-Bold",
                fontSize=8,
                leading=9,
                alignment=TA_CENTER,
                textColor=colors.HexColor("#5A7086"),
            )
            style_taux_valeur = ParagraphStyle(
                "RendementTauxValeur",
                fontName="Helvetica-Bold",
                fontSize=10.5,
                leading=12,
                alignment=TA_CENTER,
                textColor=colors.HexColor("#163A59"),
            )

            logo_cell = Spacer(1, 1)
            if os.path.exists(logo_path):
                try:
                    logo_cell = Image(logo_path, width=30 * mm, height=12 * mm)
                    logo_cell.hAlign = "LEFT"
                except:
                    logo_cell = Spacer(1, 1)

            meta_table = Table([
                [
                    Paragraph("CHANTIER", style_meta_label),
                    Paragraph(nom_chantier, style_meta_valeur),
                    Paragraph("DATE", style_meta_label),
                    Paragraph(date_du_jour, style_meta_valeur),
                ]
            ], colWidths=[18 * mm, 47 * mm, 12 * mm, 22 * mm])
            meta_table.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]))

            header_table = Table([
                [logo_cell, Paragraph("RENDEMENT CHANTIER", style_titre), meta_table]
            ], colWidths=[34 * mm, 158 * mm, 74 * mm])
            header_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F4F8FB")),
                ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#D9E4EE")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (1, 0), (1, 0), "CENTER"),
                ("ALIGN", (2, 0), (2, 0), "RIGHT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]))
            elements.append(header_table)
            elements.append(Spacer(1, 5))

            valeurs_graphique = [heures_corrigees, heures_delai_consomme, heures_restantes]
            etiquettes_graphique = ["Délai\ncorrigé", "Consommé\nfinal", "Heures\nrestantes"]
            couleurs_graphique = ["#1F4E79", "#E69138", "#70AD47"]

            fig, ax = plt.subplots(figsize=(3.25, 4.1), facecolor="white")
            barres = ax.bar(
                range(3),
                valeurs_graphique,
                width=0.55,
                color=couleurs_graphique,
                edgecolor="#FFFFFF",
                linewidth=0.8,
                zorder=3,
            )

            max_valeur = max([abs(v) for v in valeurs_graphique] + [1])
            marge = max_valeur * 0.18
            ax.set_ylim(0, max(max(valeurs_graphique), 1) + marge)
            ax.set_xticks(range(3), etiquettes_graphique)
            ax.tick_params(axis="x", labelsize=7.8, colors="#173C5A", length=0, pad=6)
            ax.tick_params(axis="y", labelsize=7.5, colors="#6A7E90")
            ax.grid(axis="y", color="#D9E4EE", linewidth=0.8, alpha=0.8, zorder=0)
            ax.set_axisbelow(True)
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            ax.spines["left"].set_color("#D9E4EE")
            ax.spines["bottom"].set_color("#D9E4EE")

            for barre, valeur in zip(barres, valeurs_graphique):
                ax.text(
                    barre.get_x() + barre.get_width() / 2,
                    barre.get_height() + max(marge * 0.08, 0.6),
                    f"{valeur:.2f} h",
                    ha="center",
                    va="bottom",
                    fontsize=7.8,
                    fontweight="bold",
                    color="#173C5A",
                )

            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
                graphique_tmp = tmp.name

            plt.tight_layout()
            plt.savefig(graphique_tmp, dpi=220, bbox_inches="tight", facecolor="white")
            plt.close(fig)

            graphique = Image(graphique_tmp, width=74 * mm, height=74 * mm)
            graphique.hAlign = "CENTER"

            largeur_blocs_gauche = [95 * mm, 30 * mm]
            largeur_blocs_droite = [38 * mm, 24 * mm]

            bloc_1 = Table([
                ["CONSTRUCTION DU DÉLAI", ""],
                [Paragraph("Heures soumission", style_libelle), Paragraph(f"{heures_soumission:.2f} h", style_valeur)],
                [Paragraph("+ Heures accordées avenant en plus", style_libelle), Paragraph(f"{heures_avenant_plus:.2f} h", style_valeur)],
                [Paragraph("+ Heures accordées dû à une attente Technique", style_libelle), Paragraph(f"{heures_attente_technique:.2f} h", style_valeur)],
                [Paragraph("+ Heures accordées dû à imprévu, prévisible non visible", style_libelle), Paragraph(f"{heures_imprevu:.2f} h", style_valeur)],
                [Paragraph("= Délai corrigé en heures", style_libelle), Paragraph(f"{heures_corrigees:.2f} h", style_valeur)],
            ], colWidths=largeur_blocs_gauche)
            bloc_1.hAlign = "CENTER"
            bloc_1.setStyle(TableStyle([
                ("SPAN", (0, 0), (1, 0)),
                ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#1F4E79")),
                ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
                ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (1, 0), 12),
                ("BACKGROUND", (0, 1), (1, -1), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (1, -1), [colors.white, colors.HexColor("#F8FBFF")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B7C5D3")),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("BACKGROUND", (0, 5), (1, 5), colors.HexColor("#FFF2CC")),
                ("FONTNAME", (0, 5), (1, 5), "Helvetica-Bold"),
            ]))

            bloc_2 = Table([
                ["CONSOMMÉ", ""],
                [Paragraph("- Heures consommées", style_libelle), Paragraph(f"{heures_consommees:.2f} h", style_valeur)],
                [Paragraph("- Heures accordées avenant en moins", style_libelle), Paragraph(f"{heures_avenant_moins:.2f} h", style_valeur)],
                [Paragraph("= Heures consommées finales / délai consommé en heures", style_libelle), Paragraph(f"{heures_delai_consomme:.2f} h", style_valeur)],
            ], colWidths=largeur_blocs_gauche)
            bloc_2.hAlign = "CENTER"
            bloc_2.setStyle(TableStyle([
                ("SPAN", (0, 0), (1, 0)),
                ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#4F81BD")),
                ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
                ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (1, 0), 12),
                ("BACKGROUND", (0, 1), (1, -1), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (1, -1), [colors.white, colors.HexColor("#F7FAFD")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B7C5D3")),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("BACKGROUND", (0, 3), (1, 3), colors.HexColor("#FFF2CC")),
                ("FONTNAME", (0, 3), (1, 3), "Helvetica-Bold"),
            ]))

            bloc_3 = Table([
                ["RÉSULTAT", ""],
                [Paragraph("Heures restantes pour exécution", style_libelle), Paragraph(f"{heures_restantes:.2f} h", style_valeur)],
                [Paragraph("Écart en heures", style_libelle), Paragraph(f"{ecart_heures:.2f} h", style_valeur)],
                [Paragraph("Pourcentage d’avancement", style_libelle), Paragraph(f"{ecart_pct:.1f} %", style_valeur)],
                [Paragraph("Valorisation financière finale (Écart en heures × 55 €)", style_libelle), Paragraph(f"{valorisation_finale:,.2f} €", style_valeur)],
            ], colWidths=largeur_blocs_gauche)
            bloc_3.hAlign = "CENTER"
            bloc_3.setStyle(TableStyle([
                ("SPAN", (0, 0), (1, 0)),
                ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#C0504D")),
                ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
                ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (1, 0), 12),
                ("BACKGROUND", (0, 1), (1, -1), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (1, -1), [colors.white, colors.HexColor("#FCF7F7")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#C9B1B1")),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("BACKGROUND", (0, 1), (1, 1), colors.HexColor("#E7E7E7")),
                ("BACKGROUND", (0, 4), (1, 4), colors.HexColor("#D9EAD3")),
                ("FONTNAME", (0, 4), (1, 4), "Helvetica-Bold"),
            ]))

            colonne_gauche = Table([
                [bloc_1],
                [bloc_2],
                [bloc_3],
            ], colWidths=[125 * mm])
            colonne_gauche.setStyle(TableStyle([
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))

            bloc_graphique = Table([
                [Paragraph("RÉSUMÉ VISUEL", style_bloc_titre)],
                [graphique],
            ], colWidths=[70 * mm])
            bloc_graphique.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#173C5A")),
                ("BACKGROUND", (0, 1), (0, 1), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#B7C5D3")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]))

            indicateurs_table = Table([
                [Paragraph("Délai corrigé en heures", style_carte_label), Paragraph(f"{heures_corrigees:.2f} h", style_carte_valeur)],
                [Paragraph("Heures consommées finales", style_carte_label), Paragraph(f"{heures_delai_consomme:.2f} h", style_carte_valeur)],
                [Paragraph("Heures restantes", style_carte_label), Paragraph(f"{heures_restantes:.2f} h", style_carte_valeur)],
                [Paragraph("Écart en heures", style_carte_label), Paragraph(f"{ecart_heures:.2f} h", style_carte_valeur)],
                [Paragraph("Pourcentage d’avancement", style_carte_label), Paragraph(f"{ecart_pct:.1f} %", style_carte_valeur)],
                [Paragraph("Valorisation financière finale", style_carte_label), Paragraph(f"{valorisation_finale:,.2f} €", style_carte_valeur_finale)],
            ], colWidths=largeur_blocs_droite)
            indicateurs_table.setStyle(TableStyle([
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#F6F9FC")]),
                ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#D6E0EA")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("BACKGROUND", (0, 5), (1, 5), colors.HexColor("#D9EAD3")),
            ]))

            indicateurs_cles = Table([
                [Paragraph("INDICATEURS CLÉS", style_bloc_titre)],
                [indicateurs_table],
            ], colWidths=[70 * mm])
            indicateurs_cles.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#365F91")),
                ("BACKGROUND", (0, 1), (0, 1), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#B7C5D3")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]))

            colonne_droite = Table([
                [bloc_graphique],
                [indicateurs_cles],
            ], colWidths=[70 * mm])
            colonne_droite.setStyle(TableStyle([
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))

            corps = Table([
                [colonne_gauche, colonne_droite]
            ], colWidths=[128 * mm, 70 * mm])
            corps.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]))
            elements.append(corps)
            elements.append(Spacer(1, 4))

            footer_row_height = 18 * mm
            footer_split_height = 6.5 * mm

            taux_table = Table([
                [Paragraph("TAUX HORAIRE", style_taux_titre)],
                [Paragraph("55.00 €/h", style_taux_valeur)],
            ], colWidths=[30 * mm], rowHeights=[footer_split_height, footer_row_height - footer_split_height])
            taux_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EAF1F8")),
                ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#B7C5D3")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))

            footer_explication = Table([
                [
                    Paragraph(
                        f"Valorisation financière finale = Écart en heures × 55 €/h<br/>{ecart_heures:.2f} h × 55 €/h = {valorisation_finale:,.2f} €",
                        style_bandeau,
                    )
                ]
            ], colWidths=[168 * mm], rowHeights=[footer_row_height])
            footer_explication.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#FFF2CC")),
                ("BOX", (0, 0), (0, 0), 0.7, colors.HexColor("#D6B656")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))

            footer_table = Table([
                [footer_explication, taux_table]
            ], colWidths=[168 * mm, 30 * mm], rowHeights=[footer_row_height])
            footer_table.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]))
            elements.append(footer_table)

            doc.build(elements)

            if graphique_tmp and os.path.exists(graphique_tmp):
                try:
                    os.unlink(graphique_tmp)
                except:
                    pass

            subprocess.run(["open", pdf_path])

        except Exception as e:
            if 'graphique_tmp' in locals() and graphique_tmp and os.path.exists(graphique_tmp):
                try:
                    os.unlink(graphique_tmp)
                except:
                    pass
            messagebox.showerror("Erreur rendement chantier", str(e))

    def _chemin_chantier_selectionne(self) -> Path | None:
        sel = self.tree.selection()
        if not sel:
            return None
        return Path(sel[0])

    def refresh_liste(self) -> None:
        self.lbl_path.config(text=f"Emplacement: {dossier_chantiers()}")
        for item in self.tree.get_children():
            self.tree.delete(item)

        d = dossier_chantiers()
        fichiers = sorted(d.glob("*.json"))

        for p in fichiers:
            try:
                chantier = lire_json(p)
                info = infos_chantier(chantier)
                av = info["avancement"]
                try:
                    av_txt = f"{float(av):.0f}%"
                except Exception:
                    av_txt = str(av)

                self.tree.insert(
                    "",
                    "end",
                    iid=str(p),
                    values=(info["nom"], info["client"], info["etat"], av_txt, info["debut"], info["fin"]),
                )
            except Exception:
                continue

    def _pr_path_and_hint(self) -> tuple[Path, str]:
        p = self._chemin_chantier_selectionne()
        if not p:
            raise ValueError("Sélectionne un chantier dans la liste.")
        dossier_chantier = p.parent / p.stem
        pr_path = dossier_chantier / "data" / "prix_de_revient.xlsx"
        if not pr_path.exists():
            raise FileNotFoundError(f"PR introuvable : {pr_path}")
        hint = pr_path.name
        return pr_path, hint

    def importer_bordereau(self) -> None:
        p = self._chemin_chantier_selectionne()
        if not p:
            messagebox.showwarning("Bordereau", "Sélectionne un chantier dans la liste.")
            return

        try:
            from app.bordereau import importer_bordereau_vso_xlsx
        except Exception as e:
            messagebox.showwarning("Bordereau", f"Import bordereau indisponible.\n{e}")
            return

        xlsx_path = filedialog.askopenfilename(
            title="Choisir le bordereau VSO (Excel)",
            filetypes=[("Excel", "*.xlsx;*.xlsm")],
        )
        if not xlsx_path:
            return

        dossier_chantier = p.parent / p.stem
        bordereau_xlsx = dossier_chantier / FICHIER_PV
        try:
            shutil.copy2(xlsx_path, bordereau_xlsx)
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de copier le bordereau dans le chantier.\n{e}")
            return

        try:
            chantier = lire_json(p)
            chantier = importer_bordereau_vso_xlsx(xlsx_path, chantier)
            ecrire_json(p, chantier)
            self.refresh_liste()
            afficher_bordereau(self, chantier, "Bordereau (JSON)")
            messagebox.showinfo("Bordereau", f"Bordereau importé.\nCopie locale: {bordereau_xlsx}")
        except Exception as e:
            messagebox.showerror("Erreur", f"Import bordereau impossible.\n{e}")

    def prix_revient(self) -> None:
        p = self._chemin_chantier_selectionne()
        if not p:
            messagebox.showwarning("Prix de revient", "Sélectionne un chantier dans la liste.")
            return
        dossier_chantier = p.parent / p.stem
        try:
            open_pr(dossier_chantier)
        except Exception as e:
            messagebox.showerror("Prix de revient", str(e))

    # ======================================================
    # ✅ OBJECTIF UNIQUE : AUTOMATISATION TEXTE -> COLONNE P
    # Matière  : P3:P12
    # MainOeuvre: P17:P26
    # ======================================================
    def recherche_matiere(self) -> None:
        try:
            pr_path, hint = self._pr_path_and_hint()
            subprocess.run(["open", str(pr_path)], check=False)
            time.sleep(0.3)

            items = read_biblio_colA_openpyxl(pr_path, "Bibliothèque_Matière.xlsx")

            def on_pick(txt: str):
                target = excel_find_first_empty_cell(hint, "Chiffrage", "P3:P12")
                if not target:
                    messagebox.showwarning("Matière", "Zone P3:P12 pleine ou Excel pas prêt.")
                    return
                excel_set_cell_value(hint, "Chiffrage", target, txt)

            popup_search_20(self, "Matière → Chiffrage (P3:P12)", items, on_pick)

        except Exception as e:
            messagebox.showerror("Matière", str(e))

    def recherche_mo(self) -> None:
        try:
            pr_path, hint = self._pr_path_and_hint()
            subprocess.run(["open", str(pr_path)], check=False)
            time.sleep(0.3)

            items = read_biblio_colA_openpyxl(pr_path, "Bibliothèque_MainOeuvre")

            def on_pick(txt: str):
                target = excel_find_first_empty_cell(hint, "Chiffrage", "P17:P26")
                if not target:
                    messagebox.showwarning("Main-d'œuvre", "Zone P17:P26 pleine ou Excel pas prêt.")
                    return
                excel_set_cell_value(hint, "Chiffrage", target, txt)

            popup_search_20(self, "Main-d’œuvre → Chiffrage (P17:P26)", items, on_pick)

        except Exception as e:
            messagebox.showerror("Main-d'œuvre", str(e))

    # ---------------------------
    # Boutons non prioritaires (inchangés)
    # ---------------------------
    
    def calcul_reca(self) -> None:
        import os
        from datetime import datetime
        from tkinter import messagebox

        p = self._chemin_chantier_selectionne()
        if not p:
            messagebox.showwarning("Calcul / Reca", "Sélectionne un chantier.")
            return

        dossier = self._dossier_chantier_selectionne()
        chemin_pr = dossier / "data" / "prix_de_revient.xlsx"

        if not chemin_pr.exists():
            messagebox.showwarning(
                "Date du PR",
                "Le fichier prix_de_revient.xlsx est introuvable."
            )
            return

        timestamp = os.path.getmtime(chemin_pr)
        date_modif = datetime.fromtimestamp(timestamp).strftime("%d/%m/%Y %H:%M:%S")

        messagebox.showinfo(
            "Date du PR",
            f"Dernier enregistrement du PR :\n{date_modif}\n\n"
            "⚠️ Si tu viens de modifier le PR :\n"
            "➡️ fais CMD + S dans Excel\n"
            "➡️ puis clique sur Calcul / Reca"
        )

        try:
            self.importer_pr_dans_etat()
            messagebox.showinfo(
                "Calcul / Reca",
                "Prix de revient importé dans l'état d'avancement."
            )
        except Exception as e:
            messagebox.showerror("Erreur", str(e))

    def _dossier_chantier_selectionne(self) -> Path:
        sel = self.tree.selection()
        if not sel:
            raise ValueError("Sélectionne un chantier dans la liste.")

        item = self.tree.item(sel[0])
        vals = item.get("values", [])
        chantier_sel = str(vals[0]).strip() if vals else ""

        base = dossier_chantiers()
        return base / chantier_sel

    def importer_pr_dans_etat(self):
        dossier = self._dossier_chantier_selectionne()
        chantier_nom = dossier.name.strip().lower()
        chemin_pr = dossier / "data" / "prix_de_revient.xlsx"

        try:
            chemin_etat = next(dossier.glob("Etat_avancement*.xlsm"))
        except StopIteration:
           return

        if not chemin_pr.exists():
            return

        wb_pr = load_workbook(chemin_pr, data_only=True)
        ws_pr = wb_pr["Chiffrage"]

        wb_etat = load_workbook(chemin_etat, keep_vba=True)
        ws_etat = wb_etat["Bordereau"]
        nb_importes = 0

        def clean_article(val):
            if val is None:
                return ""

            txt = str(val).strip().replace("\xa0", " ").replace(",", ".").lower()
            txt = re.sub(r"\s+", "", txt)
            if not txt:
                return ""

            parts = []
            for part in txt.split("."):
                if not part:
                    continue

                match = re.fullmatch(r"(\d+)([a-z]+)?", part)
                if match:
                    numero = str(int(match.group(1)))
                    suffixe = match.group(2) or ""
                    parts.append(numero + suffixe)
                    continue

                cleaned = re.sub(r"[^0-9a-z]+", "", part)
                if cleaned:
                    parts.append(cleaned)

            article = ".".join(parts)
            return article if any(c.isdigit() for c in article) else ""

        def cell_to_float(val):
            if isinstance(val, (int, float)):
                return float(val)
            if val is None:
                return 0.0
            txt = str(val).strip().replace("\xa0", " ").replace(" ", "").replace(",", ".")
            try:
                return float(txt)
            except ValueError:
                return 0.0

        pr_map = {}
        for row in range(3, ws_pr.max_row + 1):
            article = clean_article(ws_pr[f"B{row}"].value)
            if not article:
                continue

            ligne_valeur = row + 27
            valeur = ws_pr[f"F{ligne_valeur}"].value

            if isinstance(valeur, (int, float)):
                pr_map[article] = float(valeur)

        for row in range(24, ws_etat.max_row + 1):
            article = clean_article(ws_etat[f"B{row}"].value)
            if not article:
                continue

            if article in pr_map:
                prix = pr_map[article]
                if chantier_nom == "001_chapelette":
                    quantite = cell_to_float(ws_etat[f"E{row}"].value)
                    montant = round(quantite * prix, 2)
                    ws_etat[f"H{row}"] = prix
                    ws_etat[f"J{row}"] = montant
                    ws_etat[f"I{row}"] = montant_en_lettres_fr(montant)
                else:
                    ws_etat[f"L{row}"] = prix
                nb_importes += 1

        wb_etat.save(chemin_etat)
        wb_pr.close()
        wb_etat.close()

        from tkinter import messagebox
        messagebox.showinfo("Import PR", f"{nb_importes} article(s) importé(s).")
  
    def ouvrir_chantier(self) -> None:
        p = self._chemin_chantier_selectionne()
        if not p:
            messagebox.showwarning("Chantier", "Sélectionne un chantier dans la liste.")
            return
        try:
            chantier = lire_json(p)
            if chantier.get("bordereau", {}).get("articles"):
                afficher_bordereau(self, chantier, "Bordereau (JSON)")
            else:
                messagebox.showinfo("Chantier", "Chantier ouvert. (Pas encore de bordereau dans ce chantier.)")
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible d'ouvrir le chantier.\n{e}")

    def dossier_du_chantier(self) -> None:
        p = self._chemin_chantier_selectionne()
        if not p:
            messagebox.showwarning("Chantier", "Sélectionne un chantier dans la liste.")
            return
        ouvrir_dossier(p.parent)

    def nouveau_chantier(self) -> None:
        win = tk.Toplevel(self)
        win.title("Nouveau chantier")
        win.geometry("820x260")
        win.resizable(False, False)

        client_var = tk.StringVar(value="")
        chantier_var = tk.StringVar(value="")
        adresse_var = tk.StringVar(value="")

        frame = ttk.Frame(win, padding=14)
        frame.grid(row=0, column=0, sticky="nsew")
        win.columnconfigure(0, weight=1)
        win.rowconfigure(0, weight=1)

        ttk.Label(frame, text="Nom du client").grid(row=0, column=0, sticky="w")
        e_client = ttk.Entry(frame, textvariable=client_var, width=90)
        e_client.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(2, 8))

        ttk.Label(frame, text="Nom du chantier").grid(row=2, column=0, sticky="w")
        e_chantier = ttk.Entry(frame, textvariable=chantier_var, width=90)
        e_chantier.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(2, 8))

        ttk.Label(frame, text="Adresse").grid(row=4, column=0, sticky="w")
        e_adresse = ttk.Entry(frame, textvariable=adresse_var, width=90)
        e_adresse.grid(row=5, column=0, columnspan=2, sticky="ew", pady=(2, 12))

        def valider():
            nom_client = e_client.get().strip()
            nom_chantier = e_chantier.get().strip()
            adresse_chantier = e_adresse.get().strip()

            self.nom_client = nom_client
            self.nom_chantier = nom_chantier
            self.adresse_chantier = adresse_chantier

            if not nom_chantier:
                messagebox.showwarning("Nouveau chantier", "Renseigne le nom du chantier.")
                return

            nom_fichier = nom_chantier.replace("/", "_").replace("\\", "_")
            base = dossier_chantiers()
            json_path = base / f"{nom_fichier}.json"
            dossier_path = base / nom_fichier

            if json_path.exists() or dossier_path.exists():
                messagebox.showwarning("Nouveau chantier", "Ce chantier existe déjà.")
                return

            chantier = {
                "chantier": nom_chantier,
                "client": nom_client,
                "adresse": adresse_chantier,
                "etat": "Devis",
                "avancement": 0,
                "bordereau": {"source": {}, "articles": {}},
            }

            try:
                dossier_path.mkdir(parents=True, exist_ok=False)
                (dossier_path / "data").mkdir(exist_ok=True)
                ecrire_json(json_path, chantier)
                self.refresh_liste()
                if self.tree.exists(str(json_path)):
                    self.tree.selection_set(str(json_path))
                    self.tree.focus(str(json_path))
                win.destroy()
                return
            except Exception as e:
                messagebox.showerror("Nouveau chantier", f"Impossible de créer le chantier.\n{e}")
                return

            win.destroy()

        btns = ttk.Frame(frame)
        btns.grid(row=6, column=0, columnspan=2, sticky="e")
        ttk.Button(btns, text="Annuler", command=win.destroy).pack(side="right")
        ttk.Button(btns, text="Valider", command=valider).pack(side="right", padx=(0, 8))

        frame.columnconfigure(0, weight=1)
        e_client.focus_set()
        win.grab_set()



    def modifier_chantier(self) -> None:
        messagebox.showinfo("Info", "Modifier chantier (à brancher)")


    def supprimer_chantier(self) -> None:
        messagebox.showinfo("Info", "Supprimer chantier (à brancher)")


    def postes_metre(self) -> None:
        messagebox.showinfo("Info", "Postes / Métré (à faire)")

print("JE SUIS DANS LE BON FICHIER")

if __name__ == "__main__":
    app = HorizonChantierApp()
    app.mainloop()
