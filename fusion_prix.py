import pandas as pd
from datetime import datetime

SOURCE = "soumission_chapelette.xlsx"
OUT = f"CALCUL_soumission_{datetime.now().strftime('%Y_%m_%d_%H%M')}.xlsx"

# ============================================================
# Utils
# ============================================================
def norm(x) -> str:
    return str(x).replace("\xa0", " ").strip().lower()

def flatten_col(col) -> str:
    """Aplatit une colonne multi-index (tuple) en string lisible (2/3/4 niveaux+)."""
    if isinstance(col, tuple):
        parts = []
        for x in col:
            s = str(x).replace("\xa0", " ").strip()
            if s and s.lower() != "nan":
                parts.append(s)
        return " - ".join(parts).strip(" -")
    return str(col).replace("\xa0", " ").strip()

def find_table_header(xlsx_path: str, max_rows_scan: int = 350):
    """Trouve l'onglet + la ligne qui contient 'Article' et 'CSTC'."""
    xls = pd.ExcelFile(xlsx_path)
    for sh in xls.sheet_names:
        preview = pd.read_excel(xls, sheet_name=sh, header=None, nrows=max_rows_scan)
        for r in range(len(preview)):
            row = preview.iloc[r].tolist()
            row_join = " | ".join(norm(c) for c in row if str(c) != "nan")
            if "article" in row_join and "cstc" in row_join:
                return sh, r
    return None, None

def read_table_with_multiheader(xlsx_path: str, sheet: str, header_row: int):
    """
    Lit le tableau en essayant 4, puis 3, puis 2 lignes d'en-têtes.
    Retourne df avec colonnes aplaties.
    """
    # Essai 4 niveaux
    df = pd.read_excel(xlsx_path, sheet_name=sheet, header=[header_row, header_row+1, header_row+2, header_row+3])
    df.columns = [flatten_col(c) for c in df.columns]

    if not any(("cstc" in norm(c) and "article" in norm(c)) for c in df.columns):
        # Essai 3 niveaux
        df = pd.read_excel(xlsx_path, sheet_name=sheet, header=[header_row, header_row+1, header_row+2])
        df.columns = [flatten_col(c) for c in df.columns]

    if not any(("cstc" in norm(c) and "article" in norm(c)) for c in df.columns):
        # Essai 2 niveaux
        df = pd.read_excel(xlsx_path, sheet_name=sheet, header=[header_row, header_row+1])
        df.columns = [flatten_col(c) for c in df.columns]

    return df

# ============================================================
# 1) Trouver l'onglet + l'entête
# ============================================================
sheet, header_row = find_table_header(SOURCE)
if sheet is None:
    raise ValueError("Impossible de trouver la ligne d'entête (Article/CSTC) dans le fichier.")

print(f"✅ Tableau trouvé dans l'onglet: '{sheet}', ligne entête (0-index): {header_row}")

# ============================================================
# 2) Lire le tableau
# ============================================================
df = read_table_with_multiheader(SOURCE, sheet, header_row)

# Enlever les colonnes 'Ellipsis' si elles existent (artifact Excel)
df = df.loc[:, [c for c in df.columns if norm(c) != "ellipsis"]]

# Détecter colonne poste (CSTC)
poste_candidates = [c for c in df.columns if ("cstc" in norm(c) and "article" in norm(c))]
if not poste_candidates:
    print("🔎 Aperçu colonnes (30 premières) :")
    print(df.columns[:30].tolist())
    raise ValueError("Colonne poste (Article/CSTC) introuvable après lecture.")

POSTE_COL = poste_candidates[0]
print(f"✅ Colonne poste détectée: '{POSTE_COL}'")

# Garder uniquement les lignes qui ont un poste
df = df[df[POSTE_COL].notna()].copy()

# ============================================================
# 3) Détecter colonne quantité soumissionnaire
# ============================================================
qty_col = None

# Priorité 1: colonne contenant 'quantit' + 'soumissionnaire'
for c in df.columns:
    cc = norm(c)
    if "quantit" in cc and "soumissionnaire" in cc:
        qty_col = c
        break

# Priorité 2: colonne contenant 'quantit' + 'q.p.' ou 'qp'
if qty_col is None:
    for c in df.columns:
        cc = norm(c)
        if "quantit" in cc and ("q.p" in cc or "qp" in cc):
            qty_col = c
            break

# Priorité 3: dernier recours -> liste
if qty_col is None or qty_col not in df.columns or norm(qty_col) == "ellipsis":
    print("⚠️ Quantité soumissionnaire non détectée automatiquement.")
    print("🔎 Colonnes contenant 'quant' :")
    for c in df.columns:
        if "quant" in norm(c):
            print(" -", c)
    # On met une quantité=1 pour ne pas planter l'export, mais on avertit
    df["__QTE__"] = 1.0
    qty_col = "__QTE__"
    print("➡️ Quantité=1 provisoire. Donne-moi le bon nom de colonne et je verrouille.")
else:
    print(f"✅ Colonne quantité détectée: '{qty_col}'")

# ============================================================
# 4) Renommer poste + quantité (fichier de calcul lisible)
# ============================================================
rename_map = {}
if POSTE_COL in df.columns:
    rename_map[POSTE_COL] = "Poste (CSTC)"
if qty_col in df.columns and qty_col != "__QTE__":
    rename_map[qty_col] = "Quantité soumissionnaire"

df = df.rename(columns=rename_map)

# Mettre à jour les variables
if "Poste (CSTC)" in df.columns:
    POSTE_COL = "Poste (CSTC)"
if "Quantité soumissionnaire" in df.columns:
    qty_col = "Quantité soumissionnaire"

# ============================================================
# 5) Ajouter colonnes de calcul (interne)
# ============================================================
colonnes_calcul = {
    # Fourniture / métrés
    "Fourniture (libellé)": "",
    "Qté fourniture": 0.0,
    "Unité fourniture": "",
    "Longueur": 0.0,
    "Largeur": 0.0,
    "Hauteur": 0.0,
    "ML": 0.0,
    "M2": 0.0,
    "M3": 0.0,
    "KG": 0.0,
    "Jour": 0.0,
    "Transport aller": 0.0,
    "Transport retour": 0.0,
    "Prioritaire": "",
    "Prix unitaire fourniture": 0.0,
    "Total fourniture": 0.0,

    # Main-d’œuvre
    "Description MO": "",
    "Nombre d'hommes": 0.0,
    "Heures / homme": 0.0,
    "Total heures": 0.0,
    "Taux horaire": 0.0,
    "Total main-d’œuvre": 0.0,

    # Résumé
    "Prix de revient (poste)": 0.0,
    "Marge % (calcul)": 0.0,
    "Prix unitaire (calcul)": 0.0,
    "Prix global (calcul)": 0.0,
}
for c, default in colonnes_calcul.items():
    if c not in df.columns:
        df[c] = default

# ============================================================
# 6) Calculs
# ============================================================
to_num = lambda s: pd.to_numeric(s, errors="coerce").fillna(0)

df["ML"] = to_num(df["Longueur"])
df["M2"] = to_num(df["Longueur"]) * to_num(df["Largeur"])
df["M3"] = df["M2"] * to_num(df["Hauteur"])

df["Total fourniture"] = to_num(df["Qté fourniture"]) * to_num(df["Prix unitaire fourniture"])

df["Total heures"] = to_num(df["Nombre d'hommes"]) * to_num(df["Heures / homme"])
df["Total main-d’œuvre"] = df["Total heures"] * to_num(df["Taux horaire"])

df["Prix de revient (poste)"] = df["Total fourniture"] + df["Total main-d’œuvre"]
df["Prix unitaire (calcul)"] = df["Prix de revient (poste)"] * (1 + to_num(df["Marge % (calcul)"]) / 100)

df["Prix global (calcul)"] = df["Prix unitaire (calcul)"] * to_num(df[qty_col])

# ============================================================
# 7) Export
# ============================================================
# ============================================================
# FORÇAGE FINAL DES COLONNES PRIX (sécurité)
# ============================================================
if "Prix de revient (poste)" not in df.columns:
    df["Prix de revient (poste)"] = (
        pd.to_numeric(df.get("Total fourniture", 0), errors="coerce").fillna(0)
        + pd.to_numeric(df.get("Total main-d’œuvre", 0), errors="coerce").fillna(0)
    )

if "Prix unitaire (calcul)" not in df.columns:
    df["Prix unitaire (calcul)"] = (
        df["Prix de revient (poste)"]
        * (1 + pd.to_numeric(df.get("Marge % (calcul)", 0), errors="coerce").fillna(0) / 100)
    )

if "Prix global (calcul)" not in df.columns:
    df["Prix global (calcul)"] = (
        df["Prix unitaire (calcul)"]
        * pd.to_numeric(df[qty_col], errors="coerce").fillna(0)
    )
# ============================================================
# EXPORT SIMPLE POUR REPORT MANUEL (marché public)
# ============================================================
colonnes_report = [
    "Poste (CSTC)",
    qty_col,
    "Prix de revient (poste)",
    "Marge % (calcul)",
    "Prix unitaire (calcul)",
    "Prix global (calcul)"
]

# garder seulement celles qui existent réellement
colonnes_report = [c for c in colonnes_report if c in df.columns]

df_report = df[colonnes_report].copy()

report_file = f"REPORT_soumission_{datetime.now().strftime('%Y_%m_%d_%H%M')}.xlsx"
df_report.to_excel(report_file, index=False)
print(f"✅ Fichier REPORT créé : {report_file}")

df.to_excel(OUT, index=False)
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# =========================
# AJOUT FORMULES EXCEL
# =========================
wb = load_workbook(OUT)
ws = wb.active

# Récupérer les en-têtes (ligne 1) -> index colonne
headers = {}
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    if val:
        headers[str(val).strip()] = col

def col_letter(name: str) -> str:
    if name not in headers:
        raise ValueError(f"Colonne '{name}' introuvable pour formules Excel.")
    return get_column_letter(headers[name])

# Noms de colonnes (doivent correspondre à ceux de ton fichier CALCUL)
C_QTE_FOURN = col_letter("Qté fourniture")
C_PU_FOURN  = col_letter("Prix unitaire fourniture")
C_TOT_FOURN = col_letter("Total fourniture")

C_NB_H      = col_letter("Nombre d'hommes")
C_H_PAR_H   = col_letter("Heures / homme")
C_TOT_H     = col_letter("Total heures")

C_TAUX      = col_letter("Taux horaire")
C_TOT_MO    = col_letter("Total main-d’œuvre")

C_REV       = col_letter("Prix de revient (poste)")
C_MARGE     = col_letter("Marge % (calcul)")
C_PU_CALC   = col_letter("Prix unitaire (calcul)")
C_QTE_S     = col_letter("Quantité soumissionnaire")
C_PG_CALC   = col_letter("Prix global (calcul)")

# Appliquer les formules sur toutes les lignes de données
for r in range(2, ws.max_row + 1):
    ws[f"{C_TOT_FOURN}{r}"] = f"={C_QTE_FOURN}{r}*{C_PU_FOURN}{r}"
    ws[f"{C_TOT_H}{r}"]     = f"={C_NB_H}{r}*{C_H_PAR_H}{r}"
    ws[f"{C_TOT_MO}{r}"]    = f"={C_TOT_H}{r}*{C_TAUX}{r}"
    ws[f"{C_REV}{r}"]       = f"={C_TOT_FOURN}{r}+{C_TOT_MO}{r}"
    ws[f"{C_PU_CALC}{r}"]   = f"={C_REV}{r}*(1+{C_MARGE}{r}/100)"
    ws[f"{C_PG_CALC}{r}"]   = f"={C_PU_CALC}{r}*{C_QTE_S}{r}"

wb.save(OUT)
print("✅ Formules Excel injectées dans le fichier CALCUL (recalcul automatique dans Excel).")

print(f"✅ Fichier de calcul créé: {OUT}")
print("ℹ️ Le fichier officiel n'a PAS été modifié.")
# =========================
# EXPORT "A REPORTER" (valeurs)
# =========================
a_reporter_file = f"A_REPORTER_{datetime.now().strftime('%Y_%m_%d_%H%M')}.xlsx"

colonnes = [
    "Poste (CSTC)",
    "Quantité soumissionnaire",
    "Prix unitaire (calcul)",
    "Prix global (calcul)"
]

# garder seulement celles qui existent
colonnes = [c for c in colonnes if c in df.columns]

df_a = df[colonnes].copy()
df_a = df_a.rename(columns={
    "Prix unitaire (calcul)": "Prix unitaire à reporter",
    "Prix global (calcul)": "Somme à reporter"
})

df_a.to_excel(a_reporter_file, index=False)
print(f"✅ Fichier prêt à reporter créé : {a_reporter_file}")
