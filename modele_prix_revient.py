from openpyxl import load_workbook
from pathlib import Path
import shutil

# 1) Chemin de TON modèle Excel (change le nom ici)
MODELE = Path("modele_prix_revient.xlsx")

# 2) Fichier de sortie (on ne touche jamais au modèle)
SORTIE = Path("test de sortie xlsx/prix_revient_calcule.xlsx")
SORTIE.parent.mkdir(parents=True, exist_ok=True)

# Copie du modèle -> sortie
shutil.copy(MODELE, SORTIE)

# Ouvrir le fichier copié
wb = load_workbook(SORTIE)
ws = wb.active  # 1ère feuille

# ⚠️ IMPORTANT :
# Ici on met les colonnes EXACTES de ton tableau (lettres Excel).
# Tu me dis si ce n'est pas ça et je corrige en 30 sec.

# Exemple (à adapter à ton fichier):
# Quantité poste = G
# Perte % = H
# Total matière = I
# PU net = J
# PU revient = K
# Marge chantier = L
# Marge vente = M
# Total ligne = N

for r in range(3, ws.max_row + 1):
    # Total matière (poste) = Qté poste * PU net * (1 + perte%)
    ws[f"I{r}"].value = f"=G{r}*J{r}*(1+H{r})"

    # PU revient (poste) = Total matière / Qté poste (si Qté poste > 0)
    ws[f"K{r}"].value = f"=IF(G{r}=0,\"\",I{r}/G{r})"

    # Total ligne = PU revient * Qté poste * (1+marge chantier+marge vente)
    ws[f"N{r}"].value = f"=IF(G{r}=0,\"\",K{r}*G{r}*(1+L{r}+M{r}))"

# Sauvegarde
wb.save(SORTIE)
print("✅ Fichier créé :", SORTIE)
