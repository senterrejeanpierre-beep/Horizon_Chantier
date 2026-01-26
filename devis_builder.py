
# devis_builder.py
# Génère un classeur "prix de revient pour soumissionner" (pro chantier)
# Feuilles: PARAMS, PR_MATIERE, PR_MO, EXECUTION, SYNTHESE, SOUMISSION
# Sortie unique: SORTIES/DEVIS_TEST.xlsx (écrasée à chaque run)

from pathlib import Path
from datetime import date
import sys

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def chemin(relatif: str) -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys._MEIPASS) / relatif
    return Path(__file__).parent / relatif


def build_prix_revient(output_path: Path, n_rows: int = 10) -> Path:  # <= 10 lignes par défaut
    modele_path = chemin("soumission_chapelette.xlsx")
  wb = load_workbook(...)
header_fill = PatternFill(...)       # ✅ alignée
 
print("MODELE:", modele_path)
print("ONGLETS:", wb.sheetnames)


    # ---------------- Styles (créés 1x)



    header_fill = PatternFill("solid", fgColor="EDEDED")
    input_fill = PatternFill("solid", fgColor="D9E8FF")
    calc_fill = PatternFill("solid", fgColor="F3F3F3")

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    prot_locked = Protection(locked=True)
    prot_unlocked = Protection(locked=False)

    money_fmt = '#,##0.00 "€"'
    pct_fmt = "0.00%"
    qty_fmt = "#,##0.00"

    def style_header_row(ws, row, ncols):
        for col in range(1, ncols + 1):
            c = ws.cell(row=row, column=col)
            c.font = bold
            c.fill = header_fill
            c.alignment = center
            c.border = border

    def style_row(ws, r, ncols):
        for col in range(1, ncols + 1):
            c = ws.cell(row=r, column=col)
            c.border = border
            c.alignment = left

    def set_widths(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ---------------- PARAMS
    ws_params = wb.create_sheet("PARAMS")
    ws_params.append(["Paramètre", "Valeur"])
    style_header_row(ws_params, 1, 2)

    params = [
        ("Taux horaire (€/h)", 55),    # B2
        ("Frais généraux %", 0.10),    # B3
        ("Marge aménagement %", 0.10), # B4
        ("Marge vente %", 0.15),       # B5
        ("TVA %", 0.21),               # B6
    ]
    for i, (k, v) in enumerate(params, start=2):
        ws_params[f"A{i}"] = k
        ws_params[f"B{i}"] = v
        ws_params[f"A{i}"].border = border
        ws_params[f"B{i}"].border = border
        ws_params[f"A{i}"].alignment = left
        ws_params[f"B{i}"].alignment = left
        if isinstance(v, float) and v <= 1:
            ws_params[f"B{i}"].number_format = pct_fmt
        else:
            ws_params[f"B{i}"].number_format = qty_fmt

    set_widths(ws_params, [28, 16])
    ws_params.freeze_panes = "A2"

    # Références paramètres
    taux_h = "PARAMS!$B$2"
    fg_pct = "PARAMS!$B$3"
    marge_am_def = "PARAMS!$B$4"
    marge_vente_def = "PARAMS!$B$5"
    tva_pct = "PARAMS!$B$6"

    # ---------------- PRIX DE REVIENT MATIERE
    ws_mat = wb.create_sheet("PR_MATIERE")
    headers_mat = [
        "N° poste", "Poste", "Texte",
        "Désignation (matière)", "Unité", "Qté poste",
        "Perte %", "PU net",
        "Total matière (poste)", "PU revient (poste)",
        "Marge aménagement %", "Marge vente %",
        "Total ligne (vente)"
    ]
    ws_mat.append(headers_mat)
    style_header_row(ws_mat, 1, len(headers_mat))
    set_widths(ws_mat, [10, 18, 14, 32, 8, 12, 10, 12, 16, 14, 18, 14, 16])
    ws_mat.freeze_panes = "A2"

    start = 2
    end = start + n_rows - 1

    # Data validation (⚠️ 1 objet DV par feuille)
    dv_pct_mat = DataValidation(type="decimal", operator="between", formula1="0", formula2="1", allow_blank=True)
    dv_pos_mat = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    ws_mat.add_data_validation(dv_pct_mat)
    ws_mat.add_data_validation(dv_pos_mat)

    for r in range(start, end + 1):
        style_row(ws_mat, r, len(headers_mat))

        # valeurs par défaut marges
        ws_mat[f"K{r}"].value = f"={marge_am_def}"
        ws_mat[f"L{r}"].value = f"={marge_vente_def}"

        ws_mat[f"I{r}"].value = f'=IFERROR(F{r}*H{r}*(1+G{r}),"")'
        ws_mat[f"J{r}"].value = f'=IFERROR(I{r}/F{r},"")'
        ws_mat[f"M{r}"].value = f'=IFERROR(J{r}*F{r}*(1+K{r})*(1+L{r}),"")'

        # Formats
        ws_mat[f"F{r}"].number_format = qty_fmt
        ws_mat[f"G{r}"].number_format = pct_fmt
        ws_mat[f"H{r}"].number_format = money_fmt
        ws_mat[f"I{r}"].number_format = money_fmt
        ws_mat[f"J{r}"].number_format = money_fmt
        ws_mat[f"K{r}"].number_format = pct_fmt
        ws_mat[f"L{r}"].number_format = pct_fmt
        ws_mat[f"M{r}"].number_format = money_fmt

        # Validations
        dv_pct_mat.add(f"G{r}")   # perte
        dv_pct_mat.add(f"K{r}")   # marge am
        dv_pct_mat.add(f"L{r}")   # marge vente
        dv_pos_mat.add(f"F{r}")   # qté
        dv_pos_mat.add(f"H{r}")   # pu net

        # Couleurs input / calc
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "K", "L"]:
            ws_mat[f"{col}{r}"].fill = input_fill
            ws_mat[f"{col}{r}"].protection = prot_unlocked
        for col in ["I", "J", "M"]:
            ws_mat[f"{col}{r}"].fill = calc_fill
            ws_mat[f"{col}{r}"].protection = prot_locked

    ws_mat.protection.sheet = True

    # ---------------- PRIX DE REVIENT MAIN D'OEUVRE
    ws_mo = wb.create_sheet("PR_MO")
    headers_mo = [
        "N° poste", "Poste", "Texte",
        "Désignation (MO)", "Nb hommes", "Heures / homme / jour",
        "Délai (jours)", "Total heures (poste)",
        "Taux horaire", "Coût MO (poste)",
        "Marge aménagement %", "Marge vente %",
        "Total ligne (vente)"
    ]
    ws_mo.append(headers_mo)
    style_header_row(ws_mo, 1, len(headers_mo))
    set_widths(ws_mo, [10, 18, 14, 34, 10, 18, 12, 16, 12, 16, 18, 14, 16])
    ws_mo.freeze_panes = "A2"

    # Data validation (⚠️ 1 objet DV par feuille)
    dv_pct_mo = DataValidation(type="decimal", operator="between", formula1="0", formula2="1", allow_blank=True)
    dv_pos_mo = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    ws_mo.add_data_validation(dv_pct_mo)
    ws_mo.add_data_validation(dv_pos_mo)

    for r in range(start, end + 1):
        style_row(ws_mo, r, len(headers_mo))

        ws_mo[f"I{r}"].value = f"={taux_h}"
        ws_mo[f"K{r}"].value = f"={marge_am_def}"
        ws_mo[f"L{r}"].value = f"={marge_vente_def}"

        ws_mo[f"H{r}"].value = f'=IFERROR(E{r}*F{r}*G{r},"")'
        ws_mo[f"J{r}"].value = f'=IFERROR(H{r}*I{r},"")'
        ws_mo[f"M{r}"].value = f'=IFERROR(J{r}*(1+K{r})*(1+L{r}),"")'

        # Formats
        ws_mo[f"E{r}"].number_format = qty_fmt
        ws_mo[f"F{r}"].number_format = qty_fmt
        ws_mo[f"G{r}"].number_format = qty_fmt
        ws_mo[f"H{r}"].number_format = qty_fmt
        ws_mo[f"I{r}"].number_format = money_fmt
        ws_mo[f"J{r}"].number_format = money_fmt
        ws_mo[f"K{r}"].number_format = pct_fmt
        ws_mo[f"L{r}"].number_format = pct_fmt
        ws_mo[f"M{r}"].number_format = money_fmt

        # Validations
        dv_pos_mo.add(f"E{r}")  # nb hommes
        dv_pos_mo.add(f"F{r}")  # h/j
        dv_pos_mo.add(f"G{r}")  # délai
        dv_pct_mo.add(f"K{r}")  # marge am
        dv_pct_mo.add(f"L{r}")  # marge vente

        # Couleurs input / calc
        for col in ["A", "B", "C", "D", "E", "F", "G", "K", "L"]:
            ws_mo[f"{col}{r}"].fill = input_fill
            ws_mo[f"{col}{r}"].protection = prot_unlocked

        # Taux horaire (I) = param mais editable -> bleu
        ws_mo[f"I{r}"].fill = input_fill
        ws_mo[f"I{r}"].protection = prot_unlocked

        for col in ["H", "J", "M"]:
            ws_mo[f"{col}{r}"].fill = calc_fill
            ws_mo[f"{col}{r}"].protection = prot_locked

    ws_mo.protection.sheet = True

    # ---------------- EXECUTION (réel chantier + écarts)
    ws_exec = wb.create_sheet("EXECUTION")
    headers_exec = [
        "N° poste", "Poste",
        "Heures réelles (poste)", "Taux horaire réel", "Coût MO réel (poste)",
        "Coût matière réel (poste)",
        "Coût réel total (poste)",
        "Coût prévu total (poste)",
        "Écart € (réel - prévu)",
        "Écart % (réel - prévu)"
    ]
    ws_exec.append(headers_exec)
    style_header_row(ws_exec, 1, len(headers_exec))
    set_widths(ws_exec, [10, 18, 18, 16, 18, 18, 18, 18, 16, 16])
    ws_exec.freeze_panes = "A2"

    # Data validation (⚠️ 1 objet DV par feuille)
    dv_pos_exec = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    ws_exec.add_data_validation(dv_pos_exec)

    for r in range(start, end + 1):
        style_row(ws_exec, r, len(headers_exec))

        # défaut taux horaire réel = taux horaire params
        ws_exec[f"D{r}"].value = f"={taux_h}"

        # Coût MO réel
        ws_exec[f"E{r}"].value = f'=IFERROR(C{r}*D{r},"")'

        # Coût réel total = MO réel + matière réelle
        ws_exec[f"G{r}"].value = f'=IFERROR(E{r}+F{r},"")'

        # Coût prévu total = PR_MATIERE (Total matière) + PR_MO (Coût MO)
        ws_exec[f"H{r}"].value = (
            f'=IFERROR('
            f'SUMIF(PR_MATIERE!$A:$A,$A{r},PR_MATIERE!$I:$I)'
            f'+SUMIF(PR_MO!$A:$A,$A{r},PR_MO!$J:$J)'
            f',0)'
        )

        ws_exec[f"I{r}"].value = f'=IFERROR(G{r}-H{r},"")'
        ws_exec[f"J{r}"].value = f'=IFERROR(I{r}/H{r},"")'

        # formats
        ws_exec[f"C{r}"].number_format = qty_fmt
        ws_exec[f"D{r}"].number_format = money_fmt
        ws_exec[f"E{r}"].number_format = money_fmt
        ws_exec[f"F{r}"].number_format = money_fmt
        ws_exec[f"G{r}"].number_format = money_fmt
        ws_exec[f"H{r}"].number_format = money_fmt
        ws_exec[f"I{r}"].number_format = money_fmt
        ws_exec[f"J{r}"].number_format = pct_fmt

        # validations
        dv_pos_exec.add(f"C{r}")
        dv_pos_exec.add(f"D{r}")
        dv_pos_exec.add(f"F{r}")

        # couleurs input/calc
        for col in ["A", "B", "C", "D", "F"]:
            ws_exec[f"{col}{r}"].fill = input_fill
            ws_exec[f"{col}{r}"].protection = prot_unlocked
        for col in ["E", "G", "H", "I", "J"]:
            ws_exec[f"{col}{r}"].fill = calc_fill
            ws_exec[f"{col}{r}"].protection = prot_locked

    ws_exec.protection.sheet = True

    # ---------------- SYNTHESE (prix de revient pour soumissionner)
    ws_syn = wb.create_sheet("SYNTHESE")
    ws_syn.append(["Élément", "Valeur"])
    style_header_row(ws_syn, 1, 2)
    set_widths(ws_syn, [28, 18])

    ws_syn["A2"] = "Total matière (prévu)"
    ws_syn["B2"] = f'=IFERROR(SUM(PR_MATIERE!I{start}:I{end}),0)'
    ws_syn["A3"] = "Total MO (prévu)"
    ws_syn["B3"] = f'=IFERROR(SUM(PR_MO!J{start}:J{end}),0)'
    ws_syn["A4"] = "Coût direct prévu"
    ws_syn["B4"] = "=IFERROR(B2+B3,0)"
    ws_syn["A5"] = "Frais généraux (prévu)"
    ws_syn["B5"] = f"=IFERROR(B4*{fg_pct},0)"
    ws_syn["A6"] = "PRIX DE REVIENT TOTAL (prévu)"
    ws_syn["B6"] = "=IFERROR(B4+B5,0)"

    ws_syn["A8"] = "Total vente (matière)"
    ws_syn["B8"] = f'=IFERROR(SUM(PR_MATIERE!M{start}:M{end}),0)'
    ws_syn["A9"] = "Total vente (MO)"
    ws_syn["B9"] = f'=IFERROR(SUM(PR_MO!M{start}:M{end}),0)'
    ws_syn["A10"] = "PRIX DE VENTE TOTAL HT"
    ws_syn["B10"] = "=IFERROR(B8+B9,0)"
    ws_syn["A11"] = "TVA %"
    ws_syn["B11"] = f"={tva_pct}"
    ws_syn["A12"] = "PRIX DE VENTE TOTAL TTC"
    ws_syn["B12"] = "=IFERROR(B10*(1+B11),0)"

    for r in [2, 3, 4, 5, 6, 8, 9, 10, 12]:
        ws_syn[f"B{r}"].number_format = money_fmt
        ws_syn[f"A{r}"].border = border
        ws_syn[f"B{r}"].border = border
        ws_syn[f"A{r}"].alignment = left
        ws_syn[f"B{r}"].alignment = left

    ws_syn["B11"].number_format = pct_fmt
    ws_syn["A11"].border = border
    ws_syn["B11"].border = border
    ws_syn["A11"].alignment = left
    ws_syn["B11"].alignment = left

    ws_syn["A6"].font = bold
    ws_syn["B6"].font = bold
    ws_syn["A10"].font = bold
    ws_syn["B10"].font = bold
    ws_syn["A12"].font = bold
    ws_syn["B12"].font = bold

    ws_syn.freeze_panes = "A2"

        # ---------------- SOUMISSION (fiche client pro A4)
    ws_soum = wb.create_sheet("SOUMISSION")

    # Mise en page
    ws_soum.page_setup.paperSize = ws_soum.PAPERSIZE_A4
    ws_soum.page_setup.orientation = ws_soum.ORIENTATION_PORTRAIT
    ws_soum.page_margins.left = 0.5
    ws_soum.page_margins.right = 0.5
    ws_soum.page_margins.top = 0.75
    ws_soum.page_margins.bottom = 0.75

    # Largeurs colonnes
    set_widths(ws_soum, [18, 30, 18, 18])

    # Zones principales
    ws_soum.merge_cells("A1:D3")   # En-tête
    ws_soum.merge_cells("A5:D5")   # Titre
    ws_soum.merge_cells("A7:B10")  # Client
    ws_soum.merge_cells("C7:D10")  # Chantier
    ws_soum.merge_cells("A12:D12") # Séparateur
    ws_soum.merge_cells("A14:C14")
    ws_soum.merge_cells("A15:C15")
    ws_soum.merge_cells("A16:C16")

    # En-tête entreprise
    ws_soum["A1"] = "ENTREPRISE XYZ\nAdresse\nTéléphone\nEmail"
    ws_soum["A1"].font = Font(bold=True, size=14)
    ws_soum["A1"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Titre
    ws_soum["A5"] = "DEVIS"
    ws_soum["A5"].font = Font(bold=True, size=20)
    ws_soum["A5"].alignment = Alignment(horizontal="center", vertical="center")

    # Infos client
    ws_soum["A7"] = "CLIENT :\n\nÀ compléter"
    ws_soum["A7"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws_soum["A7"].font = Font(size=11)

    # Infos chantier
    ws_soum["C7"] = "CHANTIER :\n\nÀ compléter"
    ws_soum["C7"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws_soum["C7"].font = Font(size=11)

    # Séparateur
    ws_soum["A12"] = ""
    ws_soum["A12"].fill = PatternFill("solid", fgColor="EDEDED")

    # Bloc prix
    ws_soum["A14"] = "Prix de revient total"
    ws_soum["D14"] = "=SYNTHESE!B6"

    ws_soum["A15"] = "Prix de vente HT"
    ws_soum["D15"] = "=SYNTHESE!B10"

    ws_soum["A16"] = "TVA"
    ws_soum["D16"] = "=SYNTHESE!B11"

    ws_soum["A17"] = "Prix de vente TTC"
    ws_soum["D17"] = "=SYNTHESE!B12"

    # Formats
    ws_soum["D14"].number_format = money_fmt
    ws_soum["D15"].number_format = money_fmt
    ws_soum["D16"].number_format = pct_fmt
    ws_soum["D17"].number_format = money_fmt

    # Style prix
    for r in [14,15,16,17]:
        ws_soum[f"A{r}"].font = Font(bold=True)
        ws_soum[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center")
        ws_soum[f"D{r}"].font = Font(bold=True)
        ws_soum[f"D{r}"].alignment = Alignment(horizontal="right", vertical="center")

    # Signature
    ws_soum.merge_cells("A20:D22")
    ws_soum["A20"] = "Signature client :\n\n\nSignature entreprise :"
    ws_soum["A20"].alignment = Alignment(horizontal="left", vertical="top")

    # Zone d'impression
    ws_soum.print_title_rows = "1:22"
    ws_soum.print_area = "A1:D22"


    # ---------------- Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main():
    root = Path(__file__).resolve().parent
    out = root / "SORTIES" / "DEVIS_TEST.xlsx"   # sortie unique
    p = build_prix_revient(out, n_rows=10)       # <= 10 lignes
    print("OK ->", p)


if __name__ == "__main__":
    main()
