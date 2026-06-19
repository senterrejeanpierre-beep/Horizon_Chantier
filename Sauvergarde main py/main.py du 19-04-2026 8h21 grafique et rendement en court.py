#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import subprocess
import shutil
import re
import time
from pathlib import Path
import tkinter as tk
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

from tkinter import ttk, filedialog, messagebox

from openpyxl import load_workbook

APP_NAME = "Horizon Chantier"

# =========================
# FICHIERS (STRICT / SANS REFONTE)
# =========================
PV_SOURCE = "soumission_chapelette__MAJ_PV_2026-01-29_13-52-57.xlsx"
PV_BACKUP = "soumission_chapelette__backup_avant_injectionPV_2026-01-29_13-52-57.xlsx"
PV_CORRIGE = "soumission_chapelette__MAJ_injection_CORRIGE_PV_2026-01-29_13-52-57.xlsx"
FICHIER_PV = PV_SOURCE


# =========================
# AppleScript helpers (Excel ouvert)
# =========================
def _osascript(script: str) -> str:
    p = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
    if p.returncode != 0:
        raise RuntimeError(p.stderr.strip() or "Erreur AppleScript")
    return (p.stdout or "").strip()


def _a1_to_col_row(a1: str) -> tuple[str, int]:
    """
    "C12" -> ("C", 12)
    """
    m = re.fullmatch(r"([A-Z]+)(\d+)", a1.strip().upper().replace("$", ""))
    if not m:
        raise ValueError(f"Adresse cellule invalide : {a1}")
    return m.group(1), int(m.group(2))


def _is_allowed_target(cell_a1: str) -> bool:
    """
    Sécurité absolue : n'autorise que
    - P3:P12  (matière)
    - P17:P26 (main-d'œuvre)
    """
    col, row = _a1_to_col_row(cell_a1)
    if col == "P" and 3 <= row <= 12:
        return True
    if col == "P" and 17 <= row <= 26:
        return True
    return False


def _normalize_excel_addr(addr: str) -> str:
    """
    Normalise:
    - "$P$3" -> "P3"
    - "P3"   -> "P3"
    """
    if not addr:
        return ""
    a = addr.strip().replace("$", "").upper()
    if re.fullmatch(r"[A-Z]+[0-9]+", a):
        return a
    return ""


def excel_find_first_empty_cell(workbook_hint: str, sheet_name: str, cell_range: str) -> str:
    """
    1ère cellule vide dans une plage (colonne unique) ex: P3:P12
    Retour "P3" ou "" si aucune.
    Robuste: retrouve le classeur par "contient" (Excel Mac change parfois le nom).
    """
    hint = (workbook_hint or "").strip()
    if not hint:
        return ""

    for _ in range(8):
        script = f'''
        tell application "Microsoft Excel"
            set wbRef to missing value
            repeat with w in workbooks
                try
                    set wn to (name of w as string)
                    if wn contains "{hint}" then
                        set wbRef to w
                        exit repeat
                    end if
                end try
            end repeat

            if wbRef is missing value then return ""

            tell wbRef
                activate
                try
                    tell worksheet "{sheet_name}"
                        activate
                        set rng to range "{cell_range}"
                        repeat with c in cells of rng
                            try
                                set v to value of c
                                set f to formula of c

                                if v is missing value then return (address of c) as string

                                try
                                    if (v as string) is "" then return (address of c) as string
                                end try

                                -- si formule et résultat vide
                                try
                                    if (f as string) is not "" then
                                        if (v as string) is "" then return (address of c) as string
                                    end if
                                end try

                            on error
                                return (address of c) as string
                            end try
                        end repeat
                    end tell
                end try
            end tell
        end tell
        return ""
        '''
        try:
            addr_raw = _osascript(script)
            addr = _normalize_excel_addr(addr_raw)
            if addr:
                return addr
        except Exception:
            pass

        time.sleep(0.2)

    return ""


def excel_set_cell_value(workbook_hint: str, sheet_name: str, cell_a1: str, value: str) -> None:
    """
    Écrit une valeur dans Excel (classeur déjà ouvert) sans fermer.
    Sécurité : n'écrit QUE dans P3:P12 ou P17:P26.
    Robuste: retrouve le classeur par "contient".
    """
    cell_a1 = (cell_a1 or "").strip().replace("$", "").upper()
    if not _is_allowed_target(cell_a1):
        raise ValueError(f"Sécurité: écriture interdite hors zones autorisées : {cell_a1}")

    hint = (workbook_hint or "").strip()
    if not hint:
        raise ValueError("Nom de classeur vide")

    safe_val = (value or "").replace('"', '\\"')

    script = f'''
    tell application "Microsoft Excel"
        set wbRef to missing value
        repeat with w in workbooks
            try
                set wn to (name of w as string)
                if wn contains "{hint}" then
                    set wbRef to w
                    exit repeat
                end if
            end try
        end repeat

        if wbRef is missing value then return

        tell wbRef
            if not (exists worksheet "{sheet_name}") then return
            tell worksheet "{sheet_name}"
                set value of range "{cell_a1}" to "{safe_val}"
            end tell
        end tell
        activate
    end tell
    '''
    _osascript(script)


# =========================
# Lecture bibliothèques (colonne A) - lecture seule
# =========================
def read_biblio_colA_openpyxl(pr_path: Path, sheet_name: str) -> list[str]:
    wb = load_workbook(pr_path, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Onglet introuvable : {sheet_name}")
        ws = wb[sheet_name]
        out: list[str] = []
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v is None:
                continue
            s = str(v).strip()
            if s:
                out.append(s)
        return out
    finally:
        wb.close()


# =========================
# Popup recherche (20 résultats)
# =========================
def popup_search_20(parent: tk.Tk, titre: str, items: list[str], on_pick) -> None:
    win = tk.Toplevel(parent)
    win.title(titre)
    win.geometry("780x520")
    win.transient(parent)
    win.grab_set()

    frm = ttk.Frame(win, padding=12)
    frm.pack(fill="both", expand=True)

    ttk.Label(frm, text=titre, font=("Helvetica", 14, "bold")).pack(anchor="w")

    q = tk.StringVar()
    entry = ttk.Entry(frm, textvariable=q)
    entry.pack(fill="x", pady=(10, 8))
    entry.focus_set()

    lb = tk.Listbox(frm, height=18)
    lb.pack(fill="both", expand=True)

    ttk.Label(frm, text="Tape 2 lettres. Double-clic ou Entrée pour valider.", foreground="#555").pack(
        anchor="w", pady=(8, 0)
    )

    current: list[str] = []

    def refresh(*_):
        text = q.get().strip().lower()
        lb.delete(0, tk.END)

        if not text:
            current[:] = items[:20]
        else:
            filtered = [s for s in items if text in s.lower()]
            current[:] = filtered[:20]

        for s in current:
            lb.insert(tk.END, s)

        if current:
            lb.selection_set(0)

    def choose(_event=None):
        sel = lb.curselection()
        if not sel:
            return
        picked = current[sel[0]]
        try:
            on_pick(picked)
        finally:
            win.destroy()

    q.trace_add("write", refresh)
    lb.bind("<Double-Button-1>", choose)
    win.bind("<Return>", choose)

    refresh()


# ============================================================
# Montant en lettres FR (euros + centimes) - inchangé
# ============================================================
_UNITS = ["zéro", "un", "deux", "trois", "quatre", "cinq", "six", "sept", "huit", "neuf"]
_TEENS = ["dix", "onze", "douze", "treize", "quatorze", "quinze", "seize", "dix-sept", "dix-huit", "dix-neuf"]
_TENS = ["", "", "vingt", "trente", "quarante", "cinquante", "soixante", "soixante", "quatre-vingt", "quatre-vingt"]


def _below_100(n: int) -> str:
    assert 0 <= n < 100
    if n < 10:
        return _UNITS[n]
    if 10 <= n < 20:
        return _TEENS[n - 10]

    ten = n // 10
    unit = n % 10

    if ten == 7:
        return "soixante" + ("-" + _below_100(10 + unit) if unit else "-dix")
    if ten == 9:
        return "quatre-vingt" + ("-" + _below_100(10 + unit) if unit else "-dix")

    base = _TENS[ten]

    if ten == 8 and unit == 0:
        return "quatre-vingts"
    if unit == 0:
        return base

    if unit == 1 and ten in (2, 3, 4, 5, 6):
        return f"{base} et un"

    return f"{base}-{_UNITS[unit]}"


def _below_1000(n: int) -> str:
    assert 0 <= n < 1000
    if n < 100:
        return _below_100(n)

    hundred = n // 100
    rest = n % 100

    if hundred == 1:
        head = "cent"
    else:
        head = f"{_UNITS[hundred]} cent"

    if rest == 0 and hundred > 1:
        return head + "s"
    if rest == 0:
        return head

    return head + " " + _below_100(rest)


def _int_to_words_fr(n: int) -> str:
    if n == 0:
        return "zéro"

    parts: list[str] = []

    billions = n // 1_000_000_000
    n %= 1_000_000_000
    millions = n // 1_000_000
    n %= 1_000_000
    thousands = n // 1000
    n %= 1000
    rest = n

    if billions:
        parts.append("un milliard" if billions == 1 else _below_1000(billions) + " milliards")

    if millions:
        parts.append("un million" if millions == 1 else _below_1000(millions) + " millions")

    if thousands:
        parts.append("mille" if thousands == 1 else _below_1000(thousands) + " mille")

    if rest:
        parts.append(_below_1000(rest))

    return " ".join(parts)


def montant_en_lettres_fr(valeur: float) -> str:
    v = round(float(valeur), 2)
    euros = int(v)
    centimes = int(round((v - euros) * 100))
    if centimes == 100:
        euros += 1
        centimes = 0

    euros_txt = _int_to_words_fr(euros)
    euro_label = "euro" if euros == 1 else "euros"

    if centimes == 0:
        return f"{euros_txt} {euro_label}"

    cent_txt = _int_to_words_fr(centimes)
    cent_label = "centime" if centimes == 1 else "centimes"
    return f"{euros_txt} {euro_label} et {cent_txt} {cent_label}"


# ============================================================
# PR / PV (inchangés)
# ============================================================
def open_pr(chantier_dir: str | Path):
    chantier = Path(chantier_dir)
    pr = chantier / "data" / "prix_de_revient.xlsx"
    if not pr.exists():
        raise FileNotFoundError(f"PR introuvable : {pr}")
    subprocess.run(["open", str(pr)], check=False)


def inject_pv(chantier_dir: str | Path) -> int:
    chantier = Path(chantier_dir)
    pr_path = chantier / "data" / "prix_de_revient.xlsx"
    pv_path = chantier / PV_SOURCE

    if not pr_path.exists():
        raise FileNotFoundError(f"PR introuvable : {pr_path}")
    if not pv_path.exists():
        raise FileNotFoundError(f"PV source introuvable : {pv_path}")

    pr_wb = load_workbook(pr_path, data_only=True)
    pr_ws = pr_wb["Chiffrage"] if "Chiffrage" in pr_wb.sheetnames else pr_wb.active


    # trouver la 1ère ligne d'article à partir de A3
    row_article = None
    for r in range(3, pr_ws.max_row + 1):
        v = pr_ws.cell(row=r, column=1).value  # colonne A
        if v is not None and str(v).strip() != "":
            row_article = r
            break

    if row_article is None:
        pr_wb.close()
        raise ValueError("Aucun ARTICLE trouvé dans la colonne A (à partir de A3).")

    article = pr_ws.cell(row=row_article, column=1).value
    designation = pr_ws.cell(row=row_article, column=2).value

    # on garde ton prix comme avant
    price = pr_ws["E32"].value

    pr_wb.close()



    if article is None or str(article) == "":
        raise ValueError("ARTICLE vide dans PR (A3).")
    if price is None or str(price) == "":
        raise ValueError("PRIX vide dans PR (E32).")

    article = str(article)
    designation = "" if designation is None else str(designation)

    price_f = float(price)
    price_2d = round(price_f, 2)
    letters = montant_en_lettres_fr(price_2d)

    wb = load_workbook(pv_path)
    ws = wb["PV"] if "PV" in wb.sheetnames else wb.active

    row_found = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        if str(v) == article:
            row_found = r
            break

    if row_found is None:
        wb.close()
        raise ValueError(f"Article introuvable dans PV source : '{article}'")

    ws.cell(row=row_found, column=2).value = designation

    g = ws.cell(row=row_found, column=7)
    g.value = price_2d
    g.number_format = "#,##0.00"

    ws.cell(row=row_found, column=8).value = letters

    wb.save(pv_path)
    wb.close()
    return 1


def sync_pv_to_copies(chantier_dir: str | Path) -> None:
    chantier = Path(chantier_dir)

    src = chantier / PV_SOURCE
    tgt_backup = chantier / PV_BACKUP
    tgt_corrige = chantier / PV_CORRIGE

    for p in (src, tgt_backup, tgt_corrige):
        if not p.exists():
            raise FileNotFoundError(f"Fichier introuvable : {p}")

    src_wb = load_workbook(src, data_only=True)
    src_ws = src_wb["PV"] if "PV" in src_wb.sheetnames else src_wb.active
    data = [row for row in src_ws.iter_rows(values_only=True)]
    src_wb.close()

    def _write(target: Path):
        wb = load_workbook(target)
        ws = wb["PV"] if "PV" in wb.sheetnames else wb.active

        for r_idx, row in enumerate(data, start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if cell.__class__.__name__ == "MergedCell":
                    continue
                cell.value = val

        wb.save(target)
        wb.close()

    _write(tgt_backup)
    _write(tgt_corrige)


# ---------------------------
# Dossiers (auto Chantier/Chantiers)
# ---------------------------
def dossier_base() -> Path:
    return Path.home() / "Desktop" / "Horizon_Chantier_Data"


def dossier_chantiers() -> Path:
    base = dossier_base()
    d1 = base / "Chantier"
    d2 = base / "Chantiers"
    if d1.exists():
        return d1
    if d2.exists():
        return d2
    d2.mkdir(parents=True, exist_ok=True)
    return d2
def ouvrir_doc(self, nom_fichier):
    sel = self.tree.selection()
    if not sel:
        messagebox.showwarning("Sélection", "Sélectionne un chantier.")
        return

    item = self.tree.item(sel[0])

    # 1) Nom sélectionné (text si présent, sinon values[0])
    chantier_sel = (item.get("text") or "").strip()
    if not chantier_sel:
        vals = item.get("values", [])
        chantier_sel = str(vals[0]).strip() if vals else ""

    base = dossier_chantiers()

    # 2) On cherche le dossier chantier qui existe vraiment
    dossier = base / chantier_sel
    if not dossier.exists():
        def norm(s):
            return s.lower().strip().replace(" ", "_").replace("-", "_")

        cible = norm(chantier_sel)
        dossier = next(
            (p for p in base.iterdir() if p.is_dir() and norm(p.name) == cible),
            None
        )

    if not dossier or not Path(dossier).exists():
        messagebox.showerror(
            "Introuvable",
            f"Dossier chantier introuvable pour :\n{chantier_sel}\n\nBase : {base}"
        )
        return

    chemin = Path(dossier) / nom_fichier

    if not chemin.exists():
        # Si on n'a pas trouvé le fichier, on tente l'autre dossier chantier
        # (ex: "soumission_chapelette" vs "Soumission Chapelette")
        def norm(s):
            return str(s).lower().strip().replace(" ", "_").replace("-", "_")

        cible = norm(chantier_sel)
        for p in base.iterdir():
            if p.is_dir() and norm(p.name) == cible:
                alt = p / nom_fichier
                if alt.exists():
                   chemin = alt
                   break

    if not chemin.exists():
        messagebox.showerror("Introuvable", f"Fichier absent :\n{chemin}")
        return

    if sys.platform == "darwin":
        subprocess.run(["open", str(chemin)], check=False)
    elif os.name == "nt":
        os.startfile(str(chemin))

def ouvrir_dossier(path: Path) -> None:
    try:
        subprocess.run(["open", str(path)], check=False)
    except Exception as e:
        messagebox.showerror("Erreur", f"Impossible d'ouvrir le dossier.\n{e}")


# ---------------------------
# Lecture chantier JSON
# ---------------------------
def lire_json(path: Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def ecrire_json(path: Path, data: dict) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def infos_chantier(chantier: dict) -> dict:
    return {
        "nom": chantier.get("nom", chantier.get("chantier", "")) or "",
        "client": chantier.get("client", "") or "",
        "etat": chantier.get("etat", chantier.get("type_travaux", "")) or "",
        "avancement": chantier.get("avancement", chantier.get("avancement_pct", 0)) or 0,
        "debut": chantier.get("date_debut", chantier.get("debut_travaux", "")) or "",
        "fin": chantier.get("date_fin", chantier.get("fin_prevue", chantier.get("date_fin_prevue", ""))) or "",
    }


# ---------------------------
# Fenêtre Bordereau (visu JSON)
# ---------------------------
def afficher_bordereau(parent: tk.Tk, chantier: dict, titre: str = "Bordereau") -> None:
    bord = chantier.get("bordereau", {})
    articles = bord.get("articles", {})

    if not articles:
        messagebox.showwarning("Bordereau", "Aucun bordereau dans ce chantier.")
        return

    win = tk.Toplevel(parent)
    win.title(titre)
    win.geometry("1200x650")

    cols = ("Article", "Libellé", "Unité", "Qté", "PU", "Total")

    tree = ttk.Treeview(win, columns=cols, show="headings")
    vsb = ttk.Scrollbar(win, orient="vertical", command=tree.yview)
    hsb = ttk.Scrollbar(win, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

    for c in cols:
        tree.heading(c, text=c)
        if c == "Libellé":
            tree.column(c, width=700, anchor="w")
        elif c == "Article":
            tree.column(c, width=160, anchor="w")
        elif c == "Unité":
            tree.column(c, width=80, anchor="w")
        else:
            tree.column(c, width=120, anchor="e")

    def v(x):
        return "" if x is None else x

    for article_id in sorted(articles.keys()):
        a = articles[article_id] or {}
        tree.insert(
            "",
            "end",
            values=(
                v(a.get("article_id", article_id)),
                v(a.get("libelle", "")),
                v(a.get("unite", "")),
                v(a.get("quantite", 0)),
                v(a.get("pu_vente", 0)),
                v(a.get("total_vente", 0)),
            ),
        )

    tree.grid(row=0, column=0, sticky="nsew")
    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")

    win.grid_rowconfigure(0, weight=1)
    win.grid_columnconfigure(0, weight=1)


# ---------------------------
# App
# ---------------------------
class HorizonChantierApp(tk.Tk):


    def __init__(self) -> None:
        super().__init__()
        self.title(APP_NAME)
        self.geometry("1400x800")
        self.minsize(1300, 950)

        self._build_ui()
        self.refresh_liste()

    def _build_ui(self) -> None:
        root = ttk.Frame(self, padding=18)
        root.pack(fill="both", expand=True)
        style = ttk.Style()
        style.configure("TitreBloc.TLabel",
                font=("Helvetica", 12, "bold"),
                foreground="#1f4e79")
                
        title = ttk.Label(root, text=APP_NAME, font=("Helvetica", 22, "bold"))
        title.pack(anchor="w")

        self.lbl_path = ttk.Label(root, text=f"Emplacement: {dossier_chantiers()}")
        self.lbl_path.pack(fill="x", pady=(4, 8))

        cols = ("Chantier", "Client", "État", "Avancement", "Début", "Fin prévue")

        zone = ttk.Frame(root)
        zone.pack(fill="both", expand=True)

        zone.columnconfigure(0, weight=0)
        zone.columnconfigure(1, weight=1)
        zone.columnconfigure(2, weight=0)
        zone.rowconfigure(0, weight=1)

        frame_gauche = ttk.Frame(zone)
        frame_gauche.grid(row=0, column=0, sticky="ns", padx=(0, 12))

        frame_centre = ttk.Frame(zone)
        frame_centre.grid(row=0, column=1, sticky="nsew")

        frame_droite = ttk.Frame(zone)
        frame_droite.grid(row=0, column=2, sticky="ns", padx=(12, 0))

        self.tree = ttk.Treeview(frame_centre, columns=cols, show="headings")
        for c in cols:
            self.tree.heading(c, text=c)
            if c == "Chantier":
                self.tree.column(c, width=260, anchor="w")
            elif c == "Client":
                self.tree.column(c, width=180, anchor="w")
            elif c == "État":
                self.tree.column(c, width=120, anchor="w")
            elif c == "Avancement":
                self.tree.column(c, width=110, anchor="e")
            else:
                self.tree.column(c, width=120, anchor="w")

        self.tree.pack(fill="both", expand=True)

        # -------------------------
        # GAUCHE
        # -------------------------

        tk.Label(frame_gauche, text="GESTION",
            font=("Helvetica", 14, "bold"),
            fg="#0052cc").pack(anchor="w", pady=(0, 2))
        ttk.Separator(frame_gauche).pack(fill="x", pady=(0, 8))

        ttk.Button(frame_gauche, text="📄 Ouvrir dans le logiciel", command=self.ouvrir_chantier).pack(fill="x", pady=3)
        ttk.Button(frame_gauche, text="➕ Nouveau chantier", command=self.nouveau_chantier).pack(fill="x", pady=3)
        ttk.Button(frame_gauche, text="✏️ Modifier", command=self.modifier_chantier).pack(fill="x", pady=3)
        ttk.Button(frame_gauche, text="🗑️ Supprimer", command=self.supprimer_chantier).pack(fill="x", pady=3)
        ttk.Button(frame_gauche, text="🔄 Rafraîchir", command=self.refresh_liste).pack(fill="x", pady=3)

        ttk.Separator(frame_gauche).pack(fill="x", pady=10)

        tk.Label(frame_gauche, text="DOCUMENTS",
         font=("Helvetica", 15, "bold"),
         fg="#0052cc").pack(anchor="w", pady=(2, 4))
        ttk.Separator(frame_gauche).pack(fill="x", pady=(0, 8))

        ttk.Button(frame_gauche, text="📁 Ce chantier", command=self.dossier_du_chantier).pack(fill="x", pady=3)

        ttk.Button(
            frame_gauche,
            text="🔎 Dossier administratif",
            command=lambda: __import__("subprocess").run([
                "open",
                f"/Users/senterre/Desktop/Horizon_Chantier_Data/Chantiers/{self.tree.item(self.tree.selection()[0])['values'][0]}/Administratif"
            ])
        ).pack(fill="x", pady=3)

        ttk.Button(
            frame_gauche,
            text="📄 Cahier des charges administratif",
            command=lambda: __import__("subprocess").Popen([
                "open",
                f"/Users/senterre/Desktop/Horizon_Chantier_Data/Chantiers/{self.tree.item(self.tree.selection()[0])['values'][0]}/Cahier_administratif.pdf"
            ])
        ).pack(fill="x", pady=3)

        ttk.Button(
           frame_gauche,
            text="🛠 Cahier des charges technique",
            command=lambda: __import__("subprocess").Popen([
                "open",
                f"/Users/senterre/Desktop/Horizon_Chantier_Data/Chantiers/{self.tree.item(self.tree.selection()[0])['values'][0]}/Cahier_technique.pdf"
            ])
        ).pack(fill="x", pady=3)

        ttk.Button(
            frame_gauche,
            text="📐 Postes / Métré",
            command=lambda: __import__("subprocess").run([
                "open",
                f"/Users/senterre/Desktop/Horizon_Chantier_Data/Chantiers/{self.tree.item(self.tree.selection()[0])['values'][0]}/Metre_detaille.pdf"
            ])
        ).pack(fill="x", pady=3)

        ttk.Button(
            frame_gauche,
            text="🦺 Plan de sécurité (PSS)",
            command=lambda: __import__("subprocess").Popen([
                "open",
                f"/Users/senterre/Desktop/Horizon_Chantier_Data/Chantiers/{self.tree.item(self.tree.selection()[0])['values'][0]}/PSS.pdf"
            ])
        ).pack(fill="x", pady=3)

        ttk.Button(
            frame_gauche,
            text="📄 Décompte intempéries",
            command=lambda: __import__("subprocess").run([
                "open",
                f"/Users/senterre/Desktop/Horizon_Chantier_Data/Chantiers/{self.tree.item(self.tree.selection()[0])['values'][0]}/Décompte_intempéries.doc"
            ])
        ).pack(fill="x", pady=3)

        # -------------------------
        # DROITE
        # -------------------------

        tk.Label(frame_droite, text="IMPORTER BORDEREAU",
         font=("Helvetica", 15, "bold"),
         fg="#0052cc").pack(anchor="w", pady=(2, 4))
        ttk.Separator(frame_droite).pack(fill="x", pady=(0, 8))

        ttk.Button(frame_droite, text="📥 Importer bordereau", command=self.importer_bordereau).pack(fill="x", pady=3)

        ttk.Separator(frame_droite).pack(fill="x", pady=10)

        tk.Label(frame_droite, text="EXECUTION",
         font=("Helvetica", 15, "bold"),
         fg="#0052cc").pack(anchor="w", pady=(2, 4))
        ttk.Separator(frame_droite).pack(fill="x", pady=(0, 8))

        ttk.Button(
            frame_droite,
            text="📅 Planning d’exécution",
            command=lambda: __import__("subprocess").Popen([
                "open",
                f"/Users/senterre/Desktop/Horizon_Chantier_Data/Chantiers/{self.tree.item(self.tree.selection()[0])['values'][0]}/Planning_execution.xlsx"
            ])
        ).pack(fill="x", pady=3)

        ttk.Separator(frame_droite).pack(fill="x", pady=10)

        tk.Label(frame_droite, text="CALCULS",
         font=("Helvetica", 15, "bold"),
         fg="#0052cc").pack(anchor="w", pady=(2, 4))
        ttk.Separator(frame_droite).pack(fill="x", pady=(0, 8))

        ttk.Button(
            frame_droite,
            text="🔎 Coût de la sécurité",
            command=lambda: __import__("subprocess").Popen([
                "open",
                f"/Users/senterre/Desktop/Horizon_Chantier_Data/Chantiers/{self.tree.item(self.tree.selection()[0])['values'][0]}/Cout_securite.xlsx"
            ])
        ).pack(fill="x", pady=3)

        ttk.Button(
            frame_droite,
            text="📐 Formule de révision",
            command=lambda: __import__("subprocess").run([
                "open",
                f"/Users/senterre/Desktop/Horizon_Chantier_Data/Chantiers/{self.tree.item(self.tree.selection()[0])['values'][0]}/Formule_révision.xlsm"
            ], check=True)
        ).pack(fill="x", pady=3)

        ttk.Button(
            frame_droite,
            text="💰 Prix de revient",
            command=lambda: __import__("subprocess").run([
                "open",
                f"/Users/senterre/Desktop/Horizon_Chantier_Data/Chantiers/{self.tree.item(self.tree.selection()[0])['values'][0]}/data/prix_de_revient.xlsx"
            ])
        ).pack(fill="x", pady=3)

        ttk.Button(frame_droite, text="🔍 Vérifier PR", command=self.verifier_pr).pack(fill="x", pady=3)
        ttk.Button(frame_droite, text="⚙️ Calcul / Reca", command=self.calcul_reca).pack(fill="x", pady=3)

        ttk.Separator(frame_droite).pack(fill="x", pady=10)

        tk.Label(frame_droite, text="ETAT D'AVANCEMENT",
         font=("Helvetica", 15, "bold"),
         fg="#0052cc").pack(anchor="w", pady=(2, 4))
        ttk.Separator(frame_droite).pack(fill="x", pady=(0, 8))

        ttk.Button(
            frame_droite,
            text="📊 État d'avancement",
            command=lambda: subprocess.Popen([
                "open",
                next(Path(
                    f"/Users/senterre/Desktop/Horizon_Chantier_Data/Chantiers/{self.tree.item(self.tree.selection()[0])['values'][0]}"
                ).glob("Etat_avancement*.xlsm")).as_posix()
            ])
        ).pack(fill="x", pady=3)

        ttk.Button(frame_droite, text="📊 Calcul état", command=self.calcul_etat_avancement).pack(fill="x", pady=3)
        ttk.Button(frame_droite, text="♻️ Clôturer état", command=self.mise_a_zero_etat).pack(fill="x", pady=3)

        ttk.Separator(frame_droite).pack(fill="x", pady=10)

        tk.Label(frame_droite, text="PILOTAGE",
         font=("Helvetica", 15, "bold"),
         fg="#0052cc").pack(anchor="w", pady=(2, 4))
        ttk.Separator(frame_droite).pack(fill="x", pady=(0, 8))

        ttk.Button(frame_droite, text="🧭 Pilotage", command=self.pilotage_chantier).pack(fill="x", pady=3)
        ttk.Button(frame_droite, text="📊 Pilotage délai", command=self.pilotage_delai).pack(fill="x", pady=3)
        ttk.Button(frame_droite, text="⏱ Rendements", command=self.rendements_chantier).pack(fill="x", pady=3)
       
        self.tree.bind("<Double-1>", lambda e: self.ouvrir_chantier())

    def calcul_etat_avancement(self):
        chemin = next(
            Path(
                f"/Users/senterre/Desktop/Horizon_Chantier_Data/Chantiers/{self.tree.item(self.tree.selection()[0])['values'][0]}"
            ).glob("Etat_avancement*.xlsm")
        ).as_posix()

        feuille = "Bordereau"

        wb = load_workbook(chemin, keep_vba=True)
        ws = wb[feuille]

        def nombre(valeur):
            if valeur in (None, "", "-", "—"):
                return 0
            try:
                return float(str(valeur).replace("€", "").replace(" ", "").replace(",", "."))
            except:
                return 0

        def ligne_bordereau_valide(ligne):
            c = ws[f"C{ligne}"].value
            j = ws[f"J{ligne}"].value
            l = ws[f"L{ligne}"].value
            p = ws[f"P{ligne}"].value
            q = ws[f"Q{ligne}"].value

            return any(v not in (None, "", "-", "—") for v in [c, j, l, p, q])

        def derniere_ligne_bordereau():
            derniere = 24
            lignes_vides = 0

            for ligne in range(24, ws.max_row + 1):
                if ligne_bordereau_valide(ligne):
                    derniere = ligne
                    lignes_vides = 0
                else:
                    lignes_vides += 1

                    # dès qu'on a plusieurs lignes vides d'affilée,
                    # on considère que le bordereau est terminé
                    if lignes_vides >= 5:
                        break

            return derniere

        fin = derniere_ligne_bordereau()

        for ligne in range(24, fin + 1):
            if not ligne_bordereau_valide(ligne):
                continue

            j = nombre(ws[f"J{ligne}"].value)
            l = nombre(ws[f"L{ligne}"].value)
            p = nombre(ws[f"P{ligne}"].value)
            q = nombre(ws[f"Q{ligne}"].value)

            r = p + q
            s = 0 if j == 0 else r / j
            t = r * l
            u = q
            v = l
            w = u * v

            ws[f"R{ligne}"] = r
            ws[f"S{ligne}"] = s
            ws[f"T{ligne}"] = t
            ws[f"U{ligne}"] = u
            ws[f"V{ligne}"] = v
            ws[f"W{ligne}"] = w

        

        wb.save(chemin)
        print("Etat recalculé")


       

    def mise_a_zero_etat(self):
        chemin = next(
            Path(
                f"/Users/senterre/Desktop/Horizon_Chantier_Data/Chantiers/{self.tree.item(self.tree.selection()[0])['values'][0]}"
            ).glob("Etat_avancement*.xlsm")
        ).as_posix()

        feuille = "Bordereau"

        wb = load_workbook(chemin, keep_vba=True)
        ws = wb[feuille]

        ligne = 24

        while ligne <= ws.max_row:
            ws[f"P{ligne}"] = ws[f"R{ligne}"].value
            ws[f"Q{ligne}"] = 0
            ligne += 1

        wb.save(chemin)
        self.calcul_etat_avancement()
        print("Clôture état effectuée")

    def verifier_pr(self):
        p = self._chemin_chantier_selectionne()
        if not p:
            messagebox.showwarning("PR", "Sélectionne un chantier dans la liste.")
            return

        dossier_chantier = p.parent / p.stem
        chemin = dossier_chantier / "data" / "prix_de_revient.xlsx"

        import os
        from datetime import datetime

        if not chemin.exists():
            messagebox.showwarning(
                "Vérifier PR",
                "Le fichier prix_de_revient.xlsx est introuvable."
            )
            return

        timestamp = os.path.getmtime(chemin)
        date_modif = datetime.fromtimestamp(timestamp).strftime("%d/%m/%Y %H:%M:%S")

        messagebox.showinfo(
            "Date du PR",
            f"Dernier enregistrement du PR :\n{date_modif}\n\n"
            "⚠️ Si tu viens de modifier le PR :\n"
            "➡️ fais CMD + S dans Excel\n"
            "➡️ puis clique sur Vérifier PR"
        )
        
        wb = load_workbook(chemin, data_only=True)
        ws = wb["Chiffrage"]

        def nombre(v):
            if v in (None, "", "-", "—"):
                return 0
            try:
                return float(str(v).replace("€", "").replace(" ", "").replace(",", "."))
            except:
                return 0

        erreurs = []

        # MATIÈRE
        for ligne in range(3, 13):
            i = nombre(ws[f"I{ligne}"].value)
            j = nombre(ws[f"J{ligne}"].value)
            h = nombre(ws[f"H{ligne}"].value)

            attendu_k = i * (1 + j)
            attendu_l = h * attendu_k

            k = nombre(ws[f"K{ligne}"].value)
            l = nombre(ws[f"L{ligne}"].value)

            if abs(k - attendu_k) > 0.01:
                erreurs.append(f"Ligne {ligne} PU corrigé faux")

            if abs(l - attendu_l) > 0.01:
                erreurs.append(f"Ligne {ligne} Total matière faux")

        # MAIN-D’ŒUVRE
        for ligne in range(16, 26):
            c = nombre(ws[f"C{ligne}"].value)
            h = nombre(ws[f"H{ligne}"].value)
            j = nombre(ws[f"J{ligne}"].value)

            attendu_i = c * h
            attendu_k = attendu_i * j

            i = nombre(ws[f"I{ligne}"].value)
            k = nombre(ws[f"K{ligne}"].value)

            if abs(i - attendu_i) > 0.01:
                erreurs.append(f"Ligne {ligne} Heures fausses")

            if abs(k - attendu_k) > 0.01:
                erreurs.append(f"Ligne {ligne} Coût MO faux")

        if erreurs:
            messagebox.showwarning("PR", "\n".join(erreurs))
        else:
            messagebox.showinfo("PR", "✔ Tous les calculs sont corrects")

        # 🔒 Sécurisation automatique des formules
        wb = load_workbook(chemin)
        ws = wb["Chiffrage"]

        # tout déverrouiller
        for row in ws.iter_rows():
            for cell in row:
                cell.protection = cell.protection.copy(locked=False)

        # verrouiller uniquement les formules
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.protection = cell.protection.copy(locked=True)

        # activer la protection
        ws.protection.sheet = True
        ws.protection.password = "1234"

        wb.save(chemin)
            
    def pilotage_chantier(self):
        chemin = next(
            Path(
                f"/Users/senterre/Desktop/Horizon_Chantier_Data/Chantiers/{self.tree.item(self.tree.selection()[0])['values'][0]}"
            ).glob("Etat_avancement*.xlsm")
        ).as_posix()

        wb = load_workbook(chemin, data_only=True)
        ws = wb["Bordereau"]
        

        def nombre(val):
            if val in (None, "", "-", "—"):
                return 0
            try:
                return float(str(val).replace("€", "").replace(" ", "").replace(",", "."))
            except:
                return 0

        def texte(val):
            return str(val or "").strip().lower()

        def trouver_ligne_libelle(mots):
            for row in range(ws.max_row, 1, -1):
                for col in range(1, ws.max_column + 1):
                    val = texte(ws.cell(row=row, column=col).value)
                    if not val:
                        continue
                    if all(mot in val for mot in mots):
                        return row, col
            return None, None

        def premiere_valeur_a_droite(row, col_depart, max_ecart=3):
            if not row or not col_depart:
                return 0

            for col in range(col_depart + 1, min(col_depart + max_ecart + 1, ws.max_column + 1)):
                val = ws.cell(row=row, column=col).value
                num = nombre(val)
                if num != 0:
                    return num
                return 0        
      
        ws_rev = wb["Révision"]
        ws_glob = wb["Revision_global"]
        ws_avenants = wb["Avenants"]

        revision = nombre(ws_rev["F32"].value)
        revision_cumulee = nombre(ws_glob["D62"].value)

        avenants_plus = nombre(ws_avenants["D62"].value)
        avenants_moins = nombre(ws_avenants["E62"].value)

        print("revision =", revision)
        print("revision_cumulee =", revision_cumulee)
        print("avenants_plus =", avenants_plus)
        print("avenants_moins =", avenants_moins)


        ligne_soumission, col_soumission = trouver_ligne_libelle(["soumission", "tva"])
        ligne_mois, col_mois = trouver_ligne_libelle(["total", "mois", "tva"])
        ligne_execute, col_execute = trouver_ligne_libelle(["exécuté"])

        if not ligne_execute:
            ligne_execute, col_execute = trouver_ligne_libelle(["execute"])


        montant_soumission = premiere_valeur_a_droite(ligne_soumission, col_soumission)
        total_mois = premiere_valeur_a_droite(ligne_mois, col_mois)
        total_realise = premiere_valeur_a_droite(ligne_execute, col_execute)

        print("ligne_mois =", ligne_mois, "col_mois =", col_mois)
        print("total_mois =", total_mois)
        print("revision =", revision)

       

        # 🔥 CALCULS PROPRES
        total_marche = montant_soumission + avenants_plus - avenants_moins
        realise_periode = total_mois 

        avancement_periode = 0 if total_marche == 0 else (realise_periode / total_marche) * 100
        avancement = 0 if total_marche == 0 else (total_realise / total_marche) * 100

        print("montant_soumission =", montant_soumission)
        print("total_mois =", total_mois)
        print("total_realise =", total_realise)
        print("realise_periode =", realise_periode)
        print("total_marche =", total_marche)
        print("avancement_periode =", avancement_periode)
        print("avancement =", avancement)


        from datetime import datetime

        date_du_jour = datetime.now().strftime("%d/%m/%Y")
        nom_chantier = self.tree.item(self.tree.selection()[0])['values'][0]

        total_marche = nombre(ws_avenants["E65"].value)
        realise_periode = total_mois 
        avancement_periode = 0 if total_marche == 0 else (realise_periode / total_marche) * 100
        avancement = 0 if total_marche == 0 else (total_realise / total_marche) * 100

        texte_popup = f"""PILOTAGE CHANTIER

       
    Chantier : {nom_chantier}
    Date     : {date_du_jour}

    Montant soumission              : {montant_soumission:,.2f} €
    Avenants cumulés en plus        : {avenants_plus:,.2f} €
    Avenants cumulés en moins       : {avenants_moins:,.2f} €
    Marché corrigé                  : {total_marche:,.2f} €

    Montant du mois à facturer      : {total_mois:,.2f} €
    Révision période                : {revision:,.2f} €
    Révision cumulée                : {revision_cumulee:,.2f} €

    Montant exécuté période         : {realise_periode:,.2f} €
    Réalisé cumulé                  : {total_realise:,.2f} €
    
    Avancement période              : {avancement_periode:.1f} %
    Avancement cumulé               : {avancement:.1f} %
    """
        
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Image
        from reportlab.lib.units import mm

        pdf_path = chemin.replace("Etat_avancement_02.xlsm", "pilotage_chantier.pdf")

        doc = SimpleDocTemplate(pdf_path)
        styles = getSampleStyleSheet()
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER

        style_titre = ParagraphStyle(
            "Titre",
            parent=styles["Title"],
            alignment=TA_CENTER,
            fontSize=22,
            textColor=colors.HexColor("#1F4E79")
        )

        elements = []
        logo_path = "/Users/senterre/Desktop/Horizon_Chantier_Data/logo_jt_bati.png"

        try:
            logo = Image(logo_path, width=35*mm, height=15*mm)
            logo.hAlign = "LEFT"
            elements.append(logo)
            elements.append(Spacer(1, 8))
        except Exception as e:
            print("Erreur logo PDF :", e)
        

        # 🔷 TITRE (un seul)
        titre = Table([["PILOTAGE CHANTIER"]], colWidths=[140*mm], rowHeights=[22*mm])
        titre.hAlign = "CENTER"

        titre.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.white),
            
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 22),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#1F4E79")),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))

        elements.append(titre)
        elements.append(Spacer(1, 1))

        # 🔹 INFOS CHANTIER
        info_data = [
            ["Chantier", nom_chantier],
            ["Date", date_du_jour],
        ]

        info_table = Table(info_data, colWidths=[50*mm, 90*mm])
        info_table.hAlign = "CENTER"

        info_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 13),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            
        ]))

        elements.append(info_table)
        elements.append(Spacer(1, 20))

        # 🔸 TABLEAU PROPRE
        
       
        total_marche = nombre(ws_avenants["E65"].value)
        realise_periode = total_mois + revision
        avancement_periode = 0 if total_marche == 0 else (realise_periode / total_marche) * 100
        avancement = 0 if total_marche == 0 else (total_realise / total_marche) * 100
        
        data = [
        ["Indicateur", "Montant"],
        ["Montant soumission", f"{montant_soumission:,.2f} €"],
        ["Avenants cumulés en plus", f"{avenants_plus:,.2f} €"],
        ["Avenants cumulés en moins", f"{avenants_moins:,.2f} €"],
        ["Marché corrigé", f"{total_marche:,.2f} €"],
        ["Montant du mois à facturer", f"{total_mois:,.2f} €"],
        ["Révision période", f"{revision:,.2f} €"],
        ["Révision cumulée", f"{revision_cumulee:,.2f} €"],
        ["Montant chantier exécuté période", f"{total_mois:,.2f} €"],
        ["Montant chantier exécuté cumulé", f"{total_realise:,.2f} €"],
        ["Avancement période", f"{avancement_periode:.1f} %"],
        ["Avancement cumulé", f"{avancement:.1f} %"],
    ]  

        table = Table(data, colWidths=[90*mm, 50*mm])
        table.hAlign = "CENTER"

        table.setStyle(TableStyle([

            # 🔵 En-tête
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),

            # ⚪ Fond général
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),

            # 📐 Grille
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),

            # 🔤 Polices
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica"),
            ("FONTNAME", (1, 1), (1, -1), "Helvetica-Bold"),

            # 🔠 Taille
            ("FONTSIZE", (0, 0), (-1, -1), 13),

            # 📍 Alignements
            ("ALIGN", (0, 0), (-1, 0), "LEFT"),
            ("ALIGN", (0, 1), (0, -1), "LEFT"),
            ("ALIGN", (1, 1), (1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

            # 📦 Espacements
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),

            # 🟡 Mois en cours à facturer
            ("BACKGROUND", (0, 5), (1, 5), colors.HexColor("#FFFF00")),

            # Texte gauche (label)
            ("TEXTCOLOR", (0, 5), (0, 5), colors.black),
            ("FONTNAME", (0, 5), (0, 5), "Helvetica-Bold"),

            # Texte droite (montant) 👇
            ("TEXTCOLOR", (1, 5), (1, 5), colors.black),
            ("FONTNAME", (1, 5), (1, 5), "Helvetica-Bold"),

            # 🔴 Avenants cumulés (visible sans agresser)
            ("BACKGROUND", (0, 2), (1, 2), colors.HexColor("#FADBD8")),
            ("TEXTCOLOR", (0, 2), (1, 2), colors.black),
            ("FONTNAME", (0, 2), (1, 2), "Helvetica-Bold"),

            # 🔷 Révision chantier (fond bleu ciel + texte noir gras)
            ("BACKGROUND", (0, 4), (1, 4), colors.HexColor("#D6EAF8")),
            ("TEXTCOLOR", (1, 4), (1, 4), colors.black),
            ("FONTNAME", (1, 4), (1, 4), "Helvetica-Bold"),

            # 🔵 Marché total
            ("TEXTCOLOR", (1, 3), (1, 3), colors.HexColor("#2E86C1")),

            # 🟢 Réalisé cumulé
            ("TEXTCOLOR", (1, 6), (1, 6), colors.HexColor("#27AE60")),
            ("FONTNAME", (1, 3), (1, 3), "Helvetica-Bold"),  # Marché total
            ("FONTNAME", (1, 6), (1, 6), "Helvetica-Bold"),  # Réalisé cumulé
            ("FONTNAME", (0, 7), (1, 7), "Helvetica-Bold"),
        ]))

        elements.append(table)

        doc.build(elements)

        __import__("subprocess").run(["open", pdf_path])

        
        messagebox.showinfo("Pilotage chantier", texte_popup)

    def pilotage_delai(self):
        import os
        from openpyxl import load_workbook
        from tkinter import messagebox
        from datetime import datetime
        from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Image
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER

        def nombre(valeur):
            try:
                if valeur is None:
                    return 0
                if isinstance(valeur, str):
                    valeur = valeur.replace("%", "").replace("€", "").replace(" ", "").replace(",", ".").strip()
                    if valeur in ("", "-", "—"):
                        return 0
                return float(valeur)
            except:
                return 0

        try:
            if not self.tree.selection():
                messagebox.showwarning("Attention", "Veuillez sélectionner un chantier.")
                return

            nom_chantier = self.tree.item(self.tree.selection()[0])["values"][0]
            dossier_chantier = f"/Users/senterre/Desktop/Horizon_Chantier_Data/Chantiers/{nom_chantier}"
            fichier_delai = os.path.join(dossier_chantier, "Delai.xlsx")

            if not os.path.exists(fichier_delai):
                messagebox.showerror("Erreur", f"Fichier introuvable :\n{fichier_delai}")
                return

            wb = load_workbook(fichier_delai, data_only=True)
            ws = wb.active

            delai_soumission = nombre(ws["O64"].value)
            jours_consommes = nombre(ws["O66"].value)
            jours_plus = nombre(ws["O68"].value)
            jours_moins = nombre(ws["O70"].value)
            intemperies = nombre(ws["O76"].value)
            conges = nombre(ws["O77"].value)

            delai_consomme = nombre(ws["O78"].value)
            heures_consommees = nombre(ws["O79"].value)
            delai_corrige = nombre(ws["O80"].value)
            delai_corrige_heures = nombre(ws["O81"].value)
            ecart_jours_plus = nombre(ws["O82"].value)
            ecart_jours_moins = nombre(ws["O83"].value)
            ecart_pct = nombre(ws["O84"].value)

            

            date_du_jour = datetime.now().strftime("%d/%m/%Y")

            
            texte_popup = f"""PILOTAGE DÉLAI

        Chantier : {nom_chantier}
        Date     : {date_du_jour}

        Délai soumission               : {delai_soumission:.2f} j
        Jours consommé                 : {jours_consommes:.2f} j
        Jours accordés avenant plus    : {jours_plus:.2f} j
        Jours accordés avenant moins   : {jours_moins:.2f} j
        Intempéries                    : {intemperies:.2f} j
        Congé/ferrier/compensatoire    : {conges:.2f} j

        Délai consommé                 : {delai_consomme:.2f} j
        Heures consommées              : {heures_consommees:.2f} h
        Délai corrigé                  : {delai_corrige:.2f} j
        Délai corrigé en heures        : {delai_corrige_heures:.2f} h
        Écart en jour en plus          : {ecart_jours_plus:.2f} j
        Écart en jour en moins         : {ecart_jours_moins:.2f} j
        Écart en %                     : {ecart_pct:.1f} %
        """

            pdf_path = os.path.join(dossier_chantier, "pilotage_delai.pdf")

            doc = SimpleDocTemplate(pdf_path)
            styles = getSampleStyleSheet()

            style_titre = ParagraphStyle(
                "Titre",
                parent=styles["Title"],
                alignment=TA_CENTER,
                fontSize=22,
                textColor=colors.HexColor("#1F4E79")
            )

            elements = []
            logo_path = "/Users/senterre/Desktop/Horizon_Chantier_Data/logo_jt_bati.png"

            try:
                logo = Image(logo_path, width=35*mm, height=15*mm)
                logo.hAlign = "LEFT"
                elements.append(logo)
                elements.append(Spacer(1, 8))
            except Exception as e:
                print("Erreur logo PDF :", e)

            titre = Table([["PILOTAGE DÉLAI"]], colWidths=[140*mm], rowHeights=[22*mm])
            titre.hAlign = "CENTER"

            titre.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 22),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#1F4E79")),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]))

            elements.append(titre)
            elements.append(Spacer(1, 1))

            info_data = [
                ["Chantier", nom_chantier],
                ["Date", date_du_jour],
            ]

            info_table = Table(info_data, colWidths=[50*mm, 90*mm])
            info_table.hAlign = "CENTER"

            info_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 13),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]))

            elements.append(info_table)
            elements.append(Spacer(1, 20))

            data = [
                ["Indicateur", "Valeur"],
                ["Délai soumission", f"{delai_soumission:.0f} j"],
                ["Jours consommé", f"{jours_consommes:.0f} j"],
                ["Jours accordés en plus", f"{jours_plus:.0f} j"],
                ["Jours accordés en moins", f"{jours_moins:.0f} j"],
                ["Intempéries", f"{intemperies:.0f} j"],
                ["Congé/ferrier/compensatoire", f"{conges:.0f} j"],
                ["Délai consommé", f"{delai_consomme:.0f} j"],
                ["Délai corrigé", f"{delai_corrige:.0f} j"],
                ["Ecart en %", f"{ecart_pct:.1f} %"],
            ]

            table = Table(data, colWidths=[90*mm, 50*mm])
            table.hAlign = "CENTER"

            table.setStyle(TableStyle([

                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),

                ("BACKGROUND", (0, 1), (-1, -1), colors.white),

                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),

                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (0, -1), "Helvetica"),
                ("FONTNAME", (1, 1), (1, -1), "Helvetica-Bold"),

                ("FONTSIZE", (0, 0), (-1, -1), 13),

                ("ALIGN", (0, 0), (-1, 0), "LEFT"),
                ("ALIGN", (0, 1), (0, -1), "LEFT"),
                ("ALIGN", (1, 1), (1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),

                # Jour accordes = équivalent avenants cumulés
                ("BACKGROUND", (0, 3), (1, 3), colors.HexColor("#FADBD8")),
                ("TEXTCOLOR", (0, 3), (1, 3), colors.black),
                ("FONTNAME", (0, 3), (1, 3), "Helvetica-Bold"),

                # Intempéries = équivalent révision chantier
                ("BACKGROUND", (0, 4), (1, 4), colors.HexColor("#D6EAF8")),
                ("TEXTCOLOR", (1, 4), (1, 4), colors.black),
                ("FONTNAME", (1, 4), (1, 4), "Helvetica-Bold"),

                # Délai consommé = valeur bleue
                ("TEXTCOLOR", (1, 6), (1, 6), colors.HexColor("#2E86C1")),

                # Délai soumission = noir + gras
                ("TEXTCOLOR", (1, 1), (1, 1), colors.black),
                ("FONTNAME", (1, 1), (1, 1), "Helvetica-Bold"),

                # Délai corrigé = jaune + noir + gras (équivalent mois chantier)
                ("BACKGROUND", (0, 7), (1, 7), colors.HexColor("#FFFF00")),
                ("TEXTCOLOR", (0, 7), (1, 7), colors.black),
                ("FONTNAME", (0, 7), (1, 7), "Helvetica-Bold"),


                # Ecart en % = ligne forte comme avancement
                ("FONTNAME", (0, 8), (1, 8), "Helvetica-Bold"),
            ]))

            elements.append(table)

            doc.build(elements)

            __import__("subprocess").run(["open", pdf_path])

            messagebox.showinfo("Pilotage délai", texte_popup)

        except Exception as e:
            messagebox.showerror("Erreur pilotage délai", str(e))

    def rendements_chantier(self):
        import os, glob, subprocess
        import matplotlib.pyplot as plt
        from openpyxl import load_workbook
        import matplotlib.patheffects as pe

        def nombre(valeur):
            if valeur is None:
                return 0
            if isinstance(valeur, (int, float)):
                return float(valeur)
            txt = str(valeur).strip().replace("€", "").replace(" ", "").replace(",", ".")
            if txt in ("", "-", "—"):
                return 0
            try:
                return float(txt)
            except:
                return 0

        selection = self.tree.selection()
        if not selection:
                return

        nom = self.tree.item(selection[0])['values'][0]
        dossier = f"/Users/senterre/Desktop/Horizon_Chantier_Data/Chantiers/{nom}"

        fichiers = glob.glob(os.path.join(dossier, "Etat_avancement*.xlsm"))
        if not fichiers:
                return

        wb = load_workbook(fichiers[0], data_only=True)
        ws = wb["Bordereau"]

        total_realise = 0
        total_mois = 0

        for row in range(24, ws.max_row + 1):
            total_realise += nombre(ws[f"T{row}"].value)
            total_mois += nombre(ws[f"W{row}"].value)

        jours = 20
        prod_jour = total_mois / jours if jours else 0
        part = (total_mois / total_realise * 100) if total_realise else 0

        # -------- GRAPHIQUE --------
        chemin = os.path.join(dossier, "rendement.png")

        fig = plt.figure(figsize=(12, 7), facecolor="white")

        fig.text(0.05, 0.92, "RENDEMENT CHANTIER", fontsize=20, fontweight="bold")
        fig.text(0.05, 0.88, f"Chantier : {nom}", fontsize=11)

        ax = fig.add_axes([0.08, 0.2, 0.55, 0.6])

        labels = ["Mois (€)", "Jour (€)", "% Mois"]
        valeurs = [total_mois, prod_jour, part]
        couleurs = ["#2E8B57", "#1F4E78", "#E69F00"]

        barres = ax.bar(labels, valeurs, color=couleurs)
        ax.set_ylim(0, max(valeurs) * 1.3 if max(valeurs) > 0 else 100)
        ax.axhline(0, linewidth=1, color='black')
        if max(valeurs) == 0:
            ax.text(1, 50, "Aucune production ce mois",
                ha='center', va='center',
                fontsize=12, color='gray')


        
        for b, v in zip(barres, valeurs):
            ax.text(b.get_x()+b.get_width()/2, v, f"{v:,.2f}",
                    
                    ha="center", va="bottom", fontsize=11,
                    color="white",
                    path_effects=[pe.withStroke(linewidth=2, foreground="black")])

        ax.set_title("Production", fontsize=14, fontweight="bold")
        ax.grid(axis="y", alpha=0.3)

        # -------- RESUME --------
        ax2 = fig.add_axes([0.7, 0.25, 0.25, 0.5])
        ax2.axis("off")

        ax2.text(0, 0.8, f"{total_mois:,.2f} €",
                 fontsize=16, fontweight="bold", color=couleurs[0],
                 path_effects=[pe.withStroke(linewidth=2, foreground="black")])

        ax2.text(0, 0.6, f"{prod_jour:,.2f} €",
                 fontsize=16, fontweight="bold", color=couleurs[1],
                 path_effects=[pe.withStroke(linewidth=2, foreground="black")])

        ax2.text(0, 0.4, f"{part:.1f} %",
                 fontsize=16, fontweight="bold", color=couleurs[2],
                 path_effects=[pe.withStroke(linewidth=2, foreground="black")])
        

        # -------- LECTURE --------
        if part < 5:
            txt = "Faible activité"
            col = "#C00000"
        elif part < 15:
            txt = "Activité normale"
            col = "#BF9000"
        else:
            txt = "Bon rendement"
            col = "#38761D"

        ax2.text(0, 0.1, txt, fontsize=13, color=col, fontweight="bold")

        plt.savefig(chemin, dpi=200, bbox_inches="tight")
        plt.close()

        subprocess.run(["open", chemin])

    def dossier_du_chantier(self):
        ...
    def _chemin_chantier_selectionne(self) -> Path | None:
        sel = self.tree.selection()
        if not sel:
            return None
        return Path(sel[0])

    def refresh_liste(self) -> None:
        self.lbl_path.config(text=f"Emplacement: {dossier_chantiers()}")
        for item in self.tree.get_children():
            self.tree.delete(item)

        d = dossier_chantiers()
        fichiers = sorted(d.glob("*.json"))

        for p in fichiers:
            try:
                chantier = lire_json(p)
                info = infos_chantier(chantier)
                av = info["avancement"]
                try:
                    av_txt = f"{float(av):.0f}%"
                except Exception:
                    av_txt = str(av)

                self.tree.insert(
                    "",
                    "end",
                    iid=str(p),
                    values=(info["nom"], info["client"], info["etat"], av_txt, info["debut"], info["fin"]),
                )
            except Exception:
                continue

    def _pr_path_and_hint(self) -> tuple[Path, str]:
        p = self._chemin_chantier_selectionne()
        if not p:
            raise ValueError("Sélectionne un chantier dans la liste.")
        dossier_chantier = p.parent / p.stem
        pr_path = dossier_chantier / "data" / "prix_de_revient.xlsx"
        if not pr_path.exists():
            raise FileNotFoundError(f"PR introuvable : {pr_path}")
        hint = pr_path.name
        return pr_path, hint

    def importer_bordereau(self) -> None:
        p = self._chemin_chantier_selectionne()
        if not p:
            messagebox.showwarning("Bordereau", "Sélectionne un chantier dans la liste.")
            return

        xlsx_path = filedialog.askopenfilename(
            title="Choisir le bordereau VSO (Excel)",
            filetypes=[("Excel", "*.xlsx;*.xlsm")],
        )
        if not xlsx_path:
            return

        dossier_chantier = p.parent / p.stem
        bordereau_xlsx = dossier_chantier / FICHIER_PV
        try:
            shutil.copy2(xlsx_path, bordereau_xlsx)
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de copier le bordereau dans le chantier.\n{e}")
            return

        try:
            from app.bordereau import importer_bordereau_vso_xlsx

        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de charger bordereau.py\n{e}")
            return

        try:
            chantier = lire_json(p)
            chantier = importer_bordereau_vso_xlsx(xlsx_path, chantier)
            ecrire_json(p, chantier)
            self.refresh_liste()
            afficher_bordereau(self, chantier, "Bordereau (JSON)")
            messagebox.showinfo("Bordereau", f"Bordereau importé.\nCopie locale: {bordereau_xlsx}")
        except Exception as e:
            messagebox.showerror("Erreur", f"Import bordereau impossible.\n{e}")

    def prix_revient(self) -> None:
        p = self._chemin_chantier_selectionne()
        if not p:
            messagebox.showwarning("Prix de revient", "Sélectionne un chantier dans la liste.")
            return
        dossier_chantier = p.parent / p.stem
        try:
            open_pr(dossier_chantier)
        except Exception as e:
            messagebox.showerror("Prix de revient", str(e))

    # ======================================================
    # ✅ OBJECTIF UNIQUE : AUTOMATISATION TEXTE -> COLONNE P
    # Matière  : P3:P12
    # MainOeuvre: P17:P26
    # ======================================================
    def recherche_matiere(self) -> None:
        try:
            pr_path, hint = self._pr_path_and_hint()
            subprocess.run(["open", str(pr_path)], check=False)
            time.sleep(0.3)

            items = read_biblio_colA_openpyxl(pr_path, "Bibliothèque_Matière.xlsx")

            def on_pick(txt: str):
                target = excel_find_first_empty_cell(hint, "Chiffrage", "P3:P12")
                if not target:
                    messagebox.showwarning("Matière", "Zone P3:P12 pleine ou Excel pas prêt.")
                    return
                excel_set_cell_value(hint, "Chiffrage", target, txt)

            popup_search_20(self, "Matière → Chiffrage (P3:P12)", items, on_pick)

        except Exception as e:
            messagebox.showerror("Matière", str(e))

    def recherche_mo(self) -> None:
        try:
            pr_path, hint = self._pr_path_and_hint()
            subprocess.run(["open", str(pr_path)], check=False)
            time.sleep(0.3)

            items = read_biblio_colA_openpyxl(pr_path, "Bibliothèque_MainOeuvre")

            def on_pick(txt: str):
                target = excel_find_first_empty_cell(hint, "Chiffrage", "P17:P26")
                if not target:
                    messagebox.showwarning("Main-d'œuvre", "Zone P17:P26 pleine ou Excel pas prêt.")
                    return
                excel_set_cell_value(hint, "Chiffrage", target, txt)

            popup_search_20(self, "Main-d’œuvre → Chiffrage (P17:P26)", items, on_pick)

        except Exception as e:
            messagebox.showerror("Main-d'œuvre", str(e))

    # ---------------------------
    # Boutons non prioritaires (inchangés)
    # ---------------------------
    
    def calcul_reca(self) -> None:
        import os
        from datetime import datetime
        from tkinter import messagebox

        p = self._chemin_chantier_selectionne()
        if not p:
            messagebox.showwarning("Calcul / Reca", "Sélectionne un chantier.")
            return

        dossier = p.parent / p.stem
        chemin_pr = dossier / "data" / "prix_de_revient.xlsx"

        if not chemin_pr.exists():
            messagebox.showwarning(
                "Date du PR",
                "Le fichier prix_de_revient.xlsx est introuvable."
            )
            return

        timestamp = os.path.getmtime(chemin_pr)
        date_modif = datetime.fromtimestamp(timestamp).strftime("%d/%m/%Y %H:%M:%S")

        
        messagebox.showinfo(
            "Date du PR",
            f"Dernier enregistrement du PR :\n{date_modif}\n\n"
            "⚠️ Si tu viens de modifier le PR :\n"
            "➡️ fais CMD + S dans Excel\n"
            "➡️ puis clique sur Calcul / Reca"
        )

        try:
            self.importer_pr_dans_etat()
            messagebox.showinfo(
                "Calcul / Reca",
                "Prix de revient importé dans l'état d'avancement."
            )
        except Exception as e:
            messagebox.showerror("Erreur", str(e))

    def importer_pr_dans_etat(self):
        p = self._chemin_chantier_selectionne()
        if not p:
           return

        dossier = p.parent / p.stem
        chemin_pr = dossier / "data" / "prix_de_revient.xlsx"

        try:
            chemin_etat = next(dossier.glob("Etat_avancement*.xlsm"))
        except StopIteration:
            return

        if not chemin_pr.exists():
            return

        wb_pr = load_workbook(chemin_pr, data_only=True)
        ws_pr = wb_pr["Chiffrage"]

        wb_etat = load_workbook(chemin_etat, keep_vba=True)
        ws_etat = wb_etat["Bordereau"]

        # 👇 ICI TON IMPORT SIMPLE
        correspondances = [
            ("F30",  "L30"),
            ("F72",  "L32"),
            ("F114", "L34"),
        ]

        for cellule_pr, cellule_etat in correspondances:
            valeur = ws_pr[cellule_pr].value
            if isinstance(valeur, (int, float)):
                ws_etat[cellule_etat] = float(valeur)

        wb_etat.save(chemin_etat)
        wb_pr.close()
        wb_etat.close()

  
    def ouvrir_chantier(self) -> None:
        p = self._chemin_chantier_selectionne()
        if not p:
            messagebox.showwarning("Chantier", "Sélectionne un chantier dans la liste.")
            return
        try:
            chantier = lire_json(p)
            if chantier.get("bordereau", {}).get("articles"):
                afficher_bordereau(self, chantier, "Bordereau (JSON)")
            else:
                messagebox.showinfo("Chantier", "Chantier ouvert. (Pas encore de bordereau dans ce chantier.)")
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible d'ouvrir le chantier.\n{e}")

    def dossier_du_chantier(self) -> None:
        p = self._chemin_chantier_selectionne()
        if not p:
            messagebox.showwarning("Chantier", "Sélectionne un chantier dans la liste.")
            return
        ouvrir_dossier(p.parent)

    def nouveau_chantier(self) -> None:
        win = tk.Toplevel(self)
        win.title("Nouveau chantier")
        win.resizable(False, False)

        dossier_var = tk.StringVar(value="")
        client_var = tk.StringVar(value="")
        chantier_var = tk.StringVar(value="")
        adresse_var = tk.StringVar(value="")

        frame = ttk.Frame(win, padding=14)
        frame.grid(row=0, column=0, sticky="nsew")
        win.columnconfigure(0, weight=1)
        win.rowconfigure(0, weight=1)

        def choisir_dossier():
            from tkinter import filedialog
            d = filedialog.askdirectory(title="Choisir le dossier du chantier")
            if d:
                dossier_var.set(d)

        ttk.Label(frame, text="Dossier (optionnel)").grid(row=0, column=0, sticky="w")
        dossier_entry = ttk.Entry(frame, textvariable=dossier_var, width=52)
        dossier_entry.grid(row=1, column=0, sticky="ew", pady=(2, 6))
        ttk.Button(frame, text="Choisir…", command=choisir_dossier).grid(row=1, column=1, sticky="w", padx=(8, 0), pady=(2, 6))

        ttk.Label(frame, text="Nom du client").grid(row=2, column=0, sticky="w")
        e_client = ttk.Entry(frame, textvariable=client_var)
        e_client.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(2, 8))

        ttk.Label(frame, text="Nom du chantier").grid(row=4, column=0, sticky="w")
        e_chantier = ttk.Entry(frame, textvariable=chantier_var)
        e_chantier.grid(row=5, column=0, columnspan=2, sticky="ew", pady=(2, 8))

        ttk.Label(frame, text="Adresse").grid(row=6, column=0, sticky="w")
        e_adresse = ttk.Entry(frame, textvariable=adresse_var)
        e_adresse.grid(row=7, column=0, columnspan=2, sticky="ew", pady=(2, 12))

        def valider():
            self.nom_client = client_var.get().strip()
            self.nom_chantier = chantier_var.get().strip()
            self.adresse_chantier = adresse_var.get().strip()
            self.dossier_chantier = dossier_var.get().strip()

            if self.dossier_chantier:
                for fn in ("charger_chantier", "ouvrir_chantier", "importer_chantier", "load_chantier"):
                    f = getattr(self, fn, None)
                    if callable(f):
                        f(self.dossier_chantier)
                        win.destroy()
                        return

            win.destroy()

        btns = ttk.Frame(frame)
        btns.grid(row=8, column=0, columnspan=2, sticky="e")
        ttk.Button(btns, text="Annuler", command=win.destroy).pack(side="right")
        ttk.Button(btns, text="Valider", command=valider).pack(side="right", padx=(0, 8))

        frame.columnconfigure(0, weight=1)
        e_client.focus_set()
        win.grab_set()



    def modifier_chantier(self) -> None:
        messagebox.showinfo("Info", "Modifier chantier (à brancher)")


    def supprimer_chantier(self) -> None:
        messagebox.showinfo("Info", "Supprimer chantier (à brancher)")


    def postes_metre(self) -> None:
        messagebox.showinfo("Info", "Postes / Métré (à faire)")

print("JE SUIS DANS LE BON FICHIER")

if __name__ == "__main__":
    app = HorizonChantierApp()
    app.mainloop()
