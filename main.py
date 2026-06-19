#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import subprocess
import shutil
import re
import time
import unicodedata
from copy import copy
from datetime import datetime
from pathlib import Path
import tkinter as tk
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

from tkinter import ttk, messagebox

from openpyxl import load_workbook

APP_NAME = "Horizon Chantier"

# =========================
# FICHIERS (STRICT / SANS REFONTE)
# =========================
PV_SOURCE = "soumission_chapelette__MAJ_PV_2026-01-29_13-52-57.xlsx"
PV_BACKUP = "soumission_chapelette__backup_avant_injectionPV_2026-01-29_13-52-57.xlsx"
PV_CORRIGE = "soumission_chapelette__MAJ_injection_CORRIGE_PV_2026-01-29_13-52-57.xlsx"
FICHIER_PV = PV_SOURCE


# =========================
# AppleScript helpers (Excel ouvert)
# =========================
def _osascript(script: str) -> str:
    p = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
    if p.returncode != 0:
        raise RuntimeError(p.stderr.strip() or "Erreur AppleScript")
    return (p.stdout or "").strip()


def _a1_to_col_row(a1: str) -> tuple[str, int]:
    """
    "C12" -> ("C", 12)
    """
    m = re.fullmatch(r"([A-Z]+)(\d+)", a1.strip().upper().replace("$", ""))
    if not m:
        raise ValueError(f"Adresse cellule invalide : {a1}")
    return m.group(1), int(m.group(2))


def _is_allowed_target(cell_a1: str) -> bool:
    """
    Sécurité absolue : n'autorise que
    - P3:P12  (matière)
    - P17:P26 (main-d'œuvre)
    """
    col, row = _a1_to_col_row(cell_a1)
    if col == "P" and 3 <= row <= 12:
        return True
    if col == "P" and 17 <= row <= 26:
        return True
    return False


def _normalize_excel_addr(addr: str) -> str:
    """
    Normalise:
    - "$P$3" -> "P3"
    - "P3"   -> "P3"
    """
    if not addr:
        return ""
    a = addr.strip().replace("$", "").upper()
    if re.fullmatch(r"[A-Z]+[0-9]+", a):
        return a
    return ""


def excel_find_first_empty_cell(workbook_hint: str, sheet_name: str, cell_range: str) -> str:
    """
    1ère cellule vide dans une plage (colonne unique) ex: P3:P12
    Retour "P3" ou "" si aucune.
    Robuste: retrouve le classeur par "contient" (Excel Mac change parfois le nom).
    """
    hint = (workbook_hint or "").strip()
    if not hint:
        return ""

    for _ in range(8):
        script = f'''
        tell application "Microsoft Excel"
            set wbRef to missing value
            repeat with w in workbooks
                try
                    set wn to (name of w as string)
                    if wn contains "{hint}" then
                        set wbRef to w
                        exit repeat
                    end if
                end try
            end repeat

            if wbRef is missing value then return ""

            tell wbRef
                activate
                try
                    tell worksheet "{sheet_name}"
                        activate
                        set rng to range "{cell_range}"
                        repeat with c in cells of rng
                            try
                                set v to value of c
                                set f to formula of c

                                try
                                    if (f as string) is not "" then
                                        -- ne jamais proposer une cellule contenant une formule
                                    else
                                        if v is missing value then return (address of c) as string
                                        try
                                            if (v as string) is "" then return (address of c) as string
                                        end try
                                    end if
                                end try

                            on error
                                return (address of c) as string
                            end try
                        end repeat
                    end tell
                end try
            end tell
        end tell
        return ""
        '''
        try:
            addr_raw = _osascript(script)
            addr = _normalize_excel_addr(addr_raw)
            if addr:
                return addr
        except Exception:
            pass

        time.sleep(0.2)

    return ""


def excel_set_cell_value(workbook_hint: str, sheet_name: str, cell_a1: str, value: str) -> None:
    """
    Écrit une valeur dans Excel (classeur déjà ouvert) sans fermer.
    Sécurité : n'écrit QUE dans P3:P12 ou P17:P26.
    Robuste: retrouve le classeur par "contient".
    """
    cell_a1 = (cell_a1 or "").strip().replace("$", "").upper()
    if not _is_allowed_target(cell_a1):
        raise ValueError(f"Sécurité: écriture interdite hors zones autorisées : {cell_a1}")

    hint = (workbook_hint or "").strip()
    if not hint:
        raise ValueError("Nom de classeur vide")

    safe_val = (value or "").replace('"', '\\"')

    script = f'''
    tell application "Microsoft Excel"
        set wbRef to missing value
        repeat with w in workbooks
            try
                set wn to (name of w as string)
                if wn contains "{hint}" then
                    set wbRef to w
                    exit repeat
                end if
            end try
        end repeat

        if wbRef is missing value then return

        tell wbRef
            if not (exists worksheet "{sheet_name}") then return
            tell worksheet "{sheet_name}"
                if ((formula of range "{cell_a1}") as string) is not "" then error "Cellule formule protégée"
                set value of range "{cell_a1}" to "{safe_val}"
            end tell
        end tell
        activate
    end tell
    '''
    _osascript(script)


def _is_archive_or_history_path(path: Path) -> bool:
    ignored = {"archives", "archive", "harchives", "harchive", "historique_sauvegardes"}
    return any(part.lower() in ignored for part in path.parts)


def _chantier_root_for_path(path: str | Path) -> Path:
    src = Path(path).resolve()
    try:
        rel = src.relative_to(dossier_chantiers().resolve())
        if rel.parts:
            return dossier_chantiers() / rel.parts[0]
    except Exception:
        pass
    return src.parent.parent if src.parent.name == "data" else src.parent


def _historique_key(chantier_dir: Path, path: str | Path) -> str:
    src = Path(path)
    try:
        rel = src.relative_to(chantier_dir)
    except ValueError:
        rel = Path(src.name)
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", "__".join(rel.parts))


def _backup_excel_before_write(path: str | Path) -> Path:
    src = Path(path)
    chantier_dir = _chantier_root_for_path(src)
    backup_dir = chantier_dir / "Historique_Sauvegardes"
    backup_dir.mkdir(exist_ok=True)

    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    key = _historique_key(chantier_dir, src)
    dst = backup_dir / f"{stamp}_{key}"
    if dst.exists():
        i = 2
        while True:
            candidate = backup_dir / f"{stamp}_{i}_{key}"
            if not candidate.exists():
                dst = candidate
                break
            i += 1

    shutil.copy2(src, dst)
    return dst


def _normaliser_libelle_excel(valeur):
    txt = str(valeur or "").strip().lower()
    txt = txt.replace("é", "e").replace("è", "e").replace("ê", "e").replace("ë", "e")
    txt = txt.replace("à", "a").replace("â", "a").replace("ä", "a")
    txt = txt.replace("ù", "u").replace("û", "u").replace("ü", "u")
    txt = txt.replace("î", "i").replace("ï", "i")
    txt = txt.replace("ô", "o").replace("ö", "o")
    txt = txt.replace("ç", "c")
    txt = re.sub(r"\s+", " ", txt)
    return txt


def _premiere_ligne_synthese_etat(ws) -> int:
    libelles_synthese = (
        "total soumission hors tva",
        "total etat cumule hors tva",
        "total du mois hors tva",
        "total des avenants cumule",
        "montant global a facturer",
        "total execute",
        "tva",
        "revision",
        "avenant",
    )

    for ligne in range(24, ws.max_row + 1):
        contenu = " ".join(
            _normaliser_libelle_excel(ws.cell(ligne, col).value)
            for col in range(1, ws.max_column + 1)
        )
        if any(libelle in contenu for libelle in libelles_synthese):
            return ligne

    return ws.max_row + 1


def _montant_diminution_avenant(montant: float) -> float:
    return abs(montant)


def _chemin_fichier_chantier(dossier_chantier: Path, nom_fichier: str) -> Path:
    chemin = dossier_chantier / nom_fichier
    if dossier_chantier.exists():
        for candidat in dossier_chantier.iterdir():
            if candidat.name == nom_fichier:
                return candidat

        attendu = unicodedata.normalize("NFC", nom_fichier)
        for candidat in dossier_chantier.iterdir():
            if unicodedata.normalize("NFC", candidat.name) == attendu:
                return candidat

    if chemin.exists():
        return chemin

    return chemin


def _lire_synthese_avenants_pilotage(ws, convertir_nombre):
    avenants_plus = _lire_montant_synthese_colonne(ws, "E", convertir_nombre)
    avenants_moins = _montant_diminution_avenant(
        _lire_montant_synthese_colonne(ws, "G", convertir_nombre)
    )
    return avenants_plus, avenants_moins


def _lire_revision_globale_pilotage(ws, convertir_nombre):
    return _lire_montant_synthese_colonne(ws, "E", convertir_nombre)


def _lire_montant_synthese_colonne(ws, colonne: str, convertir_nombre) -> float:
    for ligne in range(ws.max_row, 0, -1):
        valeur = ws[f"{colonne}{ligne}"].value
        if valeur in (None, "", "-", "—"):
            continue
        montant = convertir_nombre(valeur)
        texte_valeur = str(valeur).strip()
        if montant != 0 or texte_valeur in {"0", "0,0", "0.0", "0,00", "0.00"}:
            return montant
    return 0


def _set_cell_protection(cell, locked: bool, hidden: bool = False) -> None:
    prot = copy(cell.protection)
    prot.locked = locked
    prot.hidden = hidden
    cell.protection = prot


def _protect_formula_cells(wb, source_path: str | Path | None = None) -> None:
    filename = Path(source_path).name.lower() if source_path else ""

    for ws in wb.worksheets:
        synthese_debut = None
        if "etat_avancement" in filename and ws.title == "Bordereau":
            synthese_debut = _premiere_ligne_synthese_etat(ws)

        for row in ws.iter_rows():
            for cell in row:
                if cell.__class__.__name__ == "MergedCell":
                    continue

                in_synthese = synthese_debut is not None and cell.row >= synthese_debut
                has_formula = cell.data_type == "f" or (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                )

                if in_synthese:
                    _set_cell_protection(cell, locked=False, hidden=False)
                else:
                    _set_cell_protection(cell, locked=bool(has_formula), hidden=bool(has_formula))

        ws.protection.sheet = True
        ws.protection.set_password("1234")


def _prepare_excel_file_before_write(path: str | Path, keep_vba: bool = False) -> None:
    _backup_excel_before_write(path)
    path = Path(path)
    wb = load_workbook(path, keep_vba=keep_vba or path.suffix.lower() == ".xlsm")
    _protect_formula_cells(wb, path)
    wb.save(path)
    wb.close()


def _export_excel_to_pdf(excel_path: str | Path, pdf_path: str | Path) -> None:
    import xlwings as xw

    excel_path = Path(excel_path).resolve()
    pdf_path = Path(pdf_path)
    book = None
    opened_here = False

    for app in xw.apps:
        for candidate in app.books:
            try:
                if Path(candidate.fullname).resolve() == excel_path:
                    book = candidate
                    break
            except Exception:
                continue
        if book is not None:
            break

    if book is None:
        book = xw.Book(str(excel_path))
        opened_here = True

    try:
        book.to_pdf(str(pdf_path))
    finally:
        if opened_here:
            book.close()


def _chemin_classeur_excel_actif() -> Path | None:
    import xlwings as xw

    apps = []
    try:
        apps.append(xw.apps.active)
    except Exception:
        pass

    try:
        for app in xw.apps:
            if app not in apps:
                apps.append(app)
    except Exception:
        pass

    for app in apps:
        try:
            book = app.books.active
            chemin = Path(book.fullname)
        except Exception:
            continue
        if chemin.suffix.lower() in {".xlsx", ".xlsm", ".xls"} and chemin.exists():
            return chemin

    return None


# =========================
# Lecture bibliothèques (colonne A) - lecture seule
# =========================
def read_biblio_colA_openpyxl(pr_path: Path, sheet_name: str) -> list[str]:
    wb = load_workbook(pr_path, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Onglet introuvable : {sheet_name}")
        ws = wb[sheet_name]
        out: list[str] = []
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v is None:
                continue
            s = str(v).strip()
            if s:
                out.append(s)
        return out
    finally:
        wb.close()


# =========================
# Popup recherche (20 résultats)
# =========================
def popup_search_20(parent: tk.Tk, titre: str, items: list[str], on_pick) -> None:
    win = tk.Toplevel(parent)
    win.title(titre)
    win.geometry("780x520")
    win.transient(parent)
    win.grab_set()

    frm = ttk.Frame(win, padding=12)
    frm.pack(fill="both", expand=True)

    ttk.Label(frm, text=titre, font=("Helvetica", 14, "bold")).pack(anchor="w")

    q = tk.StringVar()
    entry = ttk.Entry(frm, textvariable=q)
    entry.pack(fill="x", pady=(10, 8))
    entry.focus_set()

    lb = tk.Listbox(frm, height=18)
    lb.pack(fill="both", expand=True)

    ttk.Label(frm, text="Tape 2 lettres. Double-clic ou Entrée pour valider.", foreground="#555").pack(
        anchor="w", pady=(8, 0)
    )

    current: list[str] = []

    def refresh(*_):
        text = q.get().strip().lower()
        lb.delete(0, tk.END)

        if not text:
            current[:] = items[:20]
        else:
            filtered = [s for s in items if text in s.lower()]
            current[:] = filtered[:20]

        for s in current:
            lb.insert(tk.END, s)

        if current:
            lb.selection_set(0)

    def choose(_event=None):
        sel = lb.curselection()
        if not sel:
            return
        picked = current[sel[0]]
        try:
            on_pick(picked)
        finally:
            win.destroy()

    q.trace_add("write", refresh)
    lb.bind("<Double-Button-1>", choose)
    win.bind("<Return>", choose)

    refresh()


# ============================================================
# Montant en lettres FR (euros + centimes) - inchangé
# ============================================================
_UNITS = ["zéro", "un", "deux", "trois", "quatre", "cinq", "six", "sept", "huit", "neuf"]
_TEENS = ["dix", "onze", "douze", "treize", "quatorze", "quinze", "seize", "dix-sept", "dix-huit", "dix-neuf"]
_TENS = ["", "", "vingt", "trente", "quarante", "cinquante", "soixante", "soixante", "quatre-vingt", "quatre-vingt"]


def _below_100(n: int) -> str:
    assert 0 <= n < 100
    if n < 10:
        return _UNITS[n]
    if 10 <= n < 20:
        return _TEENS[n - 10]

    ten = n // 10
    unit = n % 10

    if ten == 7:
        return "soixante" + ("-" + _below_100(10 + unit) if unit else "-dix")
    if ten == 9:
        return "quatre-vingt" + ("-" + _below_100(10 + unit) if unit else "-dix")

    base = _TENS[ten]

    if ten == 8 and unit == 0:
        return "quatre-vingts"
    if unit == 0:
        return base

    if unit == 1 and ten in (2, 3, 4, 5, 6):
        return f"{base} et un"

    return f"{base}-{_UNITS[unit]}"


def _below_1000(n: int) -> str:
    assert 0 <= n < 1000
    if n < 100:
        return _below_100(n)

    hundred = n // 100
    rest = n % 100

    if hundred == 1:
        head = "cent"
    else:
        head = f"{_UNITS[hundred]} cent"

    if rest == 0 and hundred > 1:
        return head + "s"
    if rest == 0:
        return head

    return head + " " + _below_100(rest)


def _int_to_words_fr(n: int) -> str:
    if n == 0:
        return "zéro"

    parts: list[str] = []

    billions = n // 1_000_000_000
    n %= 1_000_000_000
    millions = n // 1_000_000
    n %= 1_000_000
    thousands = n // 1000
    n %= 1000
    rest = n

    if billions:
        parts.append("un milliard" if billions == 1 else _below_1000(billions) + " milliards")

    if millions:
        parts.append("un million" if millions == 1 else _below_1000(millions) + " millions")

    if thousands:
        parts.append("mille" if thousands == 1 else _below_1000(thousands) + " mille")

    if rest:
        parts.append(_below_1000(rest))

    return " ".join(parts)


def montant_en_lettres_fr(valeur: float) -> str:
    v = round(float(valeur), 2)
    euros = int(v)
    centimes = int(round((v - euros) * 100))
    if centimes == 100:
        euros += 1
        centimes = 0

    euros_txt = _int_to_words_fr(euros)
    euro_label = "euro" if euros == 1 else "euros"

    if centimes == 0:
        return f"{euros_txt} {euro_label}"

    cent_txt = _int_to_words_fr(centimes)
    cent_label = "centime" if centimes == 1 else "centimes"
    return f"{euros_txt} {euro_label} et {cent_txt} {cent_label}"


# ============================================================
# PR / PV (inchangés)
# ============================================================
def open_pr(chantier_dir: str | Path):
    chantier = Path(chantier_dir)
    pr = chantier / "data" / "prix_de_revient.xlsx"
    if not pr.exists():
        raise FileNotFoundError(f"PR introuvable : {pr}")
    subprocess.run(["open", str(pr)], check=False)


def inject_pv(chantier_dir: str | Path) -> int:
    chantier = Path(chantier_dir)
    pr_path = chantier / "data" / "prix_de_revient.xlsx"
    pv_path = chantier / PV_SOURCE

    if not pr_path.exists():
        raise FileNotFoundError(f"PR introuvable : {pr_path}")
    if not pv_path.exists():
        raise FileNotFoundError(f"PV source introuvable : {pv_path}")

    pr_wb = load_workbook(pr_path, data_only=True)
    pr_ws = pr_wb["Chiffrage"] if "Chiffrage" in pr_wb.sheetnames else pr_wb.active


    # trouver la 1ère ligne d'article à partir de B3
    row_article = None
    for r in range(3, pr_ws.max_row + 1):
        v = pr_ws.cell(row=r, column=2).value  # colonne B
        if v is not None and str(v).strip() != "":
            row_article = r
            break

    if row_article is None:
        pr_wb.close()
        raise ValueError("Aucun ARTICLE trouvé dans la colonne B (à partir de B3).")

    article = pr_ws.cell(row=row_article, column=2).value
    designation = pr_ws.cell(row=row_article, column=3).value

    price = pr_ws["F32"].value

    pr_wb.close()



    if article is None or str(article) == "":
        raise ValueError("ARTICLE vide dans PR (B3).")
    if price is None or str(price) == "":
        raise ValueError("PRIX vide dans PR (F32).")

    article = str(article)
    designation = "" if designation is None else str(designation)

    price_f = float(price)
    price_2d = round(price_f, 2)
    letters = montant_en_lettres_fr(price_2d)

    _backup_excel_before_write(pv_path)
    wb = load_workbook(pv_path)
    ws = wb["PV"] if "PV" in wb.sheetnames else wb.active

    row_found = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        if str(v) == article:
            row_found = r
            break

    if row_found is None:
        wb.close()
        raise ValueError(f"Article introuvable dans PV source : '{article}'")

    ws.cell(row=row_found, column=2).value = designation

    g = ws.cell(row=row_found, column=7)
    g.value = price_2d
    g.number_format = "#,##0.00"

    ws.cell(row=row_found, column=8).value = letters

    _protect_formula_cells(wb, pv_path)
    wb.save(pv_path)
    wb.close()
    return 1


def sync_pv_to_copies(chantier_dir: str | Path) -> None:
    chantier = Path(chantier_dir)

    src = chantier / PV_SOURCE
    tgt_backup = chantier / PV_BACKUP
    tgt_corrige = chantier / PV_CORRIGE

    for p in (src, tgt_backup, tgt_corrige):
        if not p.exists():
            raise FileNotFoundError(f"Fichier introuvable : {p}")

    src_wb = load_workbook(src, data_only=True)
    src_ws = src_wb["PV"] if "PV" in src_wb.sheetnames else src_wb.active
    data = [row for row in src_ws.iter_rows(values_only=True)]
    src_wb.close()

    def _write(target: Path):
        _backup_excel_before_write(target)
        wb = load_workbook(target)
        ws = wb["PV"] if "PV" in wb.sheetnames else wb.active

        for r_idx, row in enumerate(data, start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if cell.__class__.__name__ == "MergedCell":
                    continue
                cell.value = val

        _protect_formula_cells(wb, target)
        wb.save(target)
        wb.close()

    _write(tgt_backup)
    _write(tgt_corrige)


# ---------------------------
# Dossiers (auto Chantier/Chantiers)
# ---------------------------
def dossier_base() -> Path:
    return Path.home() / "Desktop" / "Horizon_Chantier_Data"


def dossier_chantiers() -> Path:
    base = dossier_base()
    d1 = base / "Chantier"
    d2 = base / "Chantiers"
    if d1.exists():
        return d1
    if d2.exists():
        return d2
    d2.mkdir(parents=True, exist_ok=True)
    return d2
def ouvrir_doc(self, nom_fichier):
    sel = self.tree.selection()
    if not sel:
        messagebox.showwarning("Sélection", "Sélectionne un chantier.")
        return

    item = self.tree.item(sel[0])

    # 1) Nom sélectionné (text si présent, sinon values[0])
    chantier_sel = (item.get("text") or "").strip()
    if not chantier_sel:
        vals = item.get("values", [])
        chantier_sel = str(vals[0]).strip() if vals else ""

    base = dossier_chantiers()

    # 2) On cherche le dossier chantier qui existe vraiment
    dossier = base / chantier_sel
    if not dossier.exists():
        def norm(s):
            return s.lower().strip().replace(" ", "_").replace("-", "_")

        cible = norm(chantier_sel)
        dossier = next(
            (p for p in base.iterdir() if p.is_dir() and norm(p.name) == cible),
            None
        )

    if not dossier or not Path(dossier).exists():
        messagebox.showerror(
            "Introuvable",
            f"Dossier chantier introuvable pour :\n{chantier_sel}\n\nBase : {base}"
        )
        return

    chemin = Path(dossier) / nom_fichier

    if not chemin.exists():
        # Si on n'a pas trouvé le fichier, on tente l'autre dossier chantier
        # (ex: "soumission_chapelette" vs "Soumission Chapelette")
        def norm(s):
            return str(s).lower().strip().replace(" ", "_").replace("-", "_")

        cible = norm(chantier_sel)
        for p in base.iterdir():
            if p.is_dir() and norm(p.name) == cible:
                alt = p / nom_fichier
                if alt.exists():
                   chemin = alt
                   break

    if not chemin.exists():
        messagebox.showerror("Introuvable", f"Fichier absent :\n{chemin}")
        return

    if not self._preparer_ouverture_document(chemin):
        return

    if sys.platform == "darwin":
        subprocess.run(["open", str(chemin)], check=False)
    elif os.name == "nt":
        os.startfile(str(chemin))
    if self._doit_rappeler_pdf_historique(chemin):
        self._planifier_rappel_pdf_historique_ouverture()

def ouvrir_dossier(path: Path) -> None:
    try:
        subprocess.run(["open", str(path)], check=False)
    except Exception as e:
        messagebox.showerror("Erreur", f"Impossible d'ouvrir le dossier.\n{e}")


# ---------------------------
# Lecture chantier JSON
# ---------------------------
def lire_json(path: Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def ecrire_json(path: Path, data: dict) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def infos_chantier(chantier: dict) -> dict:
    return {
        "nom": chantier.get("nom", chantier.get("chantier", "")) or "",
        "client": chantier.get("client", "") or "",
        "etat": chantier.get("etat", chantier.get("type_travaux", "")) or "",
        "avancement": chantier.get("avancement", chantier.get("avancement_pct", 0)) or 0,
        "debut": chantier.get("date_debut", chantier.get("debut_travaux", "")) or "",
        "fin": chantier.get("date_fin", chantier.get("fin_prevue", chantier.get("date_fin_prevue", ""))) or "",
    }


# ---------------------------
# Fenêtre Bordereau (visu JSON)
# ---------------------------
def afficher_bordereau(parent: tk.Tk, chantier: dict, titre: str = "Bordereau") -> None:
    bord = chantier.get("bordereau", {})
    articles = bord.get("articles", {})

    if not articles:
        messagebox.showwarning("Bordereau", "Aucun bordereau dans ce chantier.")
        return

    win = tk.Toplevel(parent)
    win.title(titre)
    win.geometry("1200x650")

    cols = ("Article", "Libellé", "Unité", "Qté", "PU", "Total")

    tree = ttk.Treeview(win, columns=cols, show="headings")
    vsb = ttk.Scrollbar(win, orient="vertical", command=tree.yview)
    hsb = ttk.Scrollbar(win, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

    for c in cols:
        tree.heading(c, text=c)
        if c == "Libellé":
            tree.column(c, width=700, anchor="w")
        elif c == "Article":
            tree.column(c, width=160, anchor="w")
        elif c == "Unité":
            tree.column(c, width=80, anchor="w")
        else:
            tree.column(c, width=120, anchor="e")

    def v(x):
        return "" if x is None else x

    for article_id in sorted(articles.keys()):
        a = articles[article_id] or {}
        tree.insert(
            "",
            "end",
            values=(
                v(a.get("article_id", article_id)),
                v(a.get("libelle", "")),
                v(a.get("unite", "")),
                v(a.get("quantite", 0)),
                v(a.get("pu_vente", 0)),
                v(a.get("total_vente", 0)),
            ),
        )

    tree.grid(row=0, column=0, sticky="nsew")
    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")

    win.grid_rowconfigure(0, weight=1)
    win.grid_columnconfigure(0, weight=1)


# ---------------------------
# App
# ---------------------------
class HorizonChantierApp(tk.Tk):


    def __init__(self) -> None:
        super().__init__()
        self.title(APP_NAME)
        self.geometry("1400x800")
        self.minsize(1300, 950)
        self.dernier_fichier_excel_ouvert: Path | None = None
        self.rappel_pdf_historique_fenetre = None
        self._rappel_pdf_historique_after_id = None

        self._build_ui()
        self.refresh_liste()

    def _nom_chantier_selectionne(self) -> str | None:
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Chantier", "Sélectionne un chantier dans la liste.")
            return None
        vals = self.tree.item(sel[0]).get("values", [])
        return str(vals[0]).strip() if vals else ""

    def _ouvrir_fichier_chantier(self, nom_fichier: str) -> None:
        nom_chantier = self._nom_chantier_selectionne()
        if not nom_chantier:
            return
        chemin = _chemin_fichier_chantier(dossier_chantiers() / nom_chantier, nom_fichier)
        if not chemin.exists():
            messagebox.showerror("Introuvable", f"Fichier absent :\n{chemin}")
            return
        if not self._preparer_ouverture_document(chemin):
            return
        subprocess.Popen(["open", str(chemin)])
        if self._doit_rappeler_pdf_historique(chemin):
            self._planifier_rappel_pdf_historique_ouverture()

    def _ouvrir_premier_fichier_chantier(self, motif: str) -> None:
        nom_chantier = self._nom_chantier_selectionne()
        if not nom_chantier:
            return
        try:
            chemin = next((dossier_chantiers() / nom_chantier).glob(motif))
        except StopIteration:
            messagebox.showerror("Introuvable", f"Fichier absent :\n{motif}")
            return
        if not self._preparer_ouverture_document(chemin):
            return
        subprocess.Popen(["open", chemin.as_posix()])
        if self._doit_rappeler_pdf_historique(chemin):
            self._planifier_rappel_pdf_historique_ouverture()

    def _dossier_depuis_json(self, p: Path) -> Path:
        try:
            chantier = lire_json(p)
            info = infos_chantier(chantier)
            nom = str(info.get("nom") or "").strip()
            if nom:
                return p.parent / nom
        except Exception:
            pass
        return p.parent / p.stem

    def _memoriser_fichier_excel(self, chemin: Path, afficher_rappel: bool = True) -> None:
        if chemin.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
            self.dernier_fichier_excel_ouvert = chemin

    def _preparer_ouverture_document(self, chemin: Path) -> bool:
        if chemin.is_dir():
            if not self._dossier_contient_fichier_excel(chemin):
                return True
            return self._confirmer_ouverture_document_excel()

        if chemin.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
            return True
        if not self._confirmer_ouverture_document_excel():
            return False
        self._memoriser_fichier_excel(chemin)
        return True

    def _dossier_contient_fichier_excel(self, dossier: Path) -> bool:
        try:
            return any(
                fichier.is_file() and fichier.suffix.lower() in {".xlsx", ".xlsm", ".xls"}
                for fichier in dossier.rglob("*")
            )
        except Exception:
            return False

    def _doit_rappeler_pdf_historique(self, chemin: Path) -> bool:
        if chemin.is_dir():
            return self._dossier_contient_fichier_excel(chemin)
        return chemin.suffix.lower() in {".xlsx", ".xlsm", ".xls"}

    def _planifier_rappel_pdf_historique_ouverture(self) -> None:
        if self._rappel_pdf_historique_after_id is not None:
            return
        self._rappel_pdf_historique_after_id = self.after(1200, self._afficher_rappel_pdf_historique_ouverture)

    def _afficher_rappel_pdf_historique_ouverture(self) -> None:
        self._rappel_pdf_historique_after_id = None
        if self.rappel_pdf_historique_fenetre and self.rappel_pdf_historique_fenetre.winfo_exists():
            self.rappel_pdf_historique_fenetre.lift()
            return

        win = tk.Toplevel(self)
        win.title("Sauvegarde PDF Historique")
        win.resizable(False, False)
        win.geometry("560x340")
        self.rappel_pdf_historique_fenetre = win
        try:
            win.attributes("-topmost", True)
        except Exception:
            pass

        cadre = ttk.Frame(win, padding=32)
        cadre.pack(fill="both", expand=True)

        tk.Label(
            cadre,
            text="⚠️",
            font=("Helvetica", 54, "bold"),
            fg="#b45309",
        ).pack(pady=(0, 8))
        tk.Label(
            cadre,
            text="ATTENTION - SAUVEGARDE OBLIGATOIRE",
            font=("Helvetica", 19, "bold"),
            wraplength=480,
            justify="center",
            fg="#7c2d12",
        ).pack(pady=(0, 18))
        tk.Label(
            cadre,
            text=(
                "N'oubliez pas d'enregistrer vos modifications dans Excel.\n\n"
                "Après vos modifications importantes, créez votre PDF Historique.\n\n"
                "Ce PDF est votre copie de sécurité en cas de modification ou d'erreur ultérieure."
            ),
            font=("Helvetica", 14, "bold"),
            wraplength=490,
            justify="center",
        ).pack(pady=(0, 26))

        def fermer():
            self.rappel_pdf_historique_fenetre = None
            win.destroy()

        ttk.Button(cadre, text="✓ OK, j'ai compris", command=fermer).pack()

        win.protocol("WM_DELETE_WINDOW", fermer)
        win.update_idletasks()
        largeur_ecran = win.winfo_screenwidth()
        hauteur_ecran = win.winfo_screenheight()
        x = self.winfo_rootx() + self.winfo_width() + 16
        if x + win.winfo_width() > largeur_ecran - 24:
            x = largeur_ecran - win.winfo_width() - 24
        y = self.winfo_rooty() + 80
        if y + win.winfo_height() > hauteur_ecran - 24:
            y = max(24, hauteur_ecran - win.winfo_height() - 24)
        win.geometry(f"+{max(x, 0)}+{max(y, 0)}")

    def _confirmer_ouverture_document_excel(self) -> bool:
        confirmation = tk.BooleanVar(value=False)

        win = tk.Toplevel(self)
        win.title("Règles de travail obligatoires")
        win.transient(self)
        win.grab_set()
        win.resizable(False, False)

        cadre = ttk.Frame(win, padding=28)
        cadre.pack(fill="both", expand=True)

        tk.Label(
            cadre,
            text="⚠️ ATTENTION - RÈGLES DE TRAVAIL OBLIGATOIRES",
            font=("Helvetica", 18, "bold"),
            fg="#7c2d12",
            wraplength=620,
            justify="center",
        ).pack(pady=(0, 18))
        tk.Label(
            cadre,
            text=(
                "Vous allez ouvrir un document de travail.\n\n"
                "Avant de continuer, vous confirmez avoir compris les règles suivantes :\n\n"
                "• Enregistrer votre fichier Excel après chaque modification.\n"
                "• Créer votre PDF Historique après vos modifications importantes.\n"
                "• Respecter les cellules protégées et les zones prévues dans les modèles Horizon Chantier."
            ),
            font=("Helvetica", 13, "bold"),
            wraplength=620,
            justify="left",
        ).pack(pady=(0, 24))

        def confirmer():
            confirmation.set(True)
            win.destroy()

        ttk.Button(
            cadre,
            text="✓ J'ai compris - Ouvrir le document",
            command=confirmer,
        ).pack()

        win.protocol("WM_DELETE_WINDOW", lambda: win.destroy())
        win.bind("<Escape>", lambda _event: win.destroy())
        win.update_idletasks()
        x = self.winfo_rootx() + (self.winfo_width() - win.winfo_width()) // 2
        y = self.winfo_rooty() + (self.winfo_height() - win.winfo_height()) // 2
        win.geometry(f"+{max(x, 0)}+{max(y, 0)}")
        self.wait_window(win)
        return confirmation.get()

    def sauvegarde_pdf_historique(self) -> bool:
        chemin = _chemin_classeur_excel_actif() or self.dernier_fichier_excel_ouvert
        if not chemin or not chemin.exists():
            messagebox.showwarning(
                "Sauvegarde PDF Historique",
                "Aucun fichier Excel ouvert n'a été trouvé."
            )
            return False
        self._memoriser_fichier_excel(chemin, afficher_rappel=False)

        try:
            dossier_chantier = _chantier_root_for_path(chemin)
            historique = dossier_chantier / "Historique_Sauvegardes"
            historique.mkdir(exist_ok=True)

            nom_source = re.sub(r"\.[^.]+$", "", chemin.name)
            nom_pdf = f"{datetime.now().strftime('%Y-%m-%d_%H-%M')}_{nom_source}.pdf"
            destination = historique / nom_pdf

            _export_excel_to_pdf(chemin, destination)
            messagebox.showinfo(
                "Sauvegarde PDF Historique",
                f"PDF historique créé :\n{destination}"
            )
            return True
        except Exception as e:
            messagebox.showerror("Sauvegarde PDF Historique", f"Impossible de créer le PDF.\n{e}")
            return False

    def _confirmer_cloture_etat(self) -> bool:
        confirmation = tk.BooleanVar(value=False)

        win = tk.Toplevel(self)
        win.title("Clôture de l'État")
        win.transient(self)
        win.grab_set()
        win.resizable(False, False)

        cadre = ttk.Frame(win, padding=28)
        cadre.pack(fill="both", expand=True)

        tk.Label(
            cadre,
            text="⚠️",
            font=("Helvetica", 54, "bold"),
            fg="#b42318",
        ).pack(pady=(0, 8))
        tk.Label(
            cadre,
            text="ATTENTION - CLÔTURE IRRÉVERSIBLE",
            font=("Helvetica", 20, "bold"),
            fg="#b42318",
            wraplength=580,
            justify="center",
        ).pack(pady=(0, 18))
        tk.Label(
            cadre,
            text=(
                "Vous allez clôturer l'État d'avancement.\n\n"
                "Cette opération effectue la mise à zéro des quantités du mois "
                "et il sera impossible de revenir à l'état précédent.\n\n"
                "Avez-vous créé votre PDF Historique de sauvegarde ?"
            ),
            font=("Helvetica", 13, "bold"),
            wraplength=590,
            justify="center",
        ).pack(pady=(0, 24))

        boutons = ttk.Frame(cadre)
        boutons.pack(fill="x")

        def annuler():
            confirmation.set(False)
            win.destroy()

        def confirmer():
            confirmation.set(True)
            win.destroy()

        bouton_annuler = ttk.Button(boutons, text="❌ Annuler", command=annuler)
        bouton_annuler.pack(side="left")
        ttk.Button(
            boutons,
            text="✅ Oui, j'ai créé mon PDF Historique et je confirme la clôture",
            command=confirmer,
        ).pack(side="right", padx=(20, 0))

        win.protocol("WM_DELETE_WINDOW", annuler)
        bouton_annuler.focus_set()
        win.bind("<Return>", lambda _event: None)
        win.bind("<Escape>", lambda _event: annuler())
        win.update_idletasks()
        x = self.winfo_rootx() + (self.winfo_width() - win.winfo_width()) // 2
        y = self.winfo_rooty() + (self.winfo_height() - win.winfo_height()) // 2
        win.geometry(f"+{max(x, 0)}+{max(y, 0)}")
        self.wait_window(win)
        return confirmation.get()

    def _build_ui(self) -> None:
        root = ttk.Frame(self, padding=18)
        root.pack(fill="both", expand=True)
        style = ttk.Style()
        style.configure("TitreBloc.TLabel",
                font=("Helvetica", 12, "bold"),
                foreground="#1f4e79")
                
        title = ttk.Label(root, text=APP_NAME, font=("Helvetica", 22, "bold"))
        title.pack(anchor="w")

        self.lbl_path = ttk.Label(root, text=f"Emplacement: {dossier_chantiers()}")
        self.lbl_path.pack(fill="x", pady=(4, 8))

        cols = ("Chantier", "Client", "État", "Avancement", "Début", "Fin prévue")

        zone = ttk.Frame(root)
        zone.pack(fill="both", expand=True)

        zone.columnconfigure(0, weight=0)
        zone.columnconfigure(1, weight=1)
        zone.columnconfigure(2, weight=0)
        zone.rowconfigure(0, weight=1)

        frame_gauche = ttk.Frame(zone)
        frame_gauche.grid(row=0, column=0, sticky="ns", padx=(0, 12))

        frame_centre = ttk.Frame(zone)
        frame_centre.grid(row=0, column=1, sticky="nsew")

        frame_droite = ttk.Frame(zone)
        frame_droite.grid(row=0, column=2, sticky="ns", padx=(12, 0))

        self.tree = ttk.Treeview(frame_centre, columns=cols, show="headings")
        for c in cols:
            self.tree.heading(c, text=c)
            if c == "Chantier":
                self.tree.column(c, width=260, anchor="w")
            elif c == "Client":
                self.tree.column(c, width=180, anchor="w")
            elif c == "État":
                self.tree.column(c, width=120, anchor="w")
            elif c == "Avancement":
                self.tree.column(c, width=110, anchor="e")
            else:
                self.tree.column(c, width=120, anchor="w")

        self.tree.pack(fill="both", expand=True)

        # -------------------------
        # GAUCHE
        # -------------------------

        ttk.Separator(frame_gauche).pack(fill="x", pady=10)
        tk.Label(frame_gauche, text="AIDE",
            font=("Helvetica", 15, "bold"),
            fg="#C62828").pack(anchor="w", pady=(2, 4))
        ttk.Separator(frame_gauche).pack(fill="x", pady=(0, 8))
        ttk.Button(frame_gauche, text="Aide / Dépannage", command=self.ouvrir_aide).pack(fill="x", pady=3)

        ttk.Separator(frame_gauche).pack(fill="x", pady=10)
        tk.Label(frame_gauche, text="GESTION",
            font=("Helvetica", 14, "bold"),
            fg="#0052cc").pack(anchor="w", pady=(0, 2))
        ttk.Separator(frame_gauche).pack(fill="x", pady=(0, 8))

        ttk.Button(frame_gauche, text="📄 Ouvrir dans le logiciel", command=self.ouvrir_chantier).pack(fill="x", pady=3)
        ttk.Button(frame_gauche, text="➕ Nouveau chantier", command=self.nouveau_chantier).pack(fill="x", pady=3)
        ttk.Button(frame_gauche, text="✏️ Modifier", command=self.modifier_chantier).pack(fill="x", pady=3)
        ttk.Button(frame_gauche, text="🗑️ Supprimer", command=self.supprimer_chantier).pack(fill="x", pady=3)
        ttk.Button(frame_gauche, text="🔄 Rafraîchir", command=self.refresh_liste).pack(fill="x", pady=3)
        tk.Button(
            frame_gauche,
            text="Sauvegarde PDF Historique",
            command=self.sauvegarde_pdf_historique,
            bg="yellow",
            background="yellow",
            fg="black",
            foreground="black",
            activebackground="yellow",
            activeforeground="black",
            font=("Helvetica", 12, "bold"),
            relief="raised",
            bd=2,
            highlightbackground="yellow",
            highlightcolor="yellow",
            highlightthickness=2,
            padx=8,
            pady=6,
        ).pack(fill="x", pady=(6, 4))

        ttk.Separator(frame_gauche).pack(fill="x", pady=10)

        tk.Label(frame_gauche, text="DOCUMENTS",
         font=("Helvetica", 15, "bold"),
         fg="#0052cc").pack(anchor="w", pady=(2, 4))
        ttk.Separator(frame_gauche).pack(fill="x", pady=(0, 8))

        ttk.Button(frame_gauche, text="📁 Ce chantier", command=self.dossier_du_chantier).pack(fill="x", pady=3)

        ttk.Button(
            frame_gauche,
            text="🔎 Dossier administratif",
            command=lambda: self._ouvrir_fichier_chantier("Administratif")
        ).pack(fill="x", pady=3)

        ttk.Button(
            frame_gauche,
            text="📄 Cahier des charges administratif",
            command=lambda: self._ouvrir_fichier_chantier("Cahier_administratif.pdf")
        ).pack(fill="x", pady=3)

        ttk.Button(
           frame_gauche,
            text="🛠 Cahier des charges technique",
            command=lambda: self._ouvrir_fichier_chantier("Cahier_technique.pdf")
        ).pack(fill="x", pady=3)

        ttk.Button(
            frame_gauche,
            text="📐 Postes / Métré",
            command=lambda: self._ouvrir_fichier_chantier("Metre_detaille.pdf")
        ).pack(fill="x", pady=3)

        ttk.Button(
            frame_gauche,
            text="🦺 Plan de sécurité (PSS)",
            command=lambda: self._ouvrir_fichier_chantier("PSS.pdf")
        ).pack(fill="x", pady=3)

        ttk.Button(
            frame_gauche,
            text="📄 Décompte intempéries",
            command=lambda: self._ouvrir_fichier_chantier("Décompte_intempéries.doc")
        ).pack(fill="x", pady=3)

        # -------------------------
        # DROITE
        # -------------------------

        

        ttk.Separator(frame_droite).pack(fill="x", pady=10)

        tk.Label(frame_droite, text="EXECUTION",
         font=("Helvetica", 15, "bold"),
         fg="#0052cc").pack(anchor="w", pady=(2, 4))
        ttk.Separator(frame_droite).pack(fill="x", pady=(0, 8))

        ttk.Button(
            frame_droite,
            text="📅 Planning d’exécution",
            command=lambda: self._ouvrir_fichier_chantier("Planning_execution.xlsx")
        ).pack(fill="x", pady=3)

        ttk.Separator(frame_droite).pack(fill="x", pady=10)

        tk.Label(frame_droite, text="CALCULS",
         font=("Helvetica", 15, "bold"),
         fg="#0052cc").pack(anchor="w", pady=(2, 4))
        ttk.Separator(frame_droite).pack(fill="x", pady=(0, 8))

        ttk.Button(
            frame_droite,
            text="🔎 Coût de la sécurité",
            command=lambda: self._ouvrir_fichier_chantier("Cout_securite.xlsx")
        ).pack(fill="x", pady=3)

        ttk.Button(
            frame_droite,
            text="📐 Formule de révision",
            command=lambda: self._ouvrir_fichier_chantier("Formule_révision.xlsm")
        ).pack(fill="x", pady=3)

        ttk.Button(
            frame_droite,
            text="💰 Prix de revient",
            command=lambda: self._ouvrir_fichier_chantier("data/prix_de_revient.xlsx")
        ).pack(fill="x", pady=3)

        ttk.Button(frame_droite, text="🔍 Vérifier PR", command=self.verifier_pr).pack(fill="x", pady=3)
        ttk.Button(frame_droite, text="⚙️ Calcul / Reca", command=self.calcul_reca).pack(fill="x", pady=3)

        ttk.Separator(frame_droite).pack(fill="x", pady=10)

        tk.Label(frame_droite, text="ETAT D'AVANCEMENT",
         font=("Helvetica", 15, "bold"),
         fg="#0052cc").pack(anchor="w", pady=(2, 4))
        ttk.Separator(frame_droite).pack(fill="x", pady=(0, 8))

        ttk.Button(
            frame_droite,
            text="📊 État d'avancement",
            command=lambda: self._ouvrir_premier_fichier_chantier("Etat_avancement*.xlsm")
        ).pack(fill="x", pady=3)

        ttk.Button(frame_droite, text="📊 Calcul état", command=self.calcul_etat_avancement).pack(fill="x", pady=3)
        btn_cloturer = tk.Label(
            frame_droite,
            text="♻️ Clôturer état",
            bg="#EF8F8F",
            fg="black",
            font=("Helvetica", 14, "bold"),
            relief="flat",
            bd=0,
            padx=8,
            pady=6,
            cursor="hand2",
        )
        btn_cloturer.bind("<Button-1>", lambda _event: self.mise_a_zero_etat())
        btn_cloturer.pack(fill="x", pady=3)

        ttk.Separator(frame_droite).pack(fill="x", pady=10)

        tk.Label(frame_droite, text="PILOTAGE",
         font=("Helvetica", 15, "bold"),
         fg="#0052cc").pack(anchor="w", pady=(2, 4))
        ttk.Separator(frame_droite).pack(fill="x", pady=(0, 8))

        ttk.Button(frame_droite, text="🧭 Pilotage", command=self.pilotage_chantier).pack(fill="x", pady=3)
        ttk.Button(frame_droite, text="📊 Pilotage délai", command=self.pilotage_delai).pack(fill="x", pady=3)
        ttk.Button(frame_droite, text="⏱ Rendements", command=self.rendements_chantier).pack(fill="x", pady=3)
       
       

    def ouvrir_aide(self) -> None:
        existing = getattr(self, "_help_window", None)
        if existing and existing.winfo_exists():
            existing.deiconify()
            existing.lift()
            existing.focus_force()
            return

        win = tk.Toplevel(self)
        self._help_window = win
        win.title("Aide / Mode d’emploi - Horizon Chantier")
        win.geometry("900x600")
        win.minsize(820, 520)
        win.transient(self)
        topmost_var = tk.BooleanVar(value=True)

        def appliquer_toujours_visible() -> None:
            try:
                win.attributes("-topmost", topmost_var.get())
            except Exception:
                pass

        def fermer() -> None:
            self._help_window = None
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", fermer)
        appliquer_toujours_visible()

        header = ttk.Frame(win, padding=(18, 16, 18, 10))
        header.pack(fill="x")

        ttk.Label(
            header,
            text="Aide / Mode d’emploi - Horizon Chantier",
            font=("Helvetica", 20, "bold"),
            foreground="#163A59",
        ).pack(anchor="w")

        ttk.Label(
            header,
            text="Choisir le guide à afficher puis suivre les consignes sans modifier les formules Excel.",
            font=("Helvetica", 10),
            foreground="#4F6475",
        ).pack(anchor="w", pady=(4, 0))

        zone_choix = ttk.Frame(header)
        zone_choix.pack(anchor="w", pady=(10, 0))
        ttk.Label(zone_choix, text="Aide :", font=("Helvetica", 10, "bold")).pack(side="left", padx=(0, 8))
        aide_var = tk.StringVar(value="Mode d’emploi")
        choix_aide = ttk.Combobox(
            zone_choix,
            textvariable=aide_var,
            values=["Mode d’emploi", "Aide en cas de blocage"],
            state="readonly",
            width=24,
        )
        choix_aide.pack(side="left")

        ttk.Separator(win).pack(fill="x")

        corps = ttk.Frame(win, padding=(14, 12, 14, 8))
        corps.pack(fill="both", expand=True)
        corps.columnconfigure(0, weight=1)
        corps.rowconfigure(0, weight=1)

        texte = tk.Text(
            corps,
            wrap="word",
            font=("Helvetica", 11),
            relief="flat",
            bd=0,
            padx=22,
            pady=18,
            spacing1=2,
            spacing3=7,
            background="#FFFFFF",
            foreground="#1E1E1E",
            insertwidth=0,
        )
        texte.grid(row=0, column=0, sticky="nsew")

        scroll = ttk.Scrollbar(corps, orient="vertical", command=texte.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        texte.configure(yscrollcommand=scroll.set)

        texte.tag_configure("titre", font=("Helvetica", 19, "bold"), foreground="#163A59", spacing3=14)
        texte.tag_configure("intro", font=("Helvetica", 11), foreground="#405261", spacing3=10)
        texte.tag_configure("section", font=("Helvetica", 14, "bold"), foreground="#1F4E79", spacing1=14, spacing3=8)
        texte.tag_configure("texte", font=("Helvetica", 11), foreground="#1E1E1E", spacing3=6)
        texte.tag_configure("liste", font=("Helvetica", 11), foreground="#1E1E1E", lmargin1=22, lmargin2=42, spacing3=4)
        texte.tag_configure("regle", font=("Helvetica", 12, "bold"), foreground="#8A1C1C", background="#FFF3CD", spacing1=8, spacing3=10, lmargin1=10, lmargin2=10)
        texte.tag_configure("formule", font=("Courier", 11, "bold"), foreground="#163A59", spacing3=4, lmargin1=22, lmargin2=42)
        texte.tag_configure("ok", font=("Helvetica", 11, "bold"), foreground="#1E6B34", spacing3=4)
        texte.tag_configure("attention", font=("Helvetica", 11, "bold"), foreground="#9C4F00", spacing3=4)
        texte.tag_configure("danger", font=("Helvetica", 11, "bold"), foreground="#B42318", spacing3=4)
        texte.tag_configure("aide", font=("Helvetica", 11, "bold"), foreground="#5B3F99", spacing3=4)
        texte.tag_configure("fin", font=("Helvetica", 10, "italic"), foreground="#4F6475", spacing1=8, spacing3=10)

        contenu_mode_emploi = [
            ("titre", "Mode d’emploi par bouton - Horizon Chantier\n"),
            ("intro", "Ce guide reprend les boutons exactement comme ils apparaissent dans Horizon Chantier. Lire la section du bouton utilisé.\n"),
            ("regle", "RÈGLE PRINCIPALE : ne jamais modifier les formules, validations, feuilles, colonnes ou structures des fichiers Excel.\n"),
            ("section", "\nRègle générale de saisie\n"),
            ("regle", "Dans tous les fichiers Excel utilisés par Horizon Chantier, les cellules grisées sont les zones de saisie prévues pour l’utilisateur.\n"),
            ("liste", "✅ L’utilisateur peut uniquement compléter ou modifier ces zones grisées.\n"),
            ("liste", "❌ Toutes les autres cellules contiennent des calculs, des formules ou des informations générées par le logiciel.\n"),
            ("liste", "❌ Elles ne doivent jamais être modifiées manuellement.\n"),
            ("liste", "✅ L’utilisateur n’a pas à contrôler ou corriger les calculs du logiciel.\n"),
            ("liste", "✅ Il doit uniquement introduire les données demandées dans les zones prévues.\n"),
            ("attention", "⚠️ Cette règle est une base obligatoire pour utiliser correctement Horizon Chantier.\n"),

            ("section", "\nAIDE\n"),
            ("section", "\nAide / Dépannage\n"),
            ("texte", "Ouvre cette fenêtre de mode d’emploi et de dépannage.\n"),

            ("section", "\nGESTION\n"),
            ("section", "\n📄 Ouvrir dans le logiciel\n"),
            ("texte", "Ouvre la fiche du chantier sélectionné dans Horizon Chantier.\n"),
            ("liste", "✅ Sélectionner un chantier dans la liste avant de cliquer.\n"),
            ("liste", "✅ Utiliser ce bouton pour consulter les informations du chantier dans le logiciel.\n"),

            ("section", "\n➕ Nouveau chantier\n"),
            ("texte", "Crée un nouveau chantier complet avec les documents de travail prévus.\n"),
            ("liste", "✅ Renseigner le client, le nom du chantier et les informations demandées.\n"),
            ("liste", "✅ Horizon prépare automatiquement le dossier et les documents du chantier.\n"),
            ("liste", "❌ Ne pas créer manuellement un dossier chantier à la place du logiciel.\n"),

            ("section", "\n✏️ Modifier\n"),
            ("texte", "Permet de modifier les informations générales du chantier sélectionné.\n"),
            ("liste", "✅ Utiliser ce bouton pour mettre à jour les informations administratives du chantier.\n"),
            ("liste", "❌ Ne pas utiliser ce bouton pour modifier les fichiers Excel.\n"),

            ("section", "\n🗑️ Supprimer\n"),
            ("texte", "Supprime le chantier sélectionné si l’action est confirmée.\n"),
            ("liste", "⚠️ À utiliser uniquement si la suppression est réellement voulue.\n"),
            ("liste", "❌ Ne pas supprimer un chantier actif ou contenant des documents utiles.\n"),

            ("section", "\n🔄 Rafraîchir\n"),
            ("texte", "Recharge la liste des chantiers affichés dans Horizon Chantier.\n"),
            ("liste", "✅ À utiliser après création, modification ou si un chantier n’apparaît pas immédiatement.\n"),

            ("section", "\nSauvegarde PDF Historique\n"),
            ("texte", "Crée le PDF de sécurité du document Excel sur lequel vous venez de travailler.\n"),
            ("liste", "✅ Ouvrir le fichier Excel du chantier.\n"),
            ("liste", "✅ Modifier le fichier dans Excel.\n"),
            ("liste", "✅ Enregistrer le fichier dans Excel.\n"),
            ("liste", "✅ Revenir dans Horizon Chantier.\n"),
            ("liste", "✅ Cliquer sur Sauvegarde PDF Historique.\n"),
            ("liste", "✅ Le PDF est créé dans Historique_Sauvegardes avec la date, l’heure et le nom du document.\n"),
            ("attention", "⚠️ Si plusieurs fichiers Excel sont ouverts, afficher d’abord le bon document Excel avant de lancer la sauvegarde PDF.\n"),

            ("section", "\nDOCUMENTS\n"),
            ("section", "\n📁 Ce chantier\n"),
            ("texte", "Ouvre le dossier complet du chantier sélectionné.\n"),
            ("liste", "✅ Cliquer sur Ce chantier.\n"),
            ("liste", "✅ Choisir le fichier Excel ou document souhaité dans le dossier.\n"),
            ("liste", "✅ Travailler dans Excel puis enregistrer.\n"),
            ("liste", "✅ Revenir dans Horizon Chantier et cliquer sur Sauvegarde PDF Historique si le fichier Excel a été modifié.\n"),

            ("section", "\n🔎 Dossier administratif\n"),
            ("texte", "Ouvre le dossier Administratif du chantier.\n"),
            ("liste", "✅ Utiliser ce bouton pour accéder aux documents administratifs du chantier.\n"),

            ("section", "\n📄 Cahier des charges administratif\n"),
            ("texte", "Ouvre le fichier Cahier_administratif.pdf du chantier.\n"),

            ("section", "\n🛠 Cahier des charges technique\n"),
            ("texte", "Ouvre le fichier Cahier_technique.pdf du chantier.\n"),

            ("section", "\n📐 Postes / Métré\n"),
            ("texte", "Ouvre le fichier Metre_detaille.pdf du chantier.\n"),

            ("section", "\n🦺 Plan de sécurité (PSS)\n"),
            ("texte", "Ouvre le fichier PSS.pdf du chantier.\n"),

            ("section", "\n📄 Décompte intempéries\n"),
            ("texte", "Ouvre le document Décompte_intempéries.doc du chantier.\n"),

            ("section", "\nEXECUTION\n"),
            ("section", "\n📅 Planning d’exécution\n"),
            ("texte", "Ouvre le planning d’exécution du chantier.\n"),
            ("liste", "✅ Modifier uniquement les zones prévues dans Excel.\n"),
            ("liste", "✅ Enregistrer puis créer une Sauvegarde PDF Historique si le planning a été modifié.\n"),

            ("section", "\nCALCULS\n"),
            ("section", "\n🔎 Coût de la sécurité\n"),
            ("texte", "Ouvre le document de coût de la sécurité du chantier.\n"),
            ("liste", "✅ Encoder uniquement les zones de saisie prévues.\n"),
            ("liste", "✅ Enregistrer puis créer une Sauvegarde PDF Historique après modification.\n"),

            ("section", "\n📐 Formule de révision\n"),
            ("texte", "Ouvre le document de formule de révision du chantier.\n"),
            ("liste", "✅ Utiliser les zones prévues pour les informations de révision.\n"),
            ("liste", "❌ Ne pas modifier les formules, la structure ou les protections du document.\n"),

            ("section", "\n💰 Prix de revient\n"),
            ("texte", "Ouvre le Prix de revient du chantier.\n"),
            ("liste", "✅ Encoder les informations de PR uniquement dans les zones prévues.\n"),
            ("liste", "✅ Enregistrer le PR avant Vérifier PR ou Calcul / Reca.\n"),
            ("liste", "❌ Ne pas modifier les formules, les feuilles de bibliothèque ou les listes déroulantes.\n"),

            ("section", "\n🔍 Vérifier PR\n"),
            ("texte", "Lance le contrôle du Prix de revient enregistré.\n"),
            ("liste", "✅ À lancer après avoir enregistré le PR dans Excel.\n"),
            ("liste", "✅ Si un message apparaît, lire le message et reprendre uniquement les zones de saisie prévues.\n"),
            ("liste", "❌ Ne pas modifier une formule pour faire disparaître une erreur.\n"),

            ("section", "\n⚙️ Calcul / Reca\n"),
            ("texte", "Lance la mise à jour prévue par Horizon Chantier à partir des documents enregistrés.\n"),
            ("liste", "✅ Enregistrer le PR et les fichiers concernés avant de cliquer.\n"),
            ("liste", "⚠️ Si Excel empêche l’action, enregistrer et fermer le document Excel puis relancer une seule fois.\n"),

            ("section", "\nETAT D'AVANCEMENT\n"),
            ("section", "\n📊 État d'avancement\n"),
            ("texte", "Ouvre l’État d’avancement du chantier.\n"),
            ("liste", "✅ Encoder les avancements et informations métier dans les zones prévues.\n"),
            ("liste", "✅ La synthèse reste une zone libre pour les ajustements métier prévus.\n"),
            ("liste", "✅ Enregistrer avant Calcul état ou Clôturer état.\n"),

            ("section", "\n📊 Calcul état\n"),
            ("texte", "Demande à Horizon Chantier de mettre à jour l’état d’avancement enregistré.\n"),
            ("liste", "✅ À lancer après avoir enregistré l’état d’avancement dans Excel.\n"),
            ("liste", "❌ Ne pas modifier les cellules calculées pour forcer un résultat.\n"),

            ("section", "\n♻️ Clôturer état\n"),
            ("texte", "Clôture l’état d’avancement terminé pour préparer l’état suivant.\n"),
            ("liste", "✅ À utiliser uniquement lorsque la période est validée.\n"),
            ("danger", "⚠️⚠️ ATTENTION - OPÉRATION IRRÉVERSIBLE ⚠️⚠️\n"),
            ("danger", "La fonction Clôturer état réalise une mise à zéro de l’état d’avancement pour préparer l’état suivant.\n"),
            ("danger", "Avant de lancer la clôture, il est OBLIGATOIRE de créer une Sauvegarde PDF Historique du document.\n"),
            ("liste", "✅ Vérifier que l’état d’avancement est terminé et enregistré.\n"),
            ("liste", "✅ Créer une Sauvegarde PDF Historique du document.\n"),
            ("liste", "✅ Vérifier que le PDF a bien été créé dans le dossier Historique_Sauvegardes.\n"),
            ("liste", "✅ Seulement après cette vérification, lancer le bouton Clôturer état.\n"),
            ("danger", "Si la clôture est réalisée sans PDF Historique : les quantités du mois sont remises à zéro, l’état précédent n’est plus récupérable sous sa forme de travail, et il est impossible de revenir en arrière.\n"),
            ("danger", "RÈGLE ABSOLUE : aucune clôture ne doit être réalisée sans cette sauvegarde préalable.\n"),

            ("section", "\nPILOTAGE\n"),
            ("section", "\n🧭 Pilotage\n"),
            ("texte", "Affiche ou génère la synthèse de pilotage du chantier.\n"),
            ("liste", "✅ À lancer après enregistrement et calcul des fichiers concernés.\n"),

            ("section", "\n📊 Pilotage délai\n"),
            ("texte", "Affiche ou génère la synthèse du délai chantier.\n"),
            ("liste", "✅ Vérifier que le fichier délai ou planning est à jour avant de cliquer.\n"),

            ("section", "\n⏱ Rendements\n"),
            ("texte", "Affiche ou génère la lecture des rendements du chantier.\n"),
            ("liste", "✅ Vérifier que les données de rendement sont enregistrées avant de cliquer.\n"),

            ("section", "\nRègles définitives sur les fichiers Excel\n"),
            ("liste", "❌ Ne pas modifier la structure des modèles Excel.\n"),
            ("liste", "❌ Ne pas modifier les formules.\n"),
            ("liste", "❌ Ne pas modifier les validations de données ou listes déroulantes.\n"),
            ("liste", "❌ Ne pas renommer les feuilles.\n"),
            ("liste", "✅ Les cellules de calcul doivent être préservées.\n"),
            ("liste", "✅ Les zones de saisie prévues par le logiciel sont les seules zones à modifier.\n"),
            ("liste", "✅ Respecter les zones de saisie prévues garantit une utilisation correcte du logiciel.\n"),

            ("section", "\nBon réflexe de fin de travail Excel\n"),
            ("ok", "✅ Enregistrer Excel.\n"),
            ("ok", "✅ Revenir dans Horizon Chantier.\n"),
            ("ok", "✅ Cliquer sur Sauvegarde PDF Historique.\n"),
            ("ok", "✅ Vérifier que le PDF daté est présent dans Historique_Sauvegardes.\n"),
            ("danger", "❌ Ne jamais attendre plusieurs jours avant de créer le PDF Historique d’un document important.\n"),
            ("fin", "\nFin du mode d’emploi.\n"),
        ]

        contenu_depannage_complet = [
            ("titre", "Aide en cas de blocage - Horizon Chantier\n"),
            ("intro", "Cette aide indique quoi faire si une action ne se passe pas comme prévu. L’utilisateur vérifie les étapes de base et demande de l’aide si le problème continue.\n"),
            ("regle", "RÈGLE : en cas de blocage, ne jamais modifier une formule, une structure Excel ou un fichier modèle.\n"),

            ("section", "\n📄 Ouvrir dans le logiciel\n"),
            ("liste", "✅ Si rien ne s’ouvre, vérifier qu’un chantier est sélectionné dans la liste.\n"),
            ("liste", "✅ Cliquer sur Rafraîchir puis réessayer.\n"),
            ("liste", "✅ Si le problème continue, noter le nom du chantier et demander de l’aide.\n"),

            ("section", "\n➕ Nouveau chantier\n"),
            ("liste", "✅ Si le chantier existe déjà, choisir un autre nom ou vérifier le dossier existant.\n"),
            ("liste", "✅ Si la création échoue, noter le message affiché.\n"),
            ("liste", "❌ Ne pas créer manuellement un dossier chantier pour contourner l’erreur.\n"),

            ("section", "\n✏️ Modifier\n"),
            ("liste", "✅ Vérifier qu’un chantier est sélectionné.\n"),
            ("liste", "✅ Modifier uniquement les informations générales demandées par la fenêtre.\n"),

            ("section", "\n🗑️ Supprimer\n"),
            ("liste", "⚠️ Vérifier que le bon chantier est sélectionné avant confirmation.\n"),
            ("liste", "❌ Ne pas supprimer un chantier actif ou non sauvegardé.\n"),

            ("section", "\n🔄 Rafraîchir\n"),
            ("liste", "✅ À utiliser si un chantier n’apparaît pas ou si la liste semble ancienne.\n"),
            ("liste", "✅ Si le chantier reste absent, demander de l’aide avec le nom exact du chantier.\n"),

            ("section", "\nSauvegarde PDF Historique\n"),
            ("liste", "✅ Vérifier qu’un fichier Excel est ouvert.\n"),
            ("liste", "✅ Si plusieurs classeurs sont ouverts, afficher le bon document Excel avant de revenir dans Horizon.\n"),
            ("liste", "✅ Enregistrer le fichier dans Excel avant de créer le PDF.\n"),
            ("liste", "✅ Vérifier que le PDF apparaît dans Historique_Sauvegardes du chantier.\n"),
            ("liste", "⚠️ Si Excel affiche une fenêtre, la lire et la fermer correctement avant de relancer la sauvegarde PDF.\n"),
            ("liste", "❌ Ne pas déplacer ou renommer le fichier Excel pendant la génération du PDF.\n"),

            ("section", "\n📁 Ce chantier\n"),
            ("liste", "✅ Vérifier qu’un chantier est bien sélectionné dans la liste.\n"),
            ("liste", "✅ Si le dossier ne s’ouvre pas, cliquer sur Rafraîchir puis réessayer.\n"),
            ("liste", "✅ Si le dossier ne s’ouvre toujours pas, demander de l’aide avec le nom du chantier.\n"),
            ("liste", "✅ Après ouverture d’un Excel depuis ce dossier, revenir dans Horizon pour Sauvegarde PDF Historique.\n"),

            ("section", "\n🔎 Dossier administratif\n"),
            ("liste", "✅ Si le dossier ne s’ouvre pas, revenir à Ce chantier pour accéder aux documents.\n"),
            ("liste", "✅ Si le document manque, demander de l’aide avec le nom du chantier.\n"),

            ("section", "\n📄 Cahier des charges administratif\n"),
            ("liste", "✅ Si le document ne s’ouvre pas, essayer d’abord Ce chantier.\n"),
            ("liste", "✅ Si le PDF ne s’ouvre pas, vérifier l’application PDF par défaut du poste.\n"),

            ("section", "\n🛠 Cahier des charges technique\n"),
            ("liste", "✅ Si le document ne s’ouvre pas, essayer d’abord Ce chantier.\n"),

            ("section", "\n📐 Postes / Métré\n"),
            ("liste", "✅ Si le document ne s’ouvre pas, essayer d’abord Ce chantier.\n"),

            ("section", "\n🦺 Plan de sécurité (PSS)\n"),
            ("liste", "✅ Si le document ne s’ouvre pas, essayer d’abord Ce chantier.\n"),

            ("section", "\n📄 Décompte intempéries\n"),
            ("liste", "✅ Si le document ne s’ouvre pas, essayer d’abord Ce chantier.\n"),
            ("liste", "✅ Si le document ne s’ouvre pas, vérifier l’application associée aux fichiers .doc.\n"),

            ("section", "\n📅 Planning d’exécution\n"),
            ("liste", "✅ Si le planning ne s’ouvre pas, essayer d’abord Ce chantier.\n"),
            ("liste", "✅ Enregistrer Excel avant de créer la Sauvegarde PDF Historique.\n"),
            ("liste", "❌ Ne pas modifier les formules ou la structure du planning.\n"),

            ("section", "\n🔎 Coût de la sécurité\n"),
            ("liste", "✅ Si le document ne s’ouvre pas, essayer d’abord Ce chantier.\n"),
            ("liste", "✅ Modifier uniquement les zones prévues puis enregistrer.\n"),

            ("section", "\n📐 Formule de révision\n"),
            ("liste", "✅ Si le document ne s’ouvre pas, essayer d’abord Ce chantier.\n"),
            ("liste", "✅ Si Excel bloque les macros ou l’ouverture, suivre le message Excel sans modifier le fichier.\n"),

            ("section", "\n💰 Prix de revient\n"),
            ("liste", "✅ Si le PR ne s’ouvre pas, essayer d’abord Ce chantier.\n"),
            ("liste", "✅ Enregistrer le PR avant Vérifier PR ou Calcul / Reca.\n"),
            ("liste", "✅ En cas de doute sur une valeur, ne pas modifier les formules et demander confirmation.\n"),
            ("liste", "❌ Ne pas modifier les formules, bibliothèques ou validations.\n"),

            ("section", "\n🔍 Vérifier PR\n"),
            ("liste", "✅ Enregistrer le PR dans Excel avant de cliquer sur Vérifier PR.\n"),
            ("liste", "✅ Si un message apparaît, lire le message et reprendre uniquement les zones de saisie prévues.\n"),
            ("liste", "✅ Si le fichier est verrouillé, fermer Excel puis relancer l’action.\n"),
            ("liste", "❌ Ne pas modifier une formule pour changer un résultat.\n"),

            ("section", "\n⚙️ Calcul / Reca\n"),
            ("liste", "✅ Vérifier que le bon chantier est sélectionné.\n"),
            ("liste", "✅ Enregistrer le PR et l’état d’avancement avant de cliquer.\n"),
            ("liste", "✅ Si l’action ne se termine pas, fermer Excel puis relancer une seule fois.\n"),
            ("liste", "✅ Si le problème continue, noter le message affiché et demander de l’aide.\n"),

            ("section", "\n📊 État d'avancement\n"),
            ("liste", "✅ Si l’état ne s’ouvre pas, essayer d’abord Ce chantier.\n"),
            ("liste", "✅ Enregistrer Excel avant Calcul état ou Clôturer état.\n"),
            ("liste", "✅ La synthèse reste libre pour les ajustements métier prévus.\n"),
            ("liste", "❌ Ne pas écrire dans les cellules de calcul pour forcer un total.\n"),

            ("section", "\n📊 Calcul état\n"),
            ("liste", "✅ Enregistrer l’état d’avancement dans Excel avant de cliquer.\n"),
            ("liste", "✅ Si un résultat paraît anormal, ne pas toucher aux formules : contrôler uniquement les zones de saisie prévues ou demander de l’aide.\n"),
            ("liste", "✅ Relancer une seule fois après avoir enregistré Excel.\n"),
            ("liste", "❌ Ne pas modifier les formules.\n"),

            ("section", "\n♻️ Clôturer état\n"),
            ("liste", "✅ Utiliser ce bouton uniquement quand la période est validée.\n"),
            ("danger", "⚠️⚠️ ATTENTION - OPÉRATION IRRÉVERSIBLE ⚠️⚠️\n"),
            ("danger", "Clôturer état remet à zéro l’état d’avancement pour préparer l’état suivant.\n"),
            ("liste", "✅ Avant toute clôture, vérifier que l’état d’avancement est terminé et enregistré.\n"),
            ("liste", "✅ Créer une Sauvegarde PDF Historique du document.\n"),
            ("liste", "✅ Vérifier que le PDF a bien été créé dans Historique_Sauvegardes.\n"),
            ("liste", "✅ Seulement ensuite, lancer Clôturer état.\n"),
            ("danger", "Sans PDF Historique, les quantités du mois sont remises à zéro et l’état précédent n’est plus récupérable sous sa forme de travail.\n"),
            ("danger", "Aucune clôture ne doit être réalisée sans cette sauvegarde préalable.\n"),
            ("liste", "✅ Vérifier que le bon chantier et le bon état sont sélectionnés.\n"),

            ("section", "\n🧭 Pilotage\n"),
            ("liste", "✅ Enregistrer les fichiers Excel concernés avant de cliquer.\n"),
            ("liste", "✅ Si la synthèse paraît incorrecte, vérifier que les documents concernés ont bien été enregistrés.\n"),

            ("section", "\n📊 Pilotage délai\n"),
            ("liste", "✅ Vérifier que les données délai/planning sont à jour et enregistrées.\n"),
            ("liste", "✅ Si le résultat paraît faux, vérifier que le document délai/planning est enregistré puis demander de l’aide si nécessaire.\n"),

            ("section", "\n⏱ Rendements\n"),
            ("liste", "✅ Vérifier que les données de rendement sont enregistrées.\n"),
            ("liste", "✅ Si les heures ou écarts semblent faux, vérifier que le document de rendement est enregistré puis demander de l’aide si nécessaire.\n"),

            ("section", "\nProblèmes généraux d’ouverture de documents\n"),
            ("liste", "✅ Vérifier que le bon chantier est sélectionné.\n"),
            ("liste", "✅ Essayer d’ouvrir le document avec Ce chantier.\n"),
            ("liste", "✅ Vérifier qu’aucune fenêtre Excel ne bloque l’ouverture.\n"),
            ("liste", "✅ Fermer Excel puis relancer l’ouverture depuis Horizon si nécessaire.\n"),

            ("section", "\nProblèmes généraux liés aux fichiers Excel\n"),
            ("liste", "⚠️ Si Excel signale un fichier en lecture seule, vérifier qu’il n’est pas déjà ouvert ailleurs.\n"),
            ("liste", "⚠️ Si Excel demande une récupération, enregistrer une copie avant toute manipulation risquée.\n"),
            ("liste", "✅ Ne modifier que les zones de saisie prévues.\n"),
            ("liste", "❌ Ne jamais supprimer une feuille, une colonne, une validation ou une formule.\n"),
            ("liste", "❌ Ne pas enregistrer un fichier métier sous un autre nom sans consigne.\n"),

            ("section", "\nFenêtre de rappel PDF\n"),
            ("liste", "✅ Le rappel apparaît lors de l’ouverture d’un document de travail depuis le flux habituel.\n"),
            ("liste", "✅ Avec Ce chantier, le rappel s’affiche après l’ouverture du dossier afin de rappeler la sauvegarde PDF avant le travail Excel.\n"),
            ("liste", "✅ La fenêtre reste affichée jusqu’au clic sur ✓ OK, j’ai compris.\n"),
            ("liste", "🔍 Si elle est derrière une autre fenêtre, revenir sur Horizon ou vérifier le côté de l’écran.\n"),

            ("section", "\nCellules protégées ou formules invisibles\n"),
            ("liste", "✅ Les cellules contenant des formules sont protégées pour éviter les erreurs de manipulation.\n"),
            ("liste", "✅ Les formules peuvent être masquées dans la barre de formule lorsque la feuille est protégée.\n"),
            ("liste", "✅ Les zones de saisie prévues restent modifiables.\n"),
            ("liste", "✅ La synthèse de l’état d’avancement reste libre pour les ajustements métier prévus.\n"),
            ("liste", "❌ Ne pas enlever les protections pour travailler plus vite.\n"),

            ("section", "\nErreurs courantes\n"),
            ("liste", "❌ Travailler sur le mauvais chantier sélectionné.\n"),
            ("liste", "❌ Oublier d’enregistrer Excel avant un calcul ou un PDF Historique.\n"),
            ("liste", "❌ Ouvrir plusieurs fichiers Excel et laisser le mauvais classeur actif avant la sauvegarde PDF.\n"),
            ("liste", "❌ Renommer un fichier utilisé par Horizon.\n"),
            ("liste", "❌ Copier-coller sur des cellules de calcul.\n"),
            ("liste", "❌ Modifier le dossier Modèles pour intervenir sur un chantier existant.\n"),

            ("section", "\nProcédure générale en cas de blocage\n"),
            ("liste", "1. Arrêter les modifications hasardeuses.\n"),
            ("liste", "2. Identifier le chantier sélectionné et le fichier concerné.\n"),
            ("liste", "3. Enregistrer Excel si possible.\n"),
            ("liste", "4. Créer une Sauvegarde PDF Historique si le document est lisible et important.\n"),
            ("liste", "5. Fermer les boîtes de dialogue Excel puis relancer une seule fois l’action.\n"),
            ("liste", "6. Si le problème persiste, noter le message affiché, le bouton utilisé et le nom du chantier.\n"),
            ("danger", "❌ Ne jamais modifier une formule, une validation, une feuille ou une structure pour dépanner dans l’urgence.\n"),
            ("fin", "\nFin du dépannage complet.\n"),
        ]

        def afficher_aide(*_):
            texte.configure(state="normal")
            texte.delete("1.0", "end")

            if aide_var.get() == "Aide en cas de blocage":
                contenu = contenu_depannage_complet
            else:
                contenu = contenu_mode_emploi

            for tag, bloc in contenu:
                texte.insert("end", bloc, tag)

            texte.configure(state="disabled")
            texte.yview_moveto(0)

        choix_aide.bind("<<ComboboxSelected>>", afficher_aide)
        afficher_aide()

        footer = ttk.Frame(win, padding=(14, 0, 14, 14))
        footer.pack(fill="x")
        ttk.Checkbutton(
            footer,
            text="Toujours visible",
            variable=topmost_var,
            command=appliquer_toujours_visible,
        ).pack(side="left")
        ttk.Button(footer, text="Fermer", command=fermer).pack(side="right")

    def ouvrir_aide_mode_emploi(self) -> None:
        self.ouvrir_aide()

    def calcul_etat_avancement(self):
        nom_chantier = self._nom_chantier_selectionne()
        if not nom_chantier:
            return
        chemin = next(
            Path(
                dossier_chantiers() / nom_chantier
            ).glob("Etat_avancement*.xlsm")
        ).as_posix()

        feuille = "Bordereau"

        _backup_excel_before_write(chemin)
        wb = load_workbook(chemin, keep_vba=True)
        ws = wb[feuille]

        def normaliser_libelle(valeur):
            texte = str(valeur or "").strip().lower()
            texte = texte.replace("é", "e").replace("è", "e").replace("ê", "e").replace("ë", "e")
            texte = texte.replace("à", "a").replace("â", "a").replace("ä", "a")
            texte = texte.replace("ù", "u").replace("û", "u").replace("ü", "u")
            texte = texte.replace("î", "i").replace("ï", "i")
            texte = texte.replace("ô", "o").replace("ö", "o")
            texte = texte.replace("ç", "c")
            texte = texte.replace("\n", " ")
            texte = re.sub(r"\s+", " ", texte)
            return texte

        def trouver_colonne_bordereau(alias):
            alias_normalises = {normaliser_libelle(a) for a in alias}
            max_ligne_entete = min(15, ws.max_row)
            for ligne in range(1, max_ligne_entete + 1):
                for col in range(1, ws.max_column + 1):
                    libelle = normaliser_libelle(ws.cell(ligne, col).value)
                    if libelle in alias_normalises or any(a in libelle for a in alias_normalises):
                        return ws.cell(ligne, col).column_letter
            return None

        col_quantite = trouver_colonne_bordereau({"quantite", "qte", "qte.", "qte :", "qte/", "qté"})
        col_prix_unitaire = trouver_colonne_bordereau({"prix unitaire", "pu", "p.u.", "prix unit"})

        if not col_quantite or not col_prix_unitaire:
            erreurs = []
            if not col_quantite:
                erreurs.append("colonne quantité introuvable (attendu : quantité, quantite, qté, qte)")
            if not col_prix_unitaire:
                erreurs.append("colonne prix unitaire introuvable (attendu : prix unitaire, pu, p.u., prix unit)")
            wb.close()
            messagebox.showerror("Calcul état", "Impossible de recalculer la feuille Bordereau.\n" + "\n".join(erreurs))
            return

        def nombre(valeur):
            if valeur in (None, "", "-", "—"):
                return 0
            try:
                return float(str(valeur).replace("€", "").replace(" ", "").replace(",", "."))
            except:
                return 0

        def ligne_bordereau_valide(ligne):
            c = ws[f"C{ligne}"].value
            j = ws[f"{col_quantite}{ligne}"].value
            l = ws[f"{col_prix_unitaire}{ligne}"].value
            p = ws[f"P{ligne}"].value
            q = ws[f"Q{ligne}"].value

            return any(v not in (None, "", "-", "—") for v in [c, j, l, p, q])

        def derniere_ligne_bordereau():
            derniere = 24
            lignes_vides = 0

            for ligne in range(24, ws.max_row + 1):
                if ligne_bordereau_valide(ligne):
                    derniere = ligne
                    lignes_vides = 0
                else:
                    lignes_vides += 1

                    # dès qu'on a plusieurs lignes vides d'affilée,
                    # on considère que le bordereau est terminé
                    if lignes_vides >= 5:
                        break

            return derniere

        fin = derniere_ligne_bordereau()

        for ligne in range(24, fin + 1):
            if not ligne_bordereau_valide(ligne):
                ws[f"S{ligne}"] = None
                continue

            j = nombre(ws[f"{col_quantite}{ligne}"].value)
            l = nombre(ws[f"{col_prix_unitaire}{ligne}"].value)
            p = nombre(ws[f"P{ligne}"].value)
            q = nombre(ws[f"Q{ligne}"].value)

            r = p + q
            s = 0 if j == 0 else r / j
            t = r * l
            u = q
            v = l
            w = u * v

            if j == 0 and l == 0:
                continue

            ws[f"R{ligne}"] = r
            ws[f"S{ligne}"] = f'=IF(AND(OR(P{ligne}="",P{ligne}=0),OR(Q{ligne}="",Q{ligne}=0),OR(R{ligne}="",R{ligne}=0)),"",IFERROR(R{ligne}/J{ligne},0))'
            ws[f"S{ligne}"].number_format = "0.00%"
            ws[f"T{ligne}"] = t
            ws[f"U{ligne}"] = u
            ws[f"V{ligne}"] = v
            ws[f"W{ligne}"] = w

        

        _protect_formula_cells(wb, chemin)
        wb.save(chemin)
        print("Etat recalculé")


       

    def mise_a_zero_etat(self):
        if not self._confirmer_cloture_etat():
            return

        nom_chantier = self._nom_chantier_selectionne()
        if not nom_chantier:
            return
        chemin = next(
            Path(
                dossier_chantiers() / nom_chantier
            ).glob("Etat_avancement*.xlsm")
        ).as_posix()

        feuille = "Bordereau"

        _backup_excel_before_write(chemin)
        wb = load_workbook(chemin, keep_vba=True)
        ws = wb[feuille]

        def normaliser_libelle(valeur):
            texte = str(valeur or "").strip().lower()
            texte = texte.replace("é", "e").replace("è", "e").replace("ê", "e").replace("ë", "e")
            texte = texte.replace("à", "a").replace("â", "a").replace("ä", "a")
            texte = texte.replace("ù", "u").replace("û", "u").replace("ü", "u")
            texte = texte.replace("î", "i").replace("ï", "i")
            texte = texte.replace("ô", "o").replace("ö", "o")
            texte = texte.replace("ç", "c")
            texte = texte.replace("\n", " ")
            texte = re.sub(r"\s+", " ", texte)
            return texte

        libelles_synthese = (
            "total soumission hors tva",
            "total etat cumule hors tva",
            "total du mois hors tva",
            "total des avenants cumule",
            "montant global a facturer",
            "total execute",
        )

        def est_ligne_synthese(ligne):
            contenu = " ".join(
                normaliser_libelle(ws.cell(ligne, col).value)
                for col in range(1, ws.max_column + 1)
            )
            return any(libelle in contenu for libelle in libelles_synthese)

        def premiere_ligne_synthese():
            for ligne in range(24, ws.max_row + 1):
                if est_ligne_synthese(ligne):
                    return ligne
            return ws.max_row + 1

        def ligne_bordereau_a_cloturer(ligne):
            valeurs = [
                ws[f"B{ligne}"].value,
                ws[f"C{ligne}"].value,
                ws[f"J{ligne}"].value,
                ws[f"L{ligne}"].value,
                ws[f"P{ligne}"].value,
                ws[f"Q{ligne}"].value,
            ]
            return any(v not in (None, "", "-", "—") for v in valeurs)

        fin_bordereau = premiere_ligne_synthese() - 1

        for ligne in range(24, fin_bordereau + 1):
            if not ligne_bordereau_a_cloturer(ligne):
                continue
            ws[f"P{ligne}"] = ws[f"R{ligne}"].value
            ws[f"Q{ligne}"] = 0

        _protect_formula_cells(wb, chemin)
        wb.save(chemin)
        self.calcul_etat_avancement()
        print("Clôture état effectuée")

    def verifier_pr(self):
        p = self._chemin_chantier_selectionne()
        if not p:
            messagebox.showwarning("PR", "Sélectionne un chantier dans la liste.")
            return

        dossier_chantier = self._dossier_chantier_selectionne()
        chemin = dossier_chantier / "data" / "prix_de_revient.xlsx"

        import os
        from datetime import datetime

        if not chemin.exists():
            messagebox.showwarning(
                "Vérifier PR",
                "Le fichier prix_de_revient.xlsx est introuvable."
            )
            return

        timestamp = os.path.getmtime(chemin)
        date_modif = datetime.fromtimestamp(timestamp).strftime("%d/%m/%Y %H:%M:%S")

        messagebox.showinfo(
            "Date du PR",
            f"Dernier enregistrement du PR :\n{date_modif}\n\n"
            "⚠️ Si tu viens de modifier le PR :\n"
            "➡️ fais CMD + S dans Excel\n"
            "➡️ puis clique sur Vérifier PR"
        )
        
        wb = load_workbook(chemin, data_only=True)
        ws = wb["Chiffrage"]

        def nombre(v):
            if v in (None, "", "-", "—"):
                return 0
            try:
                return float(str(v).replace("€", "").replace(" ", "").replace(",", "."))
            except:
                return 0

        erreurs = []

        # MATIÈRE
        for ligne in range(3, 13):
            i = nombre(ws[f"I{ligne}"].value)
            j = nombre(ws[f"J{ligne}"].value)
            h = nombre(ws[f"H{ligne}"].value)

            attendu_k = i * (1 + j)
            attendu_l = h * attendu_k

            k = nombre(ws[f"K{ligne}"].value)
            l = nombre(ws[f"L{ligne}"].value)

            if abs(k - attendu_k) > 0.01:
                erreurs.append(f"Ligne {ligne} PU corrigé faux")

            if abs(l - attendu_l) > 0.01:
                erreurs.append(f"Ligne {ligne} Total matière faux")

        # MAIN-D’ŒUVRE
        for ligne in range(16, 26):
            c = nombre(ws[f"C{ligne}"].value)
            h = nombre(ws[f"H{ligne}"].value)
            j = nombre(ws[f"J{ligne}"].value)

            attendu_i = c * h
            attendu_k = attendu_i * j

            i = nombre(ws[f"I{ligne}"].value)
            k = nombre(ws[f"K{ligne}"].value)

            if abs(i - attendu_i) > 0.01:
                erreurs.append(f"Ligne {ligne} Heures fausses")

            if abs(k - attendu_k) > 0.01:
                erreurs.append(f"Ligne {ligne} Coût MO faux")

        if erreurs:
            messagebox.showwarning("PR", "\n".join(erreurs))
        else:
            messagebox.showinfo("PR", "✔ Tous les calculs sont corrects")

        wb.close()


            
    def pilotage_chantier(self):
        import os
        import subprocess
        import tempfile
        from datetime import datetime
        import matplotlib.pyplot as plt
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_RIGHT, TA_CENTER

        nom_chantier = self._nom_chantier_selectionne()
        if not nom_chantier:
            return

        chemin = next(
            Path(
                dossier_chantiers() / nom_chantier
            ).glob("Etat_avancement*.xlsm")
        ).as_posix()

        wb = load_workbook(chemin, data_only=True)
        ws = wb["Bordereau"]
        

        def nombre(val):
            if val in (None, "", "-", "—"):
                return 0
            try:
                texte_nombre = str(val)
                texte_nombre = texte_nombre.replace("€", "")
                texte_nombre = texte_nombre.replace("−", "-")
                texte_nombre = texte_nombre.replace(" ", "")
                texte_nombre = texte_nombre.replace("\xa0", "")
                texte_nombre = texte_nombre.replace("\u202f", "")
                texte_nombre = texte_nombre.replace(",", ".")
                return float(texte_nombre)
            except:
                return 0

        def texte(val):
            return str(val or "").strip().lower()

        def trouver_ligne_libelle(mots):
            for row in range(ws.max_row, 1, -1):
                for col in range(1, ws.max_column + 1):
                    val = texte(ws.cell(row=row, column=col).value)
                    if not val:
                        continue
                    if all(mot in val for mot in mots):
                        return row, col
            return None, None

        def premiere_valeur_a_droite(row, col_depart, max_ecart=3):
            if not row or not col_depart:
                return 0

            col_fin = min(col_depart + max_ecart, ws.max_column)

            for col in range(col_depart + 1, col_fin + 1):
                val = ws.cell(row=row, column=col).value

                # Ignore les cellules vides ou décoratives
                if val in (None, "", "-", "—"):
                    continue

                # Si Excel renvoie déjà un nombre, on le prend tel quel
                if isinstance(val, (int, float)):
                    return float(val)

                # Sinon on passe par le parseur existant
                num = nombre(val)
                txt = str(val).strip()

                # Conserve aussi un vrai zéro explicite
                if num != 0 or txt in {"0", "0,0", "0.0", "0,00", "0.00"}:
                    return num

            return 0

        def valeur_soumission_publique(libelles):
            for libelle in libelles:
                row, col = trouver_ligne_libelle([libelle])
                if row and col:
                    return premiere_valeur_a_droite(row, col), row, col
            return 0, None, None

        montant_soumission, ligne_soumission, col_soumission = valeur_soumission_publique(
            ["total soumission hors tva"]
        )
        _total_etat_cumule, _ligne_etat_cumule, _col_etat_cumule = valeur_soumission_publique(
            ["total état cumulé hors tva", "total etat cumulé hors tva"]
        )
        total_mois, ligne_mois, col_mois = valeur_soumission_publique(
            ["total du mois hors tva", "total du mois hors tva"]
        )
        avenants_plus, _ligne_avenants, _col_avenants = valeur_soumission_publique(
            ["total des avenants cumulé", "total des avenants cumule"]
        )
        total_realise, ligne_execute, col_execute = valeur_soumission_publique(
            ["total exécuté", "total execute"]
        )
        revision, _ligne_revision, _col_revision = valeur_soumission_publique(
            ["montant de la révision", "montant de la revision"]
        )
        revision_globale = revision
        _montant_global, _ligne_global, _col_global = valeur_soumission_publique(
            ["montant global à facturer", "montant global a facturer"]
        )

        if "Avenants" in wb.sheetnames:
            ws_avenants = wb["Avenants"]
            avenants_plus, avenants_moins = _lire_synthese_avenants_pilotage(
                ws_avenants,
                nombre,
            )
        else:
            avenants_moins = 0
            chemin_avenants = Path(chemin).with_name("Avenants.xlsx")
            if chemin_avenants.exists():
                wb_avenants = load_workbook(chemin_avenants, data_only=True)
                ws_avenants = wb_avenants["Avenants"] if "Avenants" in wb_avenants.sheetnames else wb_avenants.active
                avenants_plus, avenants_moins = _lire_synthese_avenants_pilotage(
                    ws_avenants,
                    nombre,
                )
                wb_avenants.close()
        total_marche = montant_soumission + avenants_plus - avenants_moins

        chemin_revision_globale = Path(chemin).with_name("Revision_global.xlsx")
        if chemin_revision_globale.exists():
            wb_revision_globale = load_workbook(chemin_revision_globale, data_only=True)
            ws_revision_globale = wb_revision_globale.active
            revision_globale = _lire_revision_globale_pilotage(ws_revision_globale, nombre)
            wb_revision_globale.close()

        print("avenants_plus =", avenants_plus)
        print("avenants_moins =", avenants_moins)

        print("ligne_mois =", ligne_mois, "col_mois =", col_mois)
        print("total_mois =", total_mois)
       

    
        date_du_jour = datetime.now().strftime("%d/%m/%Y")
        realise_mois = total_mois
        realise_cumule = total_realise
        production_mois = realise_mois + revision
        production_cumulee = realise_cumule + revision_globale
        reste_a_facturer = total_marche - production_cumulee
        avancement = 0 if total_marche == 0 else (production_cumulee / total_marche) * 100

        texte_popup = f"""PILOTAGE CHANTIER

       
    Chantier : {nom_chantier}
    Date     : {date_du_jour}

    Montant soumission              : {montant_soumission:,.2f} €
    Avenants cumulés en plus        : {avenants_plus:,.2f} €
    Avenants cumulés en moins       : {avenants_moins:,.2f} €
    Marché total                    : {total_marche:,.2f} €
    Réalisé du mois                 : {realise_mois:,.2f} €
    Révision renseignée             : {revision:,.2f} €
    Production du mois              : {production_mois:,.2f} €
    Réalisé cumulé                  : {realise_cumule:,.2f} €
    Production cumulée              : {production_cumulee:,.2f} €
    Reste à facturer                : {reste_a_facturer:,.2f} €
    Avancement                      : {avancement:.1f} %
    """

        pdf_path = str(Path(chemin).with_name("pilotage_chantier.pdf"))
        wb.close()

        doc = SimpleDocTemplate(
            pdf_path,
            pagesize=landscape(A4),
            leftMargin=8 * mm,
            rightMargin=8 * mm,
            topMargin=5 * mm,
            bottomMargin=5 * mm,
        )

        elements = []
        graphique_tmp = None
        logo_path = str(dossier_base() / "logo_jt_bati.png")

        style_titre = ParagraphStyle(
            "PilotageTitre",
            fontName="Helvetica-Bold",
            fontSize=18,
            leading=18,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#163A59"),
        )
        style_meta_label = ParagraphStyle(
            "PilotageMetaLabel",
            fontName="Helvetica-Bold",
            fontSize=8.5,
            leading=9,
            textColor=colors.HexColor("#36556F"),
        )
        style_meta_valeur = ParagraphStyle(
            "PilotageMetaValeur",
            fontName="Helvetica",
            fontSize=8.8,
            leading=9.2,
            textColor=colors.black,
        )
        style_bloc_titre = ParagraphStyle(
            "PilotageBlocTitre",
            fontName="Helvetica-Bold",
            fontSize=9.6,
            leading=10,
            alignment=TA_CENTER,
            textColor=colors.white,
        )
        style_libelle = ParagraphStyle(
            "PilotageLibelle",
            fontName="Helvetica",
            fontSize=8.6,
            leading=9,
            textColor=colors.black,
        )
        style_valeur = ParagraphStyle(
            "PilotageValeur",
            fontName="Helvetica-Bold",
            fontSize=8.8,
            leading=9,
            alignment=TA_RIGHT,
            textColor=colors.black,
        )
        style_carte_label = ParagraphStyle(
            "PilotageCarteLabel",
            fontName="Helvetica-Bold",
            fontSize=7.6,
            leading=8,
            textColor=colors.HexColor("#5A7086"),
        )
        style_carte_valeur = ParagraphStyle(
            "PilotageCarteValeur",
            fontName="Helvetica-Bold",
            fontSize=11.2,
            leading=11.4,
            alignment=TA_RIGHT,
            textColor=colors.HexColor("#173C5A"),
        )
        style_carte_valeur_finale = ParagraphStyle(
            "PilotageCarteValeurFinale",
            fontName="Helvetica-Bold",
            fontSize=12.2,
            leading=12.2,
            alignment=TA_RIGHT,
            textColor=colors.HexColor("#2E7D32"),
        )
        style_bandeau = ParagraphStyle(
            "PilotageBandeau",
            fontName="Helvetica-Bold",
            fontSize=9.3,
            leading=9.8,
            textColor=colors.HexColor("#173C5A"),
        )
        style_avancement_titre = ParagraphStyle(
            "PilotageAvancementTitre",
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=8.2,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#5A7086"),
        )
        style_avancement_valeur = ParagraphStyle(
            "PilotageAvancementValeur",
            fontName="Helvetica-Bold",
            fontSize=10.5,
            leading=10.8,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#163A59"),
        )

        logo_cell = Spacer(1, 1)
        if os.path.exists(logo_path):
            try:
                logo_cell = Image(logo_path, width=27 * mm, height=10.8 * mm)
                logo_cell.hAlign = "LEFT"
            except:
                logo_cell = Spacer(1, 1)

        meta_table = Table([
            [
                Paragraph("CHANTIER", style_meta_label),
                Paragraph(nom_chantier, style_meta_valeur),
                Paragraph("DATE", style_meta_label),
                Paragraph(date_du_jour, style_meta_valeur),
            ]
        ], colWidths=[18 * mm, 47 * mm, 12 * mm, 22 * mm])
        meta_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))

        header_table = Table([
            [logo_cell, Paragraph("PILOTAGE CHANTIER", style_titre), meta_table]
        ], colWidths=[34 * mm, 158 * mm, 74 * mm])
        header_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F4F8FB")),
            ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#D9E4EE")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (1, 0), (1, 0), "CENTER"),
            ("ALIGN", (2, 0), (2, 0), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 3))

        valeurs_graphique = [total_marche, production_cumulee, reste_a_facturer]
        valeurs_plot = [max(total_marche, 0), max(production_cumulee, 0), max(reste_a_facturer, 0)]
        etiquettes_graphique = ["Marché\ntotal", "Production\ncumulée", "Reste à\nfacturer"]
        couleurs_graphique = ["#1F4E79", "#E69138", "#70AD47"]

        fig, ax = plt.subplots(figsize=(3.1, 3.7), facecolor="white")
        barres = ax.bar(
            range(3),
            valeurs_plot,
            width=0.55,
            color=couleurs_graphique,
            edgecolor="#FFFFFF",
            linewidth=0.8,
            zorder=3,
        )

        max_valeur = max([abs(v) for v in valeurs_plot] + [1])
        marge = max_valeur * 0.18
        ax.set_ylim(0, max(max(valeurs_plot), 1) + marge)
        ax.set_xticks(range(3), etiquettes_graphique)
        ax.tick_params(axis="x", labelsize=7.8, colors="#173C5A", length=0, pad=6)
        ax.tick_params(axis="y", labelsize=7.5, colors="#6A7E90")
        ax.grid(axis="y", color="#D9E4EE", linewidth=0.8, alpha=0.8, zorder=0)
        ax.set_axisbelow(True)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.spines["left"].set_color("#D9E4EE")
        ax.spines["bottom"].set_color("#D9E4EE")

        for barre, valeur in zip(barres, valeurs_graphique):
            ax.text(
                barre.get_x() + barre.get_width() / 2,
                barre.get_height() + max(marge * 0.08, 0.6),
                f"{valeur:,.2f} €",
                ha="center",
                va="bottom",
                fontsize=7.8,
                fontweight="bold",
                color="#173C5A",
            )

        with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
            graphique_tmp = tmp.name

        plt.tight_layout()
        plt.savefig(graphique_tmp, dpi=220, bbox_inches="tight", facecolor="white")
        plt.close(fig)

        graphique = Image(graphique_tmp, width=66 * mm, height=66 * mm)
        graphique.hAlign = "CENTER"

        largeur_blocs_gauche = [95 * mm, 30 * mm]
        largeur_blocs_droite = [38 * mm, 24 * mm]

        bloc_1 = Table([
            ["CONSTRUCTION", ""],
            [Paragraph("Montant soumission", style_libelle), Paragraph(f"{montant_soumission:,.2f} €", style_valeur)],
            [Paragraph("+ Avenants cumulés en plus", style_libelle), Paragraph(f"{avenants_plus:,.2f} €", style_valeur)],
            [Paragraph("- Avenants cumulés en moins", style_libelle), Paragraph(f"{avenants_moins:,.2f} €", style_valeur)],
            [Paragraph("= Marché total", style_libelle), Paragraph(f"{total_marche:,.2f} €", style_valeur)],
        ], colWidths=largeur_blocs_gauche)
        bloc_1.hAlign = "CENTER"
        bloc_1.setStyle(TableStyle([
            ("SPAN", (0, 0), (1, 0)),
            ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
            ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (1, 0), 12),
            ("BACKGROUND", (0, 1), (1, -1), colors.white),
            ("ROWBACKGROUNDS", (0, 1), (1, -1), [colors.white, colors.HexColor("#F8FBFF")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B7C5D3")),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
            ("BACKGROUND", (0, 4), (1, 4), colors.HexColor("#FFF2CC")),
            ("FONTNAME", (0, 4), (1, 4), "Helvetica-Bold"),
        ]))

        bloc_2 = Table([
            ["RÉALISÉ", ""],
            [Paragraph("Réalisé du mois", style_libelle), Paragraph(f"{realise_mois:,.2f} €", style_valeur)],
            [Paragraph("+ Révision renseignée", style_libelle), Paragraph(f"{revision:,.2f} €", style_valeur)],
            [Paragraph("= Production du mois", style_libelle), Paragraph(f"{production_mois:,.2f} €", style_valeur)],
            [Paragraph("Réalisé cumulé", style_libelle), Paragraph(f"{realise_cumule:,.2f} €", style_valeur)],
            [Paragraph("= Production cumulée", style_libelle), Paragraph(f"{production_cumulee:,.2f} €", style_valeur)],
        ], colWidths=largeur_blocs_gauche)
        bloc_2.hAlign = "CENTER"
        bloc_2.setStyle(TableStyle([
            ("SPAN", (0, 0), (1, 0)),
            ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#4F81BD")),
            ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
            ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (1, 0), 12),
            ("BACKGROUND", (0, 1), (1, -1), colors.white),
            ("ROWBACKGROUNDS", (0, 1), (1, -1), [colors.white, colors.HexColor("#F7FAFD")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B7C5D3")),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
            ("BACKGROUND", (0, 3), (1, 3), colors.HexColor("#FFF2CC")),
            ("FONTNAME", (0, 3), (1, 3), "Helvetica-Bold"),
            ("BACKGROUND", (0, 5), (1, 5), colors.HexColor("#FFF2CC")),
            ("FONTNAME", (0, 5), (1, 5), "Helvetica-Bold"),
        ]))

        bloc_3 = Table([
            ["RÉSULTAT", ""],
            [Paragraph("Reste à facturer", style_libelle), Paragraph(f"{reste_a_facturer:,.2f} €", style_valeur)],
            [Paragraph("Avancement", style_libelle), Paragraph(f"{avancement:.1f} %", style_valeur)],
            [Paragraph("Écart financier final", style_libelle), Paragraph(f"{reste_a_facturer:,.2f} €", style_valeur)],
        ], colWidths=largeur_blocs_gauche)
        bloc_3.hAlign = "CENTER"
        bloc_3.setStyle(TableStyle([
            ("SPAN", (0, 0), (1, 0)),
            ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#C0504D")),
            ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
            ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (1, 0), 12),
            ("BACKGROUND", (0, 1), (1, -1), colors.white),
            ("ROWBACKGROUNDS", (0, 1), (1, -1), [colors.white, colors.HexColor("#FCF7F7")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#C9B1B1")),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
            ("BACKGROUND", (0, 1), (1, 1), colors.HexColor("#D9EAD3")),
            ("FONTNAME", (0, 1), (1, 1), "Helvetica-Bold"),
            ("BACKGROUND", (0, 3), (1, 3), colors.HexColor("#D9EAD3")),
            ("FONTNAME", (0, 3), (1, 3), "Helvetica-Bold"),
        ]))

        colonne_gauche = Table([
            [bloc_1],
            [bloc_2],
            [bloc_3],
        ], colWidths=[125 * mm])
        colonne_gauche.setStyle(TableStyle([
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))

        bloc_graphique = Table([
            [Paragraph("RÉSUMÉ VISUEL", style_bloc_titre)],
            [graphique],
        ], colWidths=[70 * mm])
        bloc_graphique.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#173C5A")),
            ("BACKGROUND", (0, 1), (0, 1), colors.white),
            ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#B7C5D3")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ]))

        indicateurs_table = Table([
            [Paragraph("Marché total", style_carte_label), Paragraph(f"{total_marche:,.2f} €", style_carte_valeur)],
            [Paragraph("Production du mois", style_carte_label), Paragraph(f"{production_mois:,.2f} €", style_carte_valeur)],
            [Paragraph("Production cumulée", style_carte_label), Paragraph(f"{production_cumulee:,.2f} €", style_carte_valeur)],
            [Paragraph("Avancement %", style_carte_label), Paragraph(f"{avancement:.1f} %", style_carte_valeur)],
            [Paragraph("Reste à facturer", style_carte_label), Paragraph(f"{reste_a_facturer:,.2f} €", style_carte_valeur_finale)],
            [Paragraph("Écart financier final", style_carte_label), Paragraph(f"{reste_a_facturer:,.2f} €", style_carte_valeur_finale)],
        ], colWidths=largeur_blocs_droite)
        indicateurs_table.setStyle(TableStyle([
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#F6F9FC")]),
            ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#D6E0EA")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
            ("BACKGROUND", (0, 4), (1, 4), colors.HexColor("#D9EAD3")),
            ("BACKGROUND", (0, 5), (1, 5), colors.HexColor("#D9EAD3")),
        ]))

        indicateurs_cles = Table([
            [Paragraph("INDICATEURS CLÉS", style_bloc_titre)],
            [indicateurs_table],
        ], colWidths=[70 * mm])
        indicateurs_cles.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#365F91")),
            ("BACKGROUND", (0, 1), (0, 1), colors.white),
            ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#B7C5D3")),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ]))

        colonne_droite = Table([
            [bloc_graphique],
            [indicateurs_cles],
        ], colWidths=[70 * mm])
        colonne_droite.setStyle(TableStyle([
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))

        corps = Table([
            [colonne_gauche, colonne_droite]
        ], colWidths=[128 * mm, 70 * mm])
        corps.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        elements.append(corps)
        elements.append(Spacer(1, 2))

        footer_row_height = 14.5 * mm
        footer_split_height = 5.2 * mm

        avancement_table = Table([
            [Paragraph("AVANCEMENT", style_avancement_titre)],
            [Paragraph(f"{avancement:.1f} %", style_avancement_valeur)],
        ], colWidths=[30 * mm], rowHeights=[footer_split_height, footer_row_height - footer_split_height])
        avancement_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EAF1F8")),
            ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#B7C5D3")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))

        footer_explication = Table([
            [
                Paragraph(
                    f"Reste à facturer = Marché total - Production cumulée<br/>{total_marche:,.2f} € - {production_cumulee:,.2f} € = {reste_a_facturer:,.2f} €",
                    style_bandeau,
                )
            ]
        ], colWidths=[168 * mm], rowHeights=[footer_row_height])
        footer_explication.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#FFF2CC")),
            ("BOX", (0, 0), (0, 0), 0.7, colors.HexColor("#D6B656")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))

        footer_table = Table([
            [footer_explication, avancement_table]
        ], colWidths=[168 * mm, 30 * mm], rowHeights=[footer_row_height])
        footer_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        elements.append(footer_table)

        doc.build(elements)

        if graphique_tmp and os.path.exists(graphique_tmp):
            try:
                os.unlink(graphique_tmp)
            except:
                pass

        subprocess.run(["open", pdf_path])

        
        messagebox.showinfo("Pilotage chantier", texte_popup)
   
   
    def pilotage_delai(self):
        import os
        import subprocess
        import tempfile
        from datetime import datetime
        import matplotlib.pyplot as plt
        from openpyxl import load_workbook
        from tkinter import messagebox
        from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Image, Paragraph
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.enums import TA_RIGHT, TA_CENTER

        def nombre(valeur):
            try:
                if valeur is None:
                    return 0
                if isinstance(valeur, str):
                    valeur = valeur.replace("%", "").replace("€", "").replace(" ", "").replace(",", ".").strip()
                    if valeur in ("", "-", "—"):
                        return 0
                return float(valeur)
            except:
                return 0

        try:
            nom_chantier = self._nom_chantier_selectionne()
            if not nom_chantier:
                return

            dossier_chantier = dossier_chantiers() / nom_chantier
            fichier_delai = os.path.join(dossier_chantier, "Delai.xlsx")

            if not os.path.exists(fichier_delai):
                messagebox.showerror("Erreur", f"Fichier introuvable :\n{fichier_delai}")
                return

            wb = load_workbook(fichier_delai, data_only=True)
            ws = wb.active

            def lire_valeur_delai(libelles):
                if isinstance(libelles, str):
                    libelles = [libelles]
                for row in range(1, ws.max_row + 1):
                    libelle = str(ws[f"M{row}"].value or "").strip().lower()
                    for recherche in libelles:
                        if recherche in libelle:
                            return nombre(ws[f"O{row}"].value)
                return 0

            # Bloc haut : délai corrigé
            delai_soumission_jour = lire_valeur_delai("délai soumission jour")
            jours_avenant_plus = lire_valeur_delai("jour accordes avenant en plus")
            jours_attente_technique = lire_valeur_delai("jour accordes dû à une attente technique")
            jours_imprevu = lire_valeur_delai("jour accordes dû à imprévu, prévisible non visible")
            delai_corrige_jour = lire_valeur_delai("délai corrigé en jour")

            # Bloc bas : délai consommé
            jours_consommes = lire_valeur_delai(["jours actif consommé", "jours consommé"])
            jours_avenant_moins = lire_valeur_delai("jour accordes avenant en moins")
            jours_intemperies = lire_valeur_delai("jour intempéries")
            conges = lire_valeur_delai(["congé/ferrier/compensatoire jour", "congé/ferrier/compensatoire"])
            delai_consomme = lire_valeur_delai(["délai consommé jour", "délai consommé"])

            # Résultat
            jours_restants = lire_valeur_delai("jours restant pour exécution")
            depassement = lire_valeur_delai("ecart en jour en moins")
            ecart_pct = lire_valeur_delai("ecart en %")

            if ecart_pct <= 1 and ecart_pct not in (0,):
                ecart_pct = ecart_pct * 100

            texte = (
                f"Chantier : {nom_chantier}\n\n"

                f"DÉLAI CORRIGÉ\n"
                f"Délai soumission jour : {delai_soumission_jour:.2f} j\n"
                f"Jour accordés avenant en plus : {jours_avenant_plus:.2f} j\n"
                f"Jour accordés dû à une attente Technique : {jours_attente_technique:.2f} j\n"
                f"Jour accordés dû à imprévu, prévisible non visible : {jours_imprevu:.2f} j\n"
                f"Délai corrigé en jour : {delai_corrige_jour:.2f} j\n\n"

                f"DÉLAI CONSOMMÉ\n"
                f"Jours actif consommé : {jours_consommes:.2f} j\n"
                f"Jour accordés avenant en moins : {jours_avenant_moins:.2f} j\n"
                f"Jour intempéries : {jours_intemperies:.2f} j\n"
                f"Congé/ferrier/compensatoire : {conges:.2f} j\n"
                f"Délai consommé : {delai_consomme:.2f} j\n\n"

                f"RÉSULTAT\n"
                f"Jours restant pour exécution : {jours_restants:.2f} j\n"
                f"Ecart en jour en moins : {depassement:.2f} j\n"
                f"Ecart en % : {ecart_pct:.1f} %"
            )

            # Popup
            messagebox.showinfo("📊 Pilotage délai", texte)

            # PDF
            wb.close()
            pdf_path = os.path.join(dossier_chantier, "pilotage_delai.pdf")
            doc = SimpleDocTemplate(
                pdf_path,
                pagesize=landscape(A4),
                leftMargin=10 * mm,
                rightMargin=10 * mm,
                topMargin=8 * mm,
                bottomMargin=8 * mm,
            )

            elements = []
            graphique_tmp = None
            logo_path = str(dossier_base() / "logo_jt_bati.png")
            date_du_jour = datetime.now().strftime("%d/%m/%Y")

            style_titre = ParagraphStyle(
                "DelaiTitre",
                fontName="Helvetica-Bold",
                fontSize=18,
                leading=20,
                alignment=TA_CENTER,
                textColor=colors.HexColor("#163A59"),
            )
            style_meta_label = ParagraphStyle(
                "DelaiMetaLabel",
                fontName="Helvetica-Bold",
                fontSize=8.5,
                leading=10,
                textColor=colors.HexColor("#36556F"),
            )
            style_meta_valeur = ParagraphStyle(
                "DelaiMetaValeur",
                fontName="Helvetica",
                fontSize=8.8,
                leading=10.5,
                textColor=colors.black,
            )
            style_bloc_titre = ParagraphStyle(
                "DelaiBlocTitre",
                fontName="Helvetica-Bold",
                fontSize=9.4,
                leading=10.5,
                alignment=TA_CENTER,
                textColor=colors.white,
            )
            style_libelle = ParagraphStyle(
                "DelaiLibelle",
                fontName="Helvetica",
                fontSize=7.8,
                leading=8.8,
                textColor=colors.black,
            )
            style_valeur = ParagraphStyle(
                "DelaiValeur",
                fontName="Helvetica-Bold",
                fontSize=8.1,
                leading=9.2,
                alignment=TA_RIGHT,
                textColor=colors.black,
            )
            style_carte_label = ParagraphStyle(
                "DelaiCarteLabel",
                fontName="Helvetica-Bold",
                fontSize=7.4,
                leading=8.4,
                textColor=colors.HexColor("#5A7086"),
            )
            style_carte_valeur = ParagraphStyle(
                "DelaiCarteValeur",
                fontName="Helvetica-Bold",
                fontSize=10.8,
                leading=12,
                alignment=TA_RIGHT,
                textColor=colors.HexColor("#173C5A"),
            )
            style_carte_valeur_finale = ParagraphStyle(
                "DelaiCarteValeurFinale",
                fontName="Helvetica-Bold",
                fontSize=11.6,
                leading=12.8,
                alignment=TA_RIGHT,
                textColor=colors.HexColor("#9C3B32"),
            )
            style_bandeau = ParagraphStyle(
                "DelaiBandeau",
                fontName="Helvetica-Bold",
                fontSize=9.1,
                leading=10.5,
                textColor=colors.HexColor("#173C5A"),
            )
            style_ecart_titre = ParagraphStyle(
                "DelaiEcartTitre",
                fontName="Helvetica-Bold",
                fontSize=8,
                leading=9,
                alignment=TA_CENTER,
                textColor=colors.HexColor("#5A7086"),
            )
            style_ecart_valeur = ParagraphStyle(
                "DelaiEcartValeur",
                fontName="Helvetica-Bold",
                fontSize=10.4,
                leading=12,
                alignment=TA_CENTER,
                textColor=colors.HexColor("#163A59"),
            )

            logo_cell = Spacer(1, 1)
            if os.path.exists(logo_path):
                try:
                    logo_cell = Image(logo_path, width=30 * mm, height=12 * mm)
                    logo_cell.hAlign = "LEFT"
                except:
                    logo_cell = Spacer(1, 1)

            meta_table = Table([
                [
                    Paragraph("CHANTIER", style_meta_label),
                    Paragraph(nom_chantier, style_meta_valeur),
                    Paragraph("DATE", style_meta_label),
                    Paragraph(date_du_jour, style_meta_valeur),
                ]
            ], colWidths=[18 * mm, 47 * mm, 12 * mm, 22 * mm])
            meta_table.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]))

            header_table = Table([
                [logo_cell, Paragraph("PILOTAGE DÉLAI", style_titre), meta_table]
            ], colWidths=[34 * mm, 158 * mm, 74 * mm])
            header_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F4F8FB")),
                ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#D9E4EE")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (1, 0), (1, 0), "CENTER"),
                ("ALIGN", (2, 0), (2, 0), "RIGHT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]))
            elements.append(header_table)
            elements.append(Spacer(1, 4))

            valeurs_graphique = [delai_corrige_jour, delai_consomme, jours_restants]
            valeurs_plot = [max(delai_corrige_jour, 0), max(delai_consomme, 0), max(jours_restants, 0)]
            etiquettes_graphique = ["Délai\ncorrigé", "Délai\nconsommé", "Jours\nrestants"]
            couleurs_graphique = ["#1F4E79", "#E69138", "#70AD47"]

            fig, ax = plt.subplots(figsize=(3.2, 4.0), facecolor="white")
            barres = ax.bar(
                range(3),
                valeurs_plot,
                width=0.55,
                color=couleurs_graphique,
                edgecolor="#FFFFFF",
                linewidth=0.8,
                zorder=3,
            )

            max_valeur = max([abs(v) for v in valeurs_plot] + [1])
            marge = max_valeur * 0.18
            ax.set_ylim(0, max(max(valeurs_plot), 1) + marge)
            ax.set_xticks(range(3), etiquettes_graphique)
            ax.tick_params(axis="x", labelsize=7.6, colors="#173C5A", length=0, pad=6)
            ax.tick_params(axis="y", labelsize=7.2, colors="#6A7E90")
            ax.grid(axis="y", color="#D9E4EE", linewidth=0.8, alpha=0.8, zorder=0)
            ax.set_axisbelow(True)
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            ax.spines["left"].set_color("#D9E4EE")
            ax.spines["bottom"].set_color("#D9E4EE")

            for barre, valeur in zip(barres, valeurs_graphique):
                ax.text(
                    barre.get_x() + barre.get_width() / 2,
                    barre.get_height() + max(marge * 0.08, 0.3),
                    f"{valeur:.2f} j",
                    ha="center",
                    va="bottom",
                    fontsize=7.6,
                    fontweight="bold",
                    color="#173C5A",
                )

            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
                graphique_tmp = tmp.name

            plt.tight_layout()
            plt.savefig(graphique_tmp, dpi=220, bbox_inches="tight", facecolor="white")
            plt.close(fig)

            graphique = Image(graphique_tmp, width=72 * mm, height=72 * mm)
            graphique.hAlign = "CENTER"

            largeur_blocs_gauche = [96 * mm, 29 * mm]
            largeur_blocs_droite = [38 * mm, 24 * mm]

            bloc_1 = Table([
                ["CONSTRUCTION", ""],
                [Paragraph("Délai soumission jour", style_libelle), Paragraph(f"{delai_soumission_jour:.2f} j", style_valeur)],
                [Paragraph("+ Jour accordés avenant en plus", style_libelle), Paragraph(f"{jours_avenant_plus:.2f} j", style_valeur)],
                [Paragraph("+ Jour accordés dû à une attente Technique", style_libelle), Paragraph(f"{jours_attente_technique:.2f} j", style_valeur)],
                [Paragraph("+ Jour accordés dû à imprévu, prévisible non visible", style_libelle), Paragraph(f"{jours_imprevu:.2f} j", style_valeur)],
                [Paragraph("= Délai corrigé en jour", style_libelle), Paragraph(f"{delai_corrige_jour:.2f} j", style_valeur)],
            ], colWidths=largeur_blocs_gauche)
            bloc_1.hAlign = "CENTER"
            bloc_1.setStyle(TableStyle([
                ("SPAN", (0, 0), (1, 0)),
                ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#1F4E79")),
                ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
                ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (1, 0), 11),
                ("BACKGROUND", (0, 1), (1, -1), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (1, -1), [colors.white, colors.HexColor("#F8FBFF")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B7C5D3")),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 2.5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
                ("BACKGROUND", (0, 5), (1, 5), colors.HexColor("#FFF2CC")),
                ("FONTNAME", (0, 5), (1, 5), "Helvetica-Bold"),
            ]))

            bloc_2 = Table([
                ["CONSOMMÉ", ""],
                [Paragraph("Jours actif consommé", style_libelle), Paragraph(f"{jours_consommes:.2f} j", style_valeur)],
                [Paragraph("- Jour accordés avenant en moins", style_libelle), Paragraph(f"{jours_avenant_moins:.2f} j", style_valeur)],
                [Paragraph("- Jour intempéries", style_libelle), Paragraph(f"{jours_intemperies:.2f} j", style_valeur)],
                [Paragraph("- Congé/ferrier/compensatoire", style_libelle), Paragraph(f"{conges:.2f} j", style_valeur)],
                [Paragraph("= Délai consommé", style_libelle), Paragraph(f"{delai_consomme:.2f} j", style_valeur)],
            ], colWidths=largeur_blocs_gauche)
            bloc_2.hAlign = "CENTER"
            bloc_2.setStyle(TableStyle([
                ("SPAN", (0, 0), (1, 0)),
                ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#4F81BD")),
                ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
                ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (1, 0), 11),
                ("BACKGROUND", (0, 1), (1, -1), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (1, -1), [colors.white, colors.HexColor("#F7FAFD")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B7C5D3")),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 2.5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
                ("BACKGROUND", (0, 5), (1, 5), colors.HexColor("#FFF2CC")),
                ("FONTNAME", (0, 5), (1, 5), "Helvetica-Bold"),
            ]))

            bloc_3 = Table([
                ["RÉSULTAT", ""],
                [Paragraph("Jours restant pour exécution", style_libelle), Paragraph(f"{jours_restants:.2f} j", style_valeur)],
                [Paragraph("Ecart en jour en moins", style_libelle), Paragraph(f"{depassement:.2f} j", style_valeur)],
                [Paragraph("Ecart en %", style_libelle), Paragraph(f"{ecart_pct:.1f} %", style_valeur)],
            ], colWidths=largeur_blocs_gauche)
            bloc_3.hAlign = "CENTER"
            bloc_3.setStyle(TableStyle([
                ("SPAN", (0, 0), (1, 0)),
                ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#C0504D")),
                ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
                ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (1, 0), 11),
                ("BACKGROUND", (0, 1), (1, -1), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (1, -1), [colors.white, colors.HexColor("#FCF7F7")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#C9B1B1")),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 2.5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
                ("BACKGROUND", (0, 1), (1, 1), colors.HexColor("#E7E7E7")),
                ("BACKGROUND", (0, 3), (1, 3), colors.HexColor("#D9EAD3")),
                ("FONTNAME", (0, 3), (1, 3), "Helvetica-Bold"),
            ]))

            colonne_gauche = Table([
                [bloc_1],
                [bloc_2],
                [bloc_3],
            ], colWidths=[125 * mm])
            colonne_gauche.setStyle(TableStyle([
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))

            bloc_graphique = Table([
                [Paragraph("RÉSUMÉ VISUEL", style_bloc_titre)],
                [graphique],
            ], colWidths=[70 * mm])
            bloc_graphique.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#173C5A")),
                ("BACKGROUND", (0, 1), (0, 1), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#B7C5D3")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]))

            indicateurs_table = Table([
                [Paragraph("Délai corrigé", style_carte_label), Paragraph(f"{delai_corrige_jour:.2f} j", style_carte_valeur)],
                [Paragraph("Jours actif consommé", style_carte_label), Paragraph(f"{jours_consommes:.2f} j", style_carte_valeur)],
                [Paragraph("Délai consommé", style_carte_label), Paragraph(f"{delai_consomme:.2f} j", style_carte_valeur)],
                [Paragraph("Jours restants", style_carte_label), Paragraph(f"{jours_restants:.2f} j", style_carte_valeur)],
                [Paragraph("Ecart en %", style_carte_label), Paragraph(f"{ecart_pct:.1f} %", style_carte_valeur_finale)],
            ], colWidths=largeur_blocs_droite)
            indicateurs_table.setStyle(TableStyle([
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#F6F9FC")]),
                ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#D6E0EA")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("BACKGROUND", (0, 4), (1, 4), colors.HexColor("#D9EAD3")),
            ]))

            indicateurs_cles = Table([
                [Paragraph("INDICATEURS CLÉS", style_bloc_titre)],
                [indicateurs_table],
            ], colWidths=[70 * mm])
            indicateurs_cles.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#365F91")),
                ("BACKGROUND", (0, 1), (0, 1), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#B7C5D3")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]))

            colonne_droite = Table([
                [bloc_graphique],
                [indicateurs_cles],
            ], colWidths=[70 * mm])
            colonne_droite.setStyle(TableStyle([
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))

            corps = Table([
                [colonne_gauche, colonne_droite]
            ], colWidths=[128 * mm, 70 * mm])
            corps.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]))
            elements.append(corps)
            elements.append(Spacer(1, 4))

            footer_row_height = 18 * mm
            footer_split_height = 6.5 * mm

            ecart_table = Table([
                [Paragraph("ECART %", style_ecart_titre)],
                [Paragraph(f"{ecart_pct:.1f} %", style_ecart_valeur)],
            ], colWidths=[30 * mm], rowHeights=[footer_split_height, footer_row_height - footer_split_height])
            ecart_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EAF1F8")),
                ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#B7C5D3")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))

            footer_explication = Table([
                [
                    Paragraph(
                        f"Jours restant pour exécution = Délai corrigé - Délai consommé<br/>{delai_corrige_jour:.2f} j - {delai_consomme:.2f} j = {jours_restants:.2f} j",
                        style_bandeau,
                    )
                ]
            ], colWidths=[168 * mm], rowHeights=[footer_row_height])
            footer_explication.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#FFF2CC")),
                ("BOX", (0, 0), (0, 0), 0.7, colors.HexColor("#D6B656")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))

            footer_table = Table([
                [footer_explication, ecart_table]
            ], colWidths=[168 * mm, 30 * mm], rowHeights=[footer_row_height])
            footer_table.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]))
            elements.append(footer_table)

            doc.build(elements)

            if graphique_tmp and os.path.exists(graphique_tmp):
                try:
                    os.unlink(graphique_tmp)
                except:
                    pass

            subprocess.run(["open", pdf_path])

        except Exception as e:
            if 'graphique_tmp' in locals() and graphique_tmp and os.path.exists(graphique_tmp):
                try:
                    os.unlink(graphique_tmp)
                except:
                    pass
            messagebox.showerror("Erreur pilotage délai", str(e))

    def rendements_chantier(self):
        import os
        import subprocess
        import tempfile
        from datetime import datetime
        import matplotlib.pyplot as plt
        from openpyxl import load_workbook
        from tkinter import messagebox
        from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Image, Paragraph
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_RIGHT, TA_CENTER
        from reportlab.lib.pagesizes import A4, landscape

        def nombre(valeur):
            try:
                if valeur is None:
                    return 0
                if isinstance(valeur, str):
                    valeur = valeur.replace("%", "").replace("€", "").replace(" ", "").replace(",", ".").strip()
                    if valeur in ("", "-", "—"):
                        return 0
                return float(valeur)
            except:
                return 0

        try:
            nom_chantier = self._nom_chantier_selectionne()
            if not nom_chantier:
                return

            dossier_chantier = dossier_chantiers() / nom_chantier
            fichier_delai = os.path.join(dossier_chantier, "Rendement.xlsx")

            if not os.path.exists(fichier_delai):
                messagebox.showerror("Erreur", f"Fichier introuvable :\n{fichier_delai}")
                return

            wb = load_workbook(fichier_delai, data_only=True)
            ws = wb.active

            heures_soumission = nombre(ws["O65"].value)
            heures_avenant_plus = nombre(ws["O69"].value)
            heures_attente_technique = nombre(ws["O73"].value)
            heures_imprevu = nombre(ws["O75"].value)
            heures_consommees = nombre(ws["O67"].value)
            heures_avenant_moins = nombre(ws["O71"].value)
            heures_corrigees = heures_soumission + heures_avenant_plus + heures_attente_technique + heures_imprevu - heures_avenant_moins
            heures_consommees_finales = heures_consommees + heures_avenant_moins
            heures_delai_consomme = heures_consommees_finales
            heures_restantes = heures_corrigees - heures_delai_consomme
            ecart_heures = heures_consommees_finales - heures_corrigees
            ecart_pct = nombre(ws["O82"].value)

            if ecart_pct <= 1 and ws["O82"].value not in (None, "", 0):
                ecart_pct = ecart_pct * 100

            valorisation_finale = ecart_heures * 55
            wb.close()

            texte = (
                f"Chantier : {nom_chantier}\n\n"

                f"CONSTRUCTION DU DÉLAI\n"
                f"Heures soumission : {heures_soumission:.2f} h\n"
                f"Heures accordées avenant en plus : {heures_avenant_plus:.2f} h\n"
                f"Heures accordées dû à une attente Technique : {heures_attente_technique:.2f} h\n"
                f"Heures accordées dû à imprévu, prévisible non visible : {heures_imprevu:.2f} h\n"
                f"Délai corrigé en heures : {heures_corrigees:.2f} h\n\n"

                f"CONSOMMÉ\n"
                f"Heures consommées : {heures_consommees:.2f} h\n"
                f"Heures accordées avenant en moins : {heures_avenant_moins:.2f} h\n"
                f"Heures consommées finales / délai consommé en heures : {heures_delai_consomme:.2f} h\n\n"

                f"RÉSULTAT\n"
                f"Heures restantes pour exécution : {heures_restantes:.2f} h\n"
                f"Écart en heures : {ecart_heures:.2f} h\n"
                f"Pourcentage : {ecart_pct:.1f} %\n"
                f"Valorisation financière finale : {valorisation_finale:,.2f} €"
            )

            messagebox.showinfo("⏱ Rendement chantier", texte)

            pdf_path = os.path.join(dossier_chantier, "rendement_chantier.pdf")
            doc = SimpleDocTemplate(
                pdf_path,
                pagesize=landscape(A4),
                leftMargin=10 * mm,
                rightMargin=10 * mm,
                topMargin=8 * mm,
                bottomMargin=8 * mm,
            )

            elements = []
            graphique_tmp = None

            logo_path = str(dossier_base() / "logo_jt_bati.png")
            date_du_jour = datetime.now().strftime("%d/%m/%Y")
            style_titre = ParagraphStyle(
                "RendementTitre",
                fontName="Helvetica-Bold",
                fontSize=18,
                leading=20,
                alignment=TA_CENTER,
                textColor=colors.HexColor("#163A59"),
            )
            style_meta_label = ParagraphStyle(
                "RendementMetaLabel",
                fontName="Helvetica-Bold",
                fontSize=8.5,
                leading=10,
                textColor=colors.HexColor("#36556F"),
            )
            style_meta_valeur = ParagraphStyle(
                "RendementMetaValeur",
                fontName="Helvetica",
                fontSize=8.8,
                leading=10.5,
                textColor=colors.black,
            )
            style_bloc_titre = ParagraphStyle(
                "RendementBlocTitre",
                fontName="Helvetica-Bold",
                fontSize=9.6,
                leading=11,
                alignment=TA_CENTER,
                textColor=colors.white,
            )
            style_libelle = ParagraphStyle(
                "RendementLibelle",
                fontName="Helvetica",
                fontSize=8.6,
                leading=9.8,
                textColor=colors.black,
            )
            style_valeur = ParagraphStyle(
                "RendementValeur",
                fontName="Helvetica-Bold",
                fontSize=8.8,
                leading=10,
                alignment=TA_RIGHT,
                textColor=colors.black,
            )
            style_carte_label = ParagraphStyle(
                "RendementCarteLabel",
                fontName="Helvetica-Bold",
                fontSize=7.6,
                leading=8.6,
                textColor=colors.HexColor("#5A7086"),
            )
            style_carte_valeur = ParagraphStyle(
                "RendementCarteValeur",
                fontName="Helvetica-Bold",
                fontSize=11.2,
                leading=12.5,
                alignment=TA_RIGHT,
                textColor=colors.HexColor("#173C5A"),
            )
            style_carte_valeur_finale = ParagraphStyle(
                "RendementCarteValeurFinale",
                fontName="Helvetica-Bold",
                fontSize=12.2,
                leading=13.5,
                alignment=TA_RIGHT,
                textColor=colors.HexColor("#9C3B32"),
            )
            style_bandeau = ParagraphStyle(
                "RendementBandeau",
                fontName="Helvetica-Bold",
                fontSize=9.3,
                leading=11,
                textColor=colors.HexColor("#173C5A"),
            )
            style_bandeau_valeur = ParagraphStyle(
                "RendementBandeauValeur",
                fontName="Helvetica-Bold",
                fontSize=10.3,
                leading=11.5,
                alignment=TA_RIGHT,
                textColor=colors.HexColor("#9C3B32"),
            )
            style_taux_titre = ParagraphStyle(
                "RendementTauxTitre",
                fontName="Helvetica-Bold",
                fontSize=8,
                leading=9,
                alignment=TA_CENTER,
                textColor=colors.HexColor("#5A7086"),
            )
            style_taux_valeur = ParagraphStyle(
                "RendementTauxValeur",
                fontName="Helvetica-Bold",
                fontSize=10.5,
                leading=12,
                alignment=TA_CENTER,
                textColor=colors.HexColor("#163A59"),
            )

            logo_cell = Spacer(1, 1)
            if os.path.exists(logo_path):
                try:
                    logo_cell = Image(logo_path, width=30 * mm, height=12 * mm)
                    logo_cell.hAlign = "LEFT"
                except:
                    logo_cell = Spacer(1, 1)

            meta_table = Table([
                [
                    Paragraph("CHANTIER", style_meta_label),
                    Paragraph(nom_chantier, style_meta_valeur),
                    Paragraph("DATE", style_meta_label),
                    Paragraph(date_du_jour, style_meta_valeur),
                ]
            ], colWidths=[18 * mm, 47 * mm, 12 * mm, 22 * mm])
            meta_table.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]))

            header_table = Table([
                [logo_cell, Paragraph("RENDEMENT CHANTIER", style_titre), meta_table]
            ], colWidths=[34 * mm, 158 * mm, 74 * mm])
            header_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F4F8FB")),
                ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#D9E4EE")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (1, 0), (1, 0), "CENTER"),
                ("ALIGN", (2, 0), (2, 0), "RIGHT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]))
            elements.append(header_table)
            elements.append(Spacer(1, 5))

            valeurs_graphique = [heures_corrigees, heures_delai_consomme, heures_restantes]
            etiquettes_graphique = ["Délai\ncorrigé", "Consommé\nfinal", "Heures\nrestantes"]
            couleurs_graphique = ["#1F4E79", "#E69138", "#70AD47"]

            fig, ax = plt.subplots(figsize=(3.25, 4.1), facecolor="white")
            barres = ax.bar(
                range(3),
                valeurs_graphique,
                width=0.55,
                color=couleurs_graphique,
                edgecolor="#FFFFFF",
                linewidth=0.8,
                zorder=3,
            )

            max_valeur = max([abs(v) for v in valeurs_graphique] + [1])
            marge = max_valeur * 0.18
            ax.set_ylim(0, max(max(valeurs_graphique), 1) + marge)
            ax.set_xticks(range(3), etiquettes_graphique)
            ax.tick_params(axis="x", labelsize=7.8, colors="#173C5A", length=0, pad=6)
            ax.tick_params(axis="y", labelsize=7.5, colors="#6A7E90")
            ax.grid(axis="y", color="#D9E4EE", linewidth=0.8, alpha=0.8, zorder=0)
            ax.set_axisbelow(True)
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            ax.spines["left"].set_color("#D9E4EE")
            ax.spines["bottom"].set_color("#D9E4EE")

            for barre, valeur in zip(barres, valeurs_graphique):
                ax.text(
                    barre.get_x() + barre.get_width() / 2,
                    barre.get_height() + max(marge * 0.08, 0.6),
                    f"{valeur:.2f} h",
                    ha="center",
                    va="bottom",
                    fontsize=7.8,
                    fontweight="bold",
                    color="#173C5A",
                )

            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
                graphique_tmp = tmp.name

            plt.tight_layout()
            plt.savefig(graphique_tmp, dpi=220, bbox_inches="tight", facecolor="white")
            plt.close(fig)

            graphique = Image(graphique_tmp, width=74 * mm, height=74 * mm)
            graphique.hAlign = "CENTER"

            largeur_blocs_gauche = [95 * mm, 30 * mm]
            largeur_blocs_droite = [38 * mm, 24 * mm]

            bloc_1 = Table([
                ["CONSTRUCTION DU DÉLAI", ""],
                [Paragraph("Heures soumission", style_libelle), Paragraph(f"{heures_soumission:.2f} h", style_valeur)],
                [Paragraph("+ Heures accordées avenant en plus", style_libelle), Paragraph(f"{heures_avenant_plus:.2f} h", style_valeur)],
                [Paragraph("+ Heures accordées dû à une attente Technique", style_libelle), Paragraph(f"{heures_attente_technique:.2f} h", style_valeur)],
                [Paragraph("+ Heures accordées dû à imprévu, prévisible non visible", style_libelle), Paragraph(f"{heures_imprevu:.2f} h", style_valeur)],
                [Paragraph("= Délai corrigé en heures", style_libelle), Paragraph(f"{heures_corrigees:.2f} h", style_valeur)],
            ], colWidths=largeur_blocs_gauche)
            bloc_1.hAlign = "CENTER"
            bloc_1.setStyle(TableStyle([
                ("SPAN", (0, 0), (1, 0)),
                ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#1F4E79")),
                ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
                ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (1, 0), 12),
                ("BACKGROUND", (0, 1), (1, -1), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (1, -1), [colors.white, colors.HexColor("#F8FBFF")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B7C5D3")),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("BACKGROUND", (0, 5), (1, 5), colors.HexColor("#FFF2CC")),
                ("FONTNAME", (0, 5), (1, 5), "Helvetica-Bold"),
            ]))

            bloc_2 = Table([
                ["CONSOMMÉ", ""],
                [Paragraph("- Heures consommées", style_libelle), Paragraph(f"{heures_consommees:.2f} h", style_valeur)],
                [Paragraph("- Heures accordées avenant en moins", style_libelle), Paragraph(f"{heures_avenant_moins:.2f} h", style_valeur)],
                [Paragraph("= Heures consommées finales / délai consommé en heures", style_libelle), Paragraph(f"{heures_delai_consomme:.2f} h", style_valeur)],
            ], colWidths=largeur_blocs_gauche)
            bloc_2.hAlign = "CENTER"
            bloc_2.setStyle(TableStyle([
                ("SPAN", (0, 0), (1, 0)),
                ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#4F81BD")),
                ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
                ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (1, 0), 12),
                ("BACKGROUND", (0, 1), (1, -1), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (1, -1), [colors.white, colors.HexColor("#F7FAFD")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B7C5D3")),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("BACKGROUND", (0, 3), (1, 3), colors.HexColor("#FFF2CC")),
                ("FONTNAME", (0, 3), (1, 3), "Helvetica-Bold"),
            ]))

            bloc_3 = Table([
                ["RÉSULTAT", ""],
                [Paragraph("Heures restantes pour exécution", style_libelle), Paragraph(f"{heures_restantes:.2f} h", style_valeur)],
                [Paragraph("Écart en heures", style_libelle), Paragraph(f"{ecart_heures:.2f} h", style_valeur)],
                [Paragraph("Pourcentage d’avancement", style_libelle), Paragraph(f"{ecart_pct:.1f} %", style_valeur)],
                [Paragraph("Valorisation financière finale (Écart en heures × 55 €)", style_libelle), Paragraph(f"{valorisation_finale:,.2f} €", style_valeur)],
            ], colWidths=largeur_blocs_gauche)
            bloc_3.hAlign = "CENTER"
            bloc_3.setStyle(TableStyle([
                ("SPAN", (0, 0), (1, 0)),
                ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#C0504D")),
                ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
                ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (1, 0), 12),
                ("BACKGROUND", (0, 1), (1, -1), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (1, -1), [colors.white, colors.HexColor("#FCF7F7")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#C9B1B1")),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("BACKGROUND", (0, 1), (1, 1), colors.HexColor("#E7E7E7")),
                ("BACKGROUND", (0, 4), (1, 4), colors.HexColor("#D9EAD3")),
                ("FONTNAME", (0, 4), (1, 4), "Helvetica-Bold"),
            ]))

            colonne_gauche = Table([
                [bloc_1],
                [bloc_2],
                [bloc_3],
            ], colWidths=[125 * mm])
            colonne_gauche.setStyle(TableStyle([
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))

            bloc_graphique = Table([
                [Paragraph("RÉSUMÉ VISUEL", style_bloc_titre)],
                [graphique],
            ], colWidths=[70 * mm])
            bloc_graphique.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#173C5A")),
                ("BACKGROUND", (0, 1), (0, 1), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#B7C5D3")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]))

            indicateurs_table = Table([
                [Paragraph("Délai corrigé en heures", style_carte_label), Paragraph(f"{heures_corrigees:.2f} h", style_carte_valeur)],
                [Paragraph("Heures consommées finales", style_carte_label), Paragraph(f"{heures_delai_consomme:.2f} h", style_carte_valeur)],
                [Paragraph("Heures restantes", style_carte_label), Paragraph(f"{heures_restantes:.2f} h", style_carte_valeur)],
                [Paragraph("Écart en heures", style_carte_label), Paragraph(f"{ecart_heures:.2f} h", style_carte_valeur)],
                [Paragraph("Pourcentage d’avancement", style_carte_label), Paragraph(f"{ecart_pct:.1f} %", style_carte_valeur)],
                [Paragraph("Valorisation financière finale", style_carte_label), Paragraph(f"{valorisation_finale:,.2f} €", style_carte_valeur_finale)],
            ], colWidths=largeur_blocs_droite)
            indicateurs_table.setStyle(TableStyle([
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#F6F9FC")]),
                ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#D6E0EA")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("BACKGROUND", (0, 5), (1, 5), colors.HexColor("#D9EAD3")),
            ]))

            indicateurs_cles = Table([
                [Paragraph("INDICATEURS CLÉS", style_bloc_titre)],
                [indicateurs_table],
            ], colWidths=[70 * mm])
            indicateurs_cles.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#365F91")),
                ("BACKGROUND", (0, 1), (0, 1), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#B7C5D3")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]))

            colonne_droite = Table([
                [bloc_graphique],
                [indicateurs_cles],
            ], colWidths=[70 * mm])
            colonne_droite.setStyle(TableStyle([
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))

            corps = Table([
                [colonne_gauche, colonne_droite]
            ], colWidths=[128 * mm, 70 * mm])
            corps.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]))
            elements.append(corps)
            elements.append(Spacer(1, 4))

            footer_row_height = 18 * mm
            footer_split_height = 6.5 * mm

            taux_table = Table([
                [Paragraph("TAUX HORAIRE", style_taux_titre)],
                [Paragraph("55.00 €/h", style_taux_valeur)],
            ], colWidths=[30 * mm], rowHeights=[footer_split_height, footer_row_height - footer_split_height])
            taux_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EAF1F8")),
                ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#B7C5D3")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))

            footer_explication = Table([
                [
                    Paragraph(
                        f"Valorisation financière finale = Écart en heures × 55 €/h<br/>{ecart_heures:.2f} h × 55 €/h = {valorisation_finale:,.2f} €",
                        style_bandeau,
                    )
                ]
            ], colWidths=[168 * mm], rowHeights=[footer_row_height])
            footer_explication.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#FFF2CC")),
                ("BOX", (0, 0), (0, 0), 0.7, colors.HexColor("#D6B656")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))

            footer_table = Table([
                [footer_explication, taux_table]
            ], colWidths=[168 * mm, 30 * mm], rowHeights=[footer_row_height])
            footer_table.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]))
            elements.append(footer_table)

            doc.build(elements)

            if graphique_tmp and os.path.exists(graphique_tmp):
                try:
                    os.unlink(graphique_tmp)
                except:
                    pass

            subprocess.run(["open", pdf_path])

        except Exception as e:
            if 'graphique_tmp' in locals() and graphique_tmp and os.path.exists(graphique_tmp):
                try:
                    os.unlink(graphique_tmp)
                except:
                    pass
            messagebox.showerror("Erreur rendement chantier", str(e))

    def _chemin_chantier_selectionne(self) -> Path | None:
        sel = self.tree.selection()
        if not sel:
            return None
        return Path(sel[0])

    def refresh_liste(self) -> None:
        self.lbl_path.config(text=f"Emplacement: {dossier_chantiers()}")
        for item in self.tree.get_children():
            self.tree.delete(item)

        d = dossier_chantiers()
        fichiers = sorted(d.glob("*.json"))
        fichiers_illisibles = []

        for p in fichiers:
            if p.stat().st_size == 0:
                continue
            try:
                chantier = lire_json(p)
                info = infos_chantier(chantier)
                av = info["avancement"]
                try:
                    av_txt = f"{float(av):.0f}%"
                except Exception:
                    av_txt = str(av)

                self.tree.insert(
                    "",
                    "end",
                    iid=str(p),
                    values=(info["nom"], info["client"], info["etat"], av_txt, info["debut"], info["fin"]),
                )
            except Exception as e:
                fichiers_illisibles.append(f"{p.name} ({e})")
                continue

        if fichiers_illisibles:
            messagebox.showwarning(
                "Chantier",
                "Fichier chantier illisible :\n" + "\n".join(fichiers_illisibles)
            )

    def _pr_path_and_hint(self) -> tuple[Path, str]:
        p = self._chemin_chantier_selectionne()
        if not p:
            raise ValueError("Sélectionne un chantier dans la liste.")
        dossier_chantier = self._dossier_depuis_json(p)
        pr_path = dossier_chantier / "data" / "prix_de_revient.xlsx"
        if not pr_path.exists():
            raise FileNotFoundError(f"PR introuvable : {pr_path}")
        hint = pr_path.name
        return pr_path, hint

    def prix_revient(self) -> None:
        p = self._chemin_chantier_selectionne()
        if not p:
            messagebox.showwarning("Prix de revient", "Sélectionne un chantier dans la liste.")
            return
        dossier_chantier = self._dossier_depuis_json(p)
        try:
            pr_path = dossier_chantier / "data" / "prix_de_revient.xlsx"
            if not self._preparer_ouverture_document(pr_path):
                return
            open_pr(dossier_chantier)
            self._planifier_rappel_pdf_historique_ouverture()
        except Exception as e:
            messagebox.showerror("Prix de revient", str(e))

    # ======================================================
    # ✅ OBJECTIF UNIQUE : AUTOMATISATION TEXTE -> COLONNE P
    # Matière  : P3:P12
    # MainOeuvre: P17:P26
    # ======================================================
    def recherche_matiere(self) -> None:
        try:
            pr_path, hint = self._pr_path_and_hint()
            if not self._preparer_ouverture_document(pr_path):
                return
            _prepare_excel_file_before_write(pr_path)
            subprocess.run(["open", str(pr_path)], check=False)
            self._planifier_rappel_pdf_historique_ouverture()
            time.sleep(0.3)

            items = read_biblio_colA_openpyxl(pr_path, "Bibliothèque_Matière")

            def on_pick(txt: str):
                target = excel_find_first_empty_cell(hint, "Chiffrage", "P3:P12")
                if not target:
                    messagebox.showwarning("Matière", "Zone P3:P12 pleine ou Excel pas prêt.")
                    return
                excel_set_cell_value(hint, "Chiffrage", target, txt)

            popup_search_20(self, "Matière → Chiffrage (P3:P12)", items, on_pick)

        except Exception as e:
            messagebox.showerror("Matière", str(e))

    def recherche_mo(self) -> None:
        try:
            pr_path, hint = self._pr_path_and_hint()
            if not self._preparer_ouverture_document(pr_path):
                return
            _prepare_excel_file_before_write(pr_path)
            subprocess.run(["open", str(pr_path)], check=False)
            self._planifier_rappel_pdf_historique_ouverture()
            time.sleep(0.3)

            items = read_biblio_colA_openpyxl(pr_path, "Bibliothèque_MainOeuvre")

            def on_pick(txt: str):
                target = excel_find_first_empty_cell(hint, "Chiffrage", "P17:P26")
                if not target:
                    messagebox.showwarning("Main-d'œuvre", "Zone P17:P26 pleine ou Excel pas prêt.")
                    return
                excel_set_cell_value(hint, "Chiffrage", target, txt)

            popup_search_20(self, "Main-d’œuvre → Chiffrage (P17:P26)", items, on_pick)

        except Exception as e:
            messagebox.showerror("Main-d'œuvre", str(e))

    # ---------------------------
    # Boutons non prioritaires (inchangés)
    # ---------------------------
    
    def calcul_reca(self) -> None:
        import os
        from datetime import datetime
        from tkinter import messagebox

        p = self._chemin_chantier_selectionne()
        if not p:
            messagebox.showwarning("Calcul / Reca", "Sélectionne un chantier.")
            return

        dossier = self._dossier_chantier_selectionne()
        chemin_pr = dossier / "data" / "prix_de_revient.xlsx"

        if not chemin_pr.exists():
            messagebox.showwarning(
                "Date du PR",
                "Le fichier prix_de_revient.xlsx est introuvable."
            )
            return

        timestamp = os.path.getmtime(chemin_pr)
        date_modif = datetime.fromtimestamp(timestamp).strftime("%d/%m/%Y %H:%M:%S")

        messagebox.showinfo(
            "Date du PR",
            f"Dernier enregistrement du PR :\n{date_modif}\n\n"
            "⚠️ Si tu viens de modifier le PR :\n"
            "➡️ fais CMD + S dans Excel\n"
            "➡️ puis clique sur Calcul / Reca"
        )

        try:
            self.importer_pr_dans_etat()
            messagebox.showinfo(
                "Calcul / Reca",
                "Prix de revient importé dans l'état d'avancement."
            )
        except Exception as e:
            messagebox.showerror("Erreur", str(e))

    def _dossier_chantier_selectionne(self) -> Path:
        sel = self.tree.selection()
        if not sel:
            raise ValueError("Sélectionne un chantier dans la liste.")

        item = self.tree.item(sel[0])
        vals = item.get("values", [])
        chantier_sel = str(vals[0]).strip() if vals else ""

        base = dossier_chantiers()
        return base / chantier_sel

    def importer_pr_dans_etat(self):
        dossier = self._dossier_chantier_selectionne()
        chemin_pr = dossier / "data" / "prix_de_revient.xlsx"

        try:
            chemin_etat = next(dossier.glob("Etat_avancement*.xlsm"))
        except StopIteration:
           return

        if not chemin_pr.exists():
            return

        wb_pr = load_workbook(chemin_pr, data_only=True)
        ws_pr = wb_pr["Chiffrage"]

        _backup_excel_before_write(chemin_etat)
        wb_etat = load_workbook(chemin_etat, keep_vba=True)
        ws_etat = wb_etat["Bordereau"]
        nb_importes = 0

        def clean_article(val):
            if val is None:
                return ""

            txt = str(val).strip().replace("\xa0", " ").replace(",", ".").lower()
            txt = re.sub(r"\s+", "", txt)
            if not txt:
                return ""

            parts = []
            for part in txt.split("."):
                if not part:
                    continue

                match = re.fullmatch(r"(\d+)([a-z]+)?", part)
                if match:
                    numero = str(int(match.group(1)))
                    suffixe = match.group(2) or ""
                    parts.append(numero + suffixe)
                    continue

                cleaned = re.sub(r"[^0-9a-z]+", "", part)
                if cleaned:
                    parts.append(cleaned)

            article = ".".join(parts)
            return article if any(c.isdigit() for c in article) else ""

        def normaliser_libelle(val):
            txt = str(val or "").strip().lower()
            txt = txt.replace("é", "e").replace("è", "e").replace("ê", "e").replace("ë", "e")
            txt = txt.replace("à", "a").replace("â", "a").replace("ä", "a")
            txt = txt.replace("ù", "u").replace("û", "u").replace("ü", "u")
            txt = txt.replace("î", "i").replace("ï", "i")
            txt = txt.replace("ô", "o").replace("ö", "o")
            txt = txt.replace("ç", "c")
            txt = re.sub(r"\s+", " ", txt)
            return txt

        def est_modele_marche_public():
            for ligne in range(1, min(ws_etat.max_row, 25) + 1):
                h = normaliser_libelle(ws_etat[f"H{ligne}"].value)
                i = normaliser_libelle(ws_etat[f"I{ligne}"].value)
                j = normaliser_libelle(ws_etat[f"J{ligne}"].value)
                if "en chiffres" in h and "en lettres" in i and "somme" in j:
                    return True
            return False

        marche_public = est_modele_marche_public()

        pr_map = {}
        for row in range(3, ws_pr.max_row + 1):
            article = clean_article(ws_pr[f"B{row}"].value)
            if not article:
                continue

            ligne_valeur = row + 27
            valeur = ws_pr[f"F{ligne_valeur}"].value

            if isinstance(valeur, (int, float)):
                pr_map[article] = float(valeur)

        for row in range(24, ws_etat.max_row + 1):
            article = clean_article(ws_etat[f"B{row}"].value)
            if not article:
                continue

            if article in pr_map:
                prix = pr_map[article]
                if marche_public:
                    ws_etat[f"H{row}"] = prix
                else:
                    ws_etat[f"L{row}"] = prix
                nb_importes += 1

        _protect_formula_cells(wb_etat, chemin_etat)
        wb_etat.save(chemin_etat)
        wb_pr.close()
        wb_etat.close()

        from tkinter import messagebox
        messagebox.showinfo("Import PR", f"{nb_importes} article(s) importé(s).")
  
    def ouvrir_chantier(self) -> None:
        p = self._chemin_chantier_selectionne()
        if not p:
            messagebox.showwarning("Chantier", "Sélectionne un chantier dans la liste.")
            return
        try:
            chantier = lire_json(p)
            if chantier.get("bordereau", {}).get("articles"):
                afficher_bordereau(self, chantier, "Bordereau (JSON)")
            else:
                messagebox.showinfo("Chantier", "Chantier ouvert. (Pas encore de bordereau dans ce chantier.)")
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible d'ouvrir le chantier.\n{e}")

    def dossier_du_chantier(self) -> None:
        p = self._chemin_chantier_selectionne()
        if not p:
            messagebox.showwarning("Chantier", "Sélectionne un chantier dans la liste.")
            return
        dossier = self._dossier_depuis_json(p)
        if not self._preparer_ouverture_document(dossier):
            return
        ouvrir_dossier(dossier)
        if self._doit_rappeler_pdf_historique(dossier):
            self._planifier_rappel_pdf_historique_ouverture()

    def nouveau_chantier(self) -> None:
        win = tk.Toplevel(self)
        win.title("Nouveau chantier")
        win.geometry("820x260")
        win.resizable(False, False)

        client_var = tk.StringVar(value="")
        chantier_var = tk.StringVar(value="")
        adresse_var = tk.StringVar(value="")

        frame = ttk.Frame(win, padding=14)
        frame.grid(row=0, column=0, sticky="nsew")
        win.columnconfigure(0, weight=1)
        win.rowconfigure(0, weight=1)

        ttk.Label(frame, text="Nom du client").grid(row=0, column=0, sticky="w")
        e_client = ttk.Entry(frame, textvariable=client_var, width=90)
        e_client.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(2, 8))

        ttk.Label(frame, text="Nom du chantier").grid(row=2, column=0, sticky="w")
        e_chantier = ttk.Entry(frame, textvariable=chantier_var, width=90)
        e_chantier.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(2, 8))

        ttk.Label(frame, text="Adresse").grid(row=4, column=0, sticky="w")
        e_adresse = ttk.Entry(frame, textvariable=adresse_var, width=90)
        e_adresse.grid(row=5, column=0, columnspan=2, sticky="ew", pady=(2, 12))

        def valider():
            nom_client = e_client.get().strip()
            nom_chantier = e_chantier.get().strip()
            adresse_chantier = e_adresse.get().strip()

            self.nom_client = nom_client
            self.nom_chantier = nom_chantier
            self.adresse_chantier = adresse_chantier

            if not nom_chantier:
                messagebox.showwarning("Nouveau chantier", "Renseigne le nom du chantier.")
                return

            nom_fichier = nom_chantier.replace("/", "_").replace("\\", "_")
            base = dossier_chantiers()
            json_path = base / f"{nom_fichier}.json"
            dossier_path = base / nom_fichier

            if json_path.exists() or dossier_path.exists():
                messagebox.showwarning("Nouveau chantier", "Ce chantier existe déjà.")
                return

            chantier = {
                "chantier": nom_chantier,
                "client": nom_client,
                "adresse": adresse_chantier,
                "etat": "Devis",
                "avancement": 0,
                "bordereau": {"source": {}, "articles": {}},
            }

            try:
                dossier_path.mkdir(parents=True, exist_ok=False)
                modeles_path = base / "Modèles"
                if not modeles_path.exists():
                    raise FileNotFoundError(f"Dossier modèles introuvable : {modeles_path}")
                shutil.copytree(modeles_path, dossier_path, dirs_exist_ok=True)
                ecrire_json(json_path, chantier)
                self.refresh_liste()
                if self.tree.exists(str(json_path)):
                    self.tree.selection_set(str(json_path))
                    self.tree.focus(str(json_path))
                win.destroy()
                return
            except Exception as e:
                messagebox.showerror("Nouveau chantier", f"Impossible de créer le chantier.\n{e}")
                return

            win.destroy()

        btns = ttk.Frame(frame)
        btns.grid(row=6, column=0, columnspan=2, sticky="e")
        ttk.Button(btns, text="Annuler", command=win.destroy).pack(side="right")
        ttk.Button(btns, text="Valider", command=valider).pack(side="right", padx=(0, 8))

        frame.columnconfigure(0, weight=1)
        e_client.focus_set()
        win.grab_set()



    def modifier_chantier(self) -> None:
        messagebox.showinfo("Info", "Modifier chantier (à brancher)")


    def supprimer_chantier(self) -> None:
        messagebox.showinfo("Info", "Supprimer chantier (à brancher)")


    def postes_metre(self) -> None:
        messagebox.showinfo("Info", "Postes / Métré (à faire)")

print("JE SUIS DANS LE BON FICHIER")

if __name__ == "__main__":
    app = HorizonChantierApp()
    app.mainloop()
